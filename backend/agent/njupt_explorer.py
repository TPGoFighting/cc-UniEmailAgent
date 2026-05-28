"""
南京邮电大学全学院教师邮箱爬虫 — 探索阶段
先获取所有学院的教师列表页URL，再批量爬取
"""
import asyncio
import re
import logging
from pathlib import Path

from playwright.async_api import async_playwright

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / "f8d29d14-aa64-4781-8efa-ee32cd310ec5"

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

# 南邮已知学院及其网站域名/子域名 (从主站和各来源收集)
NJUPT_COLLEGES = [
    # 学院名称 → 网站URL (从南邮主站可获得)
    ("通信与信息工程学院", "https://ctie.njupt.edu.cn"),
    ("电子与光学工程学院/柔性电子(未来技术)学院", "https://eoe.njupt.edu.cn"),
    ("计算机学院/软件学院/网络空间安全学院", "https://cs.njupt.edu.cn"),
    ("自动化学院/人工智能学院", "https://coa.njupt.edu.cn"),
    ("材料科学与工程学院", "https://iam.njupt.edu.cn"),
    ("物联网学院", "https://iot.njupt.edu.cn"),
    ("理学院", "https://cos.njupt.edu.cn"),
    ("地理与生物信息学院", "https://cgb.njupt.edu.cn"),
    ("现代邮政学院", "https://mpc.njupt.edu.cn"),
    ("传媒与艺术学院", "https://cma.njupt.edu.cn"),
    ("管理学院", "https://sm.njupt.edu.cn"),
    ("经济学院", "https://se.njupt.edu.cn"),
    ("马克思主义学院", "https://marx.njupt.edu.cn"),
    ("社会与人口学院", "https://ssp.njupt.edu.cn"),
    ("外国语学院", "https://sfs.njupt.edu.cn"),
    ("教育科学与技术学院", "https://est.njupt.edu.cn"),
    ("贝尔英才学院", "https://bel.njupt.edu.cn"),
    ("海外教育学院", "https://oice.njupt.edu.cn"),
    ("继续教育学院", "https://jjy.njupt.edu.cn"),
    ("体育部", "https://tyb.njupt.edu.cn"),
    ("通达学院", "https://tdxy.njupt.edu.cn"),
]

# 职称关键词
TITLE_KW = [
    "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
    "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
    "长江学者", "杰青", "优青", "博士后", "实验师",
    "青年专聘教授", "校聘副教授", "预聘副教授", "特聘教授",
    "客座教授", "名誉教授", "国家级教学名师", "高级实验师",
]

# 公共邮箱前缀
PUBLIC_EMAILS = {
    "jsjsj@njupt.edu.cn", "jsjyz@njupt.edu.cn", "jsjxy@njupt.edu.cn",
    "sxydw@njupt.edu.cn", "njugcglxy@njupt.edu.cn", "gcglxydw@njupt.edu.cn",
    "njudz@njupt.edu.cn", "yb_eng@njupt.edu.cn", "arch@nju.edu.cn",
}

# 导航文本过滤
NAV_TEXTS = {
    "首页", "学院概况", "学院简介", "机构设置", "新闻动态", "新闻中心",
    "通知公告", "学术动态", "科研动态", "人才培养", "招生信息", "招生宣传",
    "党建工作", "学生工作", "团学工作", "工会工作", "校友工作", "校友之家",
    "实验室", "联系我们", "院长信箱", "下载中心", "下载专区", "规章制度",
    "师德师风", "信息公开", "网站地图", "返回首页", "设为首页", "收藏本站",
    "师资队伍", "师资概况", "教师名录", "专任教师", "非专任教师", "导师介绍",
    "智慧校园", "诚聘英才", "领导信箱", "现任领导",
    "党建思政", "党委概况", "党建活动", "理论学习", "廉政监督",
    "本科生教育", "研究生教育", "专业介绍", "研究生培养",
    "创新竞赛", "创新班", "教学动态", "科学研究", "学科建设",
    "科研平台", "科研方向", "学术交流", "学生活动", "学子风采",
    "学工队伍", "研究生管理", "院徽院训", "本科教学", "研究生教学",
    "科研工作", "师资建设", "党员发展",
    "南邮主页", "南京邮电", "校内链接", "四电四邮", "学术组织",
    "个人信息", "返回列表",
    "更多", "详情", "登录", "注册", "注销", "忘记密码",
    "下一页", "上一页", "尾页", "末页",
    "视窗", "菜单", "导航", "搜索", "友情链接",
    "计算机学院", "软件学院", "网络空间安全学院",
    "通信学院", "自动化学院", "材料学院", "物联网学院",
    "理学院", "管理学院", "经济学院",
    "教师介绍", "个人简介", "教师详细信息",
}


def is_chinese_name(text: str) -> bool:
    text = text.strip()
    if not re.fullmatch(r"[一-鿿]{2,4}", text):
        return False
    if text in NAV_TEXTS:
        return False
    bad_ends = set("报组室部处委会局办系院所馆站网栏目页版")
    if text[-1] in bad_ends:
        return False
    return True


def is_public_email(email: str) -> bool:
    if not email:
        return True
    email_lower = email.lower()
    if email_lower in PUBLIC_EMAILS:
        return True
    local = email_lower.split("@")[0]
    for kw in ["webmaster", "admin", "info@", "office@", "master@", "president@",
                "xb@", "xxgk@", "jwc@", "yjsc@", "rsc@", "gjc@", "tw@", "xsc@",
                "wxyxz", "sxy", "dw"]:
        if kw in local:
            return True
    return False


def parse_at_sign(text: str) -> str:
    text = re.sub(r"\s*\[at\]\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*\(at\)\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*#@\s*", "@", text)
    text = re.sub(r"\s*\[@\]\s*", "@", text)
    text = re.sub(r"\s*\(@\)\s*", "@", text)
    text = re.sub(r"\s*\[dot\]\s*", ".", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*\(dot\)\s*", ".", text, flags=re.IGNORECASE)
    return text


def extract_titles(text: str) -> str:
    found = []
    for kw in TITLE_KW:
        if kw in text:
            found.append(kw)
    return "、".join(found[:5])


async def find_faculty_pages(context, college_name: str, site_url: str) -> list[dict]:
    """访问学院网站，找到教师列表页和所有教师详情页入口。"""
    logger.info(f"  🔍 探索 {college_name} ({site_url})")
    teacher_entries = []
    page = await context.new_page()

    try:
        # 访问学院首页
        await page.goto(site_url, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(2)

        # 收集页面上所有链接
        all_links = await page.evaluate("""() => {
            const links = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const text = (a.textContent || '').trim();
                const href = a.href;
                if (!text || !href || href.startsWith('javascript:') || href === '#') return;
                if (href.includes('mailto:')) return;
                links.push({text: text.substring(0, 100), url: href});
            });
            return links;
        }""")

        # 找"师资队伍""教师名录"等入口
        faculty_portal_links = []
        for link in all_links:
            text = link["text"]
            if any(kw in text for kw in ["师资", "教师名录", "教师", "导师", "教授"]):
                if len(text) < 15:  # 避免长文本
                    faculty_portal_links.append(link)

        logger.info(f"      找到 {len(faculty_portal_links)} 个师资相关入口")

        # 如果找到师资入口，访问它来找详情页
        visited_teacher_pages = set()
        for portal in faculty_portal_links[:3]:  # 限制探索数量
            try:
                portal_url = portal["url"]
                if portal_url in visited_teacher_pages:
                    continue
                visited_teacher_pages.add(portal_url)

                await page.goto(portal_url, wait_util="domcontentloaded", timeout=15000)
                await asyncio.sleep(2)

                # 从这个页面收集教师链接
                portal_links = await page.evaluate("""() => {
                    const links = [];
                    document.querySelectorAll('a[href]').forEach(a => {
                        const text = (a.textContent || '').trim();
                        const href = a.href;
                        if (!text || !href || href.startsWith('javascript:') || href === '#') return;
                        if (href.includes('mailto:')) return;
                        links.push({text: text.substring(0, 100), url: href});
                    });
                    return links;
                }""")

                # 筛选教师链接
                for link in portal_links:
                    t = link["text"]
                    u = link["url"]
                    if is_chinese_name(t):
                        teacher_entries.append({"name": t, "url": u, "college": college_name})
                    elif any(kw in u.lower() for kw in ["/page.htm", "/info/", "/detail/", "/view/",
                                                          "yjs.njupt.edu.cn", "/szdw/", "article",
                                                          "content", "/teacher/", "/faculty/"]):
                        # 可能是"个人信息"等链接文本，真正指向教师详情
                        if len(t) < 15 and t not in NAV_TEXTS:
                            teacher_entries.append({"name": t, "url": u, "college": college_name})

                logger.info(f"        从 {portal['text']} 收集到 {len(teacher_entries)} 个候选")

            except Exception as e:
                logger.debug(f"        {portal['text']} 访问失败: {str(e)[:60]}")

        # 如果师资入口没找到足够链接，直接从首页找所有可能的教师详情链接
        if len(teacher_entries) < 5:
            logger.info(f"      师资入口不足，从首页搜索补充...")
            for link in all_links:
                t = link["text"]
                u = link["url"]
                if is_chinese_name(t):
                    teacher_entries.append({"name": t, "url": u, "college": college_name})

    except Exception as e:
        logger.warning(f"    ⚠️ {college_name} 探索失败: {str(e)[:80]}")
    finally:
        await page.close()

    return teacher_entries


async def scrape_profile(context, entry: dict) -> dict | None:
    """访问教师详情页提取信息。"""
    page = await context.new_page()
    try:
        url = entry["url"]
        await page.goto(url, wait_until="domcontentloaded", timeout=15000)
        await asyncio.sleep(0.8)

        # 提取姓名
        name = entry["name"]
        if not is_chinese_name(name):
            # 尝试从页面提取
            title_text = await page.evaluate("() => document.title || ''")
            m = re.match(r"^([一-鿿]{2,4})\s*[-–—|]", title_text)
            if m and is_chinese_name(m.group(1)):
                name = m.group(1)
            else:
                # 尝试从 h1 提取
                headings = await page.evaluate("""() => {
                    const hs = [];
                    document.querySelectorAll('h1,h2,h3').forEach(h => {
                        hs.push(h.textContent.trim());
                    });
                    return hs;
                }""")
                for h in headings:
                    mm = re.match(r"^([一-鿿]{2,4})", h)
                    if mm and is_chinese_name(mm.group(1)):
                        name = mm.group(1)
                        break

        if not name or not is_chinese_name(name):
            return None

        # 提取邮箱
        html = await page.evaluate("() => document.body ? document.body.innerHTML : ''")
        html = parse_at_sign(html)
        body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
        body_text = parse_at_sign(body_text)

        # 尝试从"电子邮箱:"字段提取
        email = ""
        email_match = re.search(r"电子邮箱\s*[：:]\s*(\S+@\S+)", body_text)
        if email_match:
            email = email_match.group(1).strip()
        else:
            all_text = body_text + " " + html
            emails = EMAIL_RE.findall(all_text)
            email = next((e for e in emails if not is_public_email(e)), "")

        if is_public_email(email):
            email = ""

        # 提取职称
        title = extract_titles(body_text[:3000])

        if email:
            logger.info(f"  ✅ {name} → {email} | {title[:30]}")

        return {
            "姓名": name,
            "邮箱": email,
            "学院": entry["college"],
            "职称": title,
            "主页链接": url,
        }

    except Exception as e:
        logger.debug(f"  ⚠️ {entry['name']} 页面异常: {str(e)[:60]}")
        return None
    finally:
        await page.close()


async def main():
    logger.info("=" * 60)
    logger.info("🎓 南京邮电大学 全学院教师邮箱爬虫")
    logger.info("=" * 60)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"

        # === 阶段1：探索各学院教师页面 ===
        logger.info("\n📌 阶段1：探索各学院教师页面入口")
        all_entries = []
        seen_urls = set()

        for college_name, site_url in NJUPT_COLLEGES:
            try:
                ctx = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
                entries = await find_faculty_pages(ctx, college_name, site_url)
                for e in entries:
                    if e["url"] not in seen_urls:
                        seen_urls.add(e["url"])
                        all_entries.append(e)
                await ctx.close()
                logger.info(f"  📊 {college_name}: 累计 {len([e for e in all_entries if e['college'] == college_name])} 个条目")
            except Exception as e:
                logger.warning(f"  ❌ {college_name} 整体失败: {str(e)[:80]}")

        logger.info(f"\n📊 阶段1 汇总: 共 {len(all_entries)} 个教师详情页")

        # === 阶段2：批量爬取详情页 ===
        logger.info(f"\n📌 阶段2：批量爬取详情页 (共 {len(all_entries)} 个)")
        all_results = []
        sem = asyncio.Semaphore(4)

        async def process(entry):
            async with sem:
                ctx = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
                try:
                    return await scrape_profile(ctx, entry)
                finally:
                    await ctx.close()

        batch_size = 20
        for start in range(0, len(all_entries), batch_size):
            end = min(start + batch_size, len(all_entries))
            batch = all_entries[start:end]
            logger.info(f"\n  [{start+1}-{end}/{len(all_entries)}]")
            tasks = [process(e) for e in batch]
            results = await asyncio.gather(*tasks)
            count = 0
            for r in results:
                if r:
                    all_results.append(r)
                    count += 1
            logger.info(f"    本批: {count} 条, 累计: {len(all_results)}")

        await browser.close()

    # === 阶段3：数据清洗 ===
    logger.info(f"\n{'='*60}")
    logger.info("📊 数据清洗")

    # 去重
    seen_key = set()
    clean = []
    for r in all_results:
        key = (r["姓名"], r["邮箱"])
        if key not in seen_key:
            seen_key.add(key)
            clean.append(r)

    bad_names = {"个人信息", "视窗", "登录", "菜单", "导航", "更多", "详情",
                 "首页", "返回", "上一页", "下一页", "尾页", "教师介绍"}
    final = []
    for r in clean:
        if r["姓名"] in bad_names:
            continue
        if r["姓名"] in NAV_TEXTS:
            continue
        if not is_chinese_name(r["姓名"]):
            continue
        if not r["邮箱"] and not r["主页链接"]:
            continue
        final.append(r)

    logger.info(f"  原始={len(all_results)} → 去重={len(clean)} → 过滤={len(final)}")

    with_email = [r for r in final if r["邮箱"]]
    no_email = [r for r in final if not r["邮箱"]]
    logger.info(f"   有邮箱: {len(with_email)}, 无邮箱: {len(no_email)}")

    if with_email:
        logger.info(f"\n📋 有邮箱教师（前20条）:")
        for r in with_email[:20]:
            logger.info(f"   {r['姓名']} <{r['邮箱']}> {r['职称'][:25]} [{r['学院']}]")

    # === 导出 XLSX ===
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"南京邮电大学_教师邮箱_{timestamp}.xlsx"

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教师邮箱"

        HEADERS = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]

        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        cell_font = Font(name="微软雅黑", size=10)
        cell_align = Alignment(vertical="center")

        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, r in enumerate(final, 1):
            row_data = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 55

        wb.save(xlsx_path)
        logger.info(f"\n💾 XLSX 已保存: {xlsx_path}")

    except ImportError:
        logger.error("请安装 openpyxl: pip install openpyxl")
        return None

    # 统计各学院
    from collections import Counter
    college_counts = Counter(r["学院"] for r in final)
    logger.info(f"\n📊 各学院统计:")
    for college, count in college_counts.most_common():
        logger.info(f"   {college}: {count} 人")

    logger.info(f"\n✅ 完成！共 {len(final)} 条记录 → {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    from datetime import datetime
    asyncio.run(main())
