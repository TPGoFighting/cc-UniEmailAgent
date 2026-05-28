"""
清洗南京邮电大学爬虫结果：移除导航文本、公共邮箱、非教师条目
"""
import csv
import re
from datetime import datetime
from pathlib import Path
from collections import Counter

TASK_ID = "f8d29d14-aa64-4781-8efa-ee32cd310ec5"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / TASK_ID

# 不是教师姓名的全部短语（导航/页面元素/组织名等）
NOT_A_NAME = {
    # 通信学院
    "学院简介", "学院院徽", "历史沿革", "历任领导", "人才培养", "本科教育",
    "教学动态", "师资总览", "学科科研", "学科介绍", "重点平台", "科研动态",
    "党委概况", "党建制度", "党建动态", "统战工作", "工作动态", "纪检工作",
    "纪律法规", "纪检动态", "工会概况", "学工团队", "学生活动", "招生就业",
    "奖助学金", "合作交流", "学术讲座", "国际交流", "人才招聘", "教师招聘",
    "领导信箱", "学院地址", "书记信箱",
    "师资队伍", "师资概况", "教师名录", "导师介绍", "专任教师", "非专任教师",
    "研究生导师", "博士生导师", "硕士生导师",
    # 其他学院常见导航
    "学院概况", "机构设置", "现任领导", "行政科室", "系所中心",
    "本科生培养", "研究生培养", "留学生培养", "继续教育",
    "科学研究", "学术交流", "科研成果", "科研团队", "科研平台",
    "招生信息", "就业信息", "校友之家", "校友风采", "校友动态",
    "通知公告", "新闻动态", "学院新闻", "学术动态",
    "党建工作", "团学工作", "工会工作", "学生工作", "学生社团",
    "理论学习", "实践教学", "师德师风", "规章制度", "下载中心",
    "联系方式", "联系我们", "网站地图", "返回首页",
    "主题教育", "专题教育", "实践活动", "三会一课", "党员发展",
    "入党指南", "积极分子", "预备党员", "正式党员",
    "办事指南", "表格下载", "文件下载", "资料下载",
    "招标公告", "采购信息", "人才引进", "招聘信息",
    "会议通知", "活动通知", "学术活动", "学术报告",
    # 材料学院
    "学院领导", "行政科室", "系所设置",
    # 自动化学院
    "组织机构", "教师信息", "导师名录", "科研成果", "发表论文", "授权专利",
    "科研获奖", "成果展示", "培养计划", "实践教学", "党务活动", "工会概况",
    "工会活动", "学习资源", "学生社团", "学生荣誉", "风采展示",
    "本科生录", "研究生录", "校友动态", "行政工作", "教学管理",
    "第十三周", "拟发展", "科研管理", "博士后", "行政教辅",
    # 现代邮政
    "硕士生导师", "博士生导师",
    # 马克思主义
    "思想政治教育", "马克思主义原理", "中国近现代史",
    # 一般
    "教育捐赠", "监督方式", "举报方式",
}

# 高频公共邮箱（出现在多个"教师"条目中的都是公共邮箱）
PUBLIC_EMAILS_HIGH_FREQ = {
    "ici-yb@njupt.edu.cn",   # 通信学院公共
    "jksdsf@njupt.edu.cn",   # 教科院公共
}

# 邮箱关键词（明显不是个人邮箱的）
PUBLIC_KEYWORDS = {
    "ici-yb", "jksdsf", "xcb", "yuanzhang", "shuji", "bangongshi",
    "dangwei", "tuanwei", "xuesheng", "jiaowu", "keyan", "renshi",
    "caiwu", "houqin", "tushuguan", "xinxi", "bangong",
}

def is_real_person_name(text: str) -> bool:
    """判断是否为真实人名（非导航文本）"""
    text = text.strip()
    # 必须是2-4个中文字符
    if not re.fullmatch(r"[一-鿿]{2,4}", text):
        return False
    # 排除导航文本
    if text in NOT_A_NAME:
        return False
    # 排除组织/机构后缀
    if text[-1] in "报组室部处委会局办系院所馆站网栏目页版":
        return False
    if any(text.endswith(s) for s in ["学院", "大学", "中心", "研究所", "实验室"]):
        return False
    # 排除包含导航关键词的
    nav_patterns = ["学院", "教学", "科研", "学科", "招生", "就业",
                    "党建", "工会", "学生", "校友", "师资", "导师",
                    "领导", "行政", "管理", "办事", "通知", "新闻",
                    "人才", "招聘", "活动", "下载", "服务", "工作",
                    "举报", "教育", "捐", "监督"]
    for p in nav_patterns:
        if p in text:
            if len(text) <= 3:  # 2-3字导航词
                return False
    return True


def is_bad_email(email: str) -> bool:
    """判断是否为非个人邮箱"""
    if not email:
        return False
    email_lower = email.lower()
    # 高频公共邮箱
    if email_lower in PUBLIC_EMAILS_HIGH_FREQ:
        return True
    # 包含公共邮箱关键词
    for kw in PUBLIC_KEYWORDS:
        if kw in email_lower:
            return True
    # 排除明显非个人邮箱域名
    if "@" in email_lower:
        domain = email_lower.split("@")[1]
        if domain not in ("njupt.edu.cn", "njupt.edu", "nju.edu.cn",
                           "126.com", "163.com", "qq.com", "gmail.com",
                           "hotmail.com", "outlook.com", "foxmail.com",
                           "aliyun.com", "sina.com", "yeah.net", "139.com",
                           "189.cn", "wo.cn", "live.cn", "msn.com",
                           "yahoo.com", "yahoo.com.cn", "sohu.com",
                           "tom.com", "21cn.com", "vip.qq.com", "vip.163.com"):
            if ".edu" not in domain:
                return True
    return False


def main():
    # 读取最新的 xlsx
    import openpyxl

    xlsx_files = sorted(OUTPUT_DIR.glob("南京邮电大学_教师邮箱_*.xlsx"))
    if not xlsx_files:
        print("未找到xlsx文件")
        return

    latest_xlsx = xlsx_files[-1]
    print(f"读取: {latest_xlsx}")

    wb = openpyxl.load_workbook(latest_xlsx)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # 姓名不为空
            data.append({
                "序号": row[0],
                "姓名": str(row[1]).strip(),
                "邮箱": str(row[2]).strip() if row[2] else "",
                "学院": str(row[3]).strip() if row[3] else "",
                "职称": str(row[4]).strip() if row[4] else "",
                "主页链接": str(row[5]).strip() if row[5] else "",
            })

    print(f"原始: {len(data)} 条")

    # === 清洗 ===
    # 统计高频邮箱（>3个不同"姓名"使用同一邮箱 → 公共邮箱）
    email_name_counts = Counter()
    for r in data:
        if r["邮箱"]:
            email_name_counts[r["邮箱"]] += 1

    # 同一邮箱出现>=4次 → 标记为公共邮箱
    extra_public = {e for e, c in email_name_counts.items() if c >= 4}
    print(f"  高频邮箱(≥4次): {len(extra_public)}")

    # 过滤
    clean = []
    removed_no_name = 0
    removed_public_email = 0
    removed_foreign = 0

    for r in data:
        name = r["姓名"]
        email = r["邮箱"]

        # 1. 必须是真人名
        if not is_real_person_name(name):
            removed_no_name += 1
            continue

        # 2. 邮箱不能是公共邮箱
        if email and (is_bad_email(email) or email.lower() in extra_public):
            # 保留记录但清空邮箱
            r["邮箱"] = ""
            clean.append(r)
            removed_public_email += 1
            continue

        # 3. 过滤包含英文/数字的"姓名"
        if re.search(r"[a-zA-Z0-9]", name):
            removed_foreign += 1
            continue

        clean.append(r)

    print(f"  非人名: {removed_no_name}")
    print(f"  公共邮箱: {removed_public_email}")
    print(f"  含英文: {removed_foreign}")
    print(f"  清洗后: {len(clean)} 条")

    with_email = [r for r in clean if r["邮箱"]]
    no_email = [r for r in clean if not r["邮箱"]]
    print(f"  有邮箱: {len(with_email)}, 无邮箱: {len(no_email)}")

    # 学院统计
    cc = Counter(r["学院"] for r in clean)
    print(f"\n📊 各学院统计:")
    for c, cnt in cc.most_common():
        e = len([r for r in clean if r["学院"] == c and r["邮箱"]])
        print(f"   {c}: {cnt} (有邮箱: {e})")

    # 输出清洗版
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # XLSX
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = "教师邮箱"
    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]

    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    cf = Font(name="微软雅黑", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = tb

    for i, r in enumerate(clean, 1):
        row = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
        for col, val in enumerate(row, 1):
            cell = ws_new.cell(row=i+1, column=col, value=val)
            cell.font = cf; cell.alignment = Alignment(vertical="center"); cell.border = tb

    ws_new.column_dimensions["A"].width = 8
    ws_new.column_dimensions["B"].width = 18
    ws_new.column_dimensions["C"].width = 35
    ws_new.column_dimensions["D"].width = 35
    ws_new.column_dimensions["E"].width = 20
    ws_new.column_dimensions["F"].width = 60

    xlsx_clean = OUTPUT_DIR / f"南京邮电大学_教师邮箱_清洗版_{ts}.xlsx"
    wb_new.save(xlsx_clean)
    print(f"\n💾 清洗版XLSX: {xlsx_clean}")

    # 有邮箱名单
    email_only = [r for r in clean if r["邮箱"]]
    xlsx_email = OUTPUT_DIR / f"南京邮电大学_有邮箱教师_{ts}.xlsx"
    wb_e = openpyxl.Workbook()
    ws_e = wb_e.active
    ws_e.title = "有邮箱教师"
    for col, h in enumerate(headers, 1):
        cell = ws_e.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = tb
    for i, r in enumerate(email_only, 1):
        row = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
        for col, val in enumerate(row, 1):
            cell = ws_e.cell(row=i+1, column=col, value=val)
            cell.font = cf; cell.alignment = Alignment(vertical="center"); cell.border = tb
    ws_e.column_dimensions["A"].width = 8
    ws_e.column_dimensions["B"].width = 18
    ws_e.column_dimensions["C"].width = 35
    ws_e.column_dimensions["D"].width = 35
    ws_e.column_dimensions["E"].width = 20
    ws_e.column_dimensions["F"].width = 60
    wb_e.save(xlsx_email)
    print(f"💾 仅有邮箱: {xlsx_email}")

    print(f"\n✅ 完成! {len(clean)} 条, {len(with_email)} 邮箱, {len(cc)} 学院")

    print(f"\n[FILES]")
    print(f"{xlsx_clean.name} | 南京邮电大学教师邮箱完整数据-清洗版 (共{len(clean)}条, {len(with_email)}个邮箱, {len(cc)}个学院)")
    print(f"{xlsx_email.name} | 南京邮电大学教师邮箱-仅有邮箱 (共{len(email_only)}条)")
    print(f"[/FILES]")


if __name__ == "__main__":
    main()
