"""Claude Code Agent Runtime — 通过子进程调用 claude CLI，智能驱动浏览器任务。"""

import asyncio
import json
import logging
import os
import re
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import AsyncGenerator

from agent.tracing import create_run, end_run, start_span, end_span
from agent.guardrails import check_output, sanitize_output

logger = logging.getLogger(__name__)
# SUBDEBUG 等详细日志只在 DEBUG 级别输出，避免控制台刷屏
logger.setLevel(logging.WARNING)

MAX_STEPS = 10000
TIMEOUT_SECONDS = 3600
CLAUDE_STARTUP_TIMEOUT = 30

# 允许 claude CLI 使用的工具白名单（替代 bypassPermissions）
ALLOWED_TOOLS = ["Read", "Edit", "Write", "Bash", "Glob", "Grep", "WebSearch", "WebFetch"]


class ClaudeAgent:
    """Claude Code Agent 包装器。

    通过 asyncio 子进程调用 `claude -p --output-format stream-json`，
    解析流式 JSON 输出，实时推送日志。

    如果 claude CLI 不可用，自动降级到 PlaywrightAgent。
    """

    def __init__(self):
        self._last_stderr = ""
        self.active_procs = {}
        self._check_claude()

    def _check_claude(self) -> bool:
        import shutil

        path = shutil.which("claude")
        if not path:
            logger.warning("claude CLI 未找到，将使用回退模式")
            return False
        logger.info(f"claude CLI 已就绪: {path}")
        return True

    def stop_task(self, task_id: str) -> bool:
        """终止正在运行的任务子进程。"""
        proc = self.active_procs.pop(task_id, None)
        if proc:
            try:
                proc.kill()
                logger.info(f"成功终止 Task {task_id} 的子进程")
                return True
            except Exception as e:
                logger.error(f"终止 Task {task_id} 的子进程失败: {e}")
        return False

    def _timestamp(self) -> str:
        return datetime.now().strftime("%H:%M:%S")

    async def execute(
        self,
        message: str,
        task_id: str = "",
        is_continuation: bool = False,
        is_crawl_session: bool = True,
        current_user_message: str | None = None,
        intent_result=None,
        **kwargs
    ) -> AsyncGenerator[dict, None]:
        """执行任务。先尝试 Claude Code，失败则回退到 Playwright。

        task_id 用于任务隔离：输出文件写入 outputs/{task_id}/ 子目录。
        is_continuation 为 True 时跳过爬取策略注入（上下文已由 main.py 构建）。"""
        # 如果不是爬取任务，直接进行智能分流（免拉起 CLI 进程以达到毫秒级响应）
        if not is_crawl_session:
            eval_msg = current_user_message or message
            has_key = os.environ.get("DEEPSEEK_API_KEY") or os.environ.get("OPENAI_API_KEY") or os.environ.get("ANTHROPIC_API_KEY")
            if has_key:
                async for log in self._respond_conversational(eval_msg):
                    yield log
                return
            else:
                # 无 Key 时，先发一条友好的秒回优化提示，然后自动降级使用 Claude CLI 获取真实回答
                yield {
                    "type": "text",
                    "message": "💡 **提示**：检测到当前尚未配置 `DEEPSEEK_API_KEY`。系统已自动为你降级为 Claude CLI 流程（冷启动及索引需要约 10 秒）。\n\n*   **秒回优化建议**：在 `backend/.env` 中配置 `DEEPSEEK_API_KEY` 环境变量即可彻底激活 DeepSeek 毫秒级流式秒回通道！\n\n---\n\n正在为你生成回答...\n",
                    "timestamp": self._timestamp()
                }

        # LangSmith 三阶段追踪：startup → execute → finish
        run_id = create_run("claude_execute", {"task_id": task_id, "message": message[:200]})
        span_startup = start_span("claude_startup", {"task_id": task_id, "phase": "init"})
        end_span(span_startup)
        _claude_error = None
        try:
            span_process = start_span("claude_process", {"task_id": task_id, "phase": "execute"})
            try:
                async for log in self._run_claude(message, task_id, is_continuation):
                    yield log
            except Exception:
                end_span(span_process, error="执行异常")
                raise
            else:
                end_span(span_process)
            span_finish = start_span("claude_finish", {"task_id": task_id, "phase": "cleanup"})
            end_span(span_finish)
        except Exception as e:
            _claude_error = f"{type(e).__name__}: {str(e)[:200]}"
            import traceback as _tb
            tb_lines = _tb.format_exc().replace("\n", " | ")
            logger.warning(f"V2 Claude Code 执行失败: {_claude_error} || TRACEBACK: {tb_lines}")

            # 仅爬取任务才回退到 Playwright，追问/普通问答不回退
            if is_crawl_session and not is_continuation:
                yield {
                    "type": "log",
                    "message": f"V2 Claude Code 不可用（{_claude_error}），切换到内置浏览器 Agent",
                    "timestamp": self._timestamp(),
                }
                from agent.playwright_agent import PlaywrightAgent
                fallback = PlaywrightAgent()
                try:
                    async for log in fallback.execute(message, task_id):
                        yield log
                except Exception as pe:
                    logger.error(f"V2 Playwright 回退也失败: {type(pe).__name__}: {str(pe)[:200]}")
                    yield {
                        "type": "error",
                        "message": f"V2 Playwright 浏览器 Agent 也执行失败: {type(pe).__name__}: {str(pe)[:200]}",
                        "timestamp": self._timestamp(),
                    }
            else:
                yield {
                    "type": "error",
                    "message": f"V2 Claude Code 执行失败（{_claude_error}）。这不是爬取任务，请检查 claude CLI 是否正常工作。",
                    "timestamp": self._timestamp(),
                }
        finally:
            end_run(run_id, error=_claude_error)

    async def execute_query(
        self,
        message: str,
        task_id: str,
        task_output_dir: str = "",
    ) -> AsyncGenerator[dict, None]:
        """处理简单问答意图：读取当前任务输出文件，用 LLM 做统计分析后回答。

        绝不触发爬虫代码，只读已有文件进行分析。无 LLM 时回退本地统计。
        """
        from agent.exporter import _BASE_OUTPUT_DIR
        import csv as _csv

        run_id = create_run("claude_query", {"task_id": task_id, "message": message[:200]})
        span_analysis = start_span("claude_analysis", {"task_id": task_id, "query": message[:100]})

        output_dir = Path(task_output_dir) if task_output_dir else _BASE_OUTPUT_DIR / task_id

        # ── 扫描当前任务目录下的数据文件 ──
        data_files: list[Path] = []
        if output_dir.exists():
            for f in output_dir.iterdir():
                if f.suffix.lower() in (".csv", ".xlsx") and f.stat().st_size > 100:
                    data_files.append(f)

        if not data_files:
            # 常识问答（网址、官网等）→ 直接回答
            knowledge_qa = [
                (r"(南京邮电大学|南邮).*网址|(南京邮电大学|南邮).*官网|nuist.*url|nuist.*website", "南京邮电大学官网是 https://www.njupt.edu.cn"),
                (r"(南京大学|南大).*网址|(南京大学|南大).*官网", "南京大学官网是 https://www.nju.edu.cn"),
                (r"(东南大学|东大).*网址|(东南大学|东大).*官网", "东南大学官网是 https://www.seu.edu.cn"),
                (r"(南京理工大学|南理工).*网址|(南京理工大学|南理工).*官网", "南京理工大学官网是 https://www.njust.edu.cn"),
                (r"(南京航空航天大学|南航).*网址|(南京航空航天大学|南航).*官网", "南京航空航天大学官网是 https://www.nuaa.edu.cn"),
                (r"(清华大学|北大).*网址|(清华大学|北大).*官网", "清华大学官网是 https://www.tsinghua.edu.cn"),
                (r"(北京大学|北大).*网址|(北京大学|北大).*官网", "北京大学官网是 https://www.pku.edu.cn"),
            ]
            answer = None
            for pattern, response in knowledge_qa:
                if re.search(pattern, message, re.IGNORECASE):
                    answer = response
                    break
            if answer:
                yield {"type": "text", "message": answer, "timestamp": self._timestamp()}
                end_span(span_analysis)
                end_run(run_id)
                yield {"type": "done", "message": "已回答", "timestamp": self._timestamp()}
                return

            # 打招呼/闲聊 → 友好回应
            greeting_patterns = [
                r"^(你好|您好|hi|hello|hey|嗨|早上好|下午好|晚上好|在吗|在不在)[!！.。]*$",
                r"^(你好|您好|hi|hello)\s*[,，]?\s*(啊|呀|吗|吧)?[!！？?]*$",
            ]
            is_greeting = any(re.match(p, message.strip(), re.IGNORECASE) for p in greeting_patterns)
            if is_greeting:
                yield {
                    "type": "text",
                    "message": "你好！我是 UniEmail Agent，可以帮你爬取高校教师的邮箱信息。\n\n你可以这样跟我说话：\n- 「帮我抓取南京大学计算机学院教师邮箱」\n- 「补充清华大学计算机系缺失的邮箱」\n- 「导出北京大学已抓取的数据为 CSV」",
                    "timestamp": self._timestamp(),
                }
            else:
                yield {
                    "type": "text",
                    "message": "当前任务还没有生成数据文件，请先执行爬取任务后再查询。",
                    "timestamp": self._timestamp(),
                }
            yield {"type": "done", "message": "无数据可分析", "timestamp": self._timestamp()}
            end_span(span_analysis)
            end_run(run_id)
            return

        # ── 读取文件内容做分析 ──
        file_stats: list[dict] = []
        all_rows: list[dict] = []
        for fp in sorted(data_files):
            rows = []
            if fp.suffix.lower() == ".csv":
                try:
                    with open(fp, "r", encoding="utf-8-sig", errors="replace") as f:
                        reader = _csv.DictReader(f)
                        for row in reader:
                            rows.append(row)
                except Exception:
                    pass
            elif fp.suffix.lower() == ".xlsx":
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(fp, read_only=True)
                    ws = wb.active
                    headers = [str(c.value or "") for c in next(ws.iter_rows())]
                    for row in ws.iter_rows(values_only=True):
                        rows.append(dict(zip(headers, [str(v or "") for v in row])))
                    wb.close()
                except Exception:
                    pass

            # 统计
            total = len(rows)
            has_email = sum(1 for r in rows if "@" in (r.get("邮箱", "") or r.get("email", "")))
            dept_col = "学院" if "学院" in (rows[0] if rows else {}) else "department"
            dept_counts: dict[str, int] = {}
            for r in rows:
                d = r.get(dept_col, "未知").strip() or "未知"
                dept_counts[d] = dept_counts.get(d, 0) + 1

            file_stats.append({
                "file": fp.name,
                "total": total,
                "with_email": has_email,
                "email_rate": f"{has_email / total * 100:.1f}%" if total > 0 else "0%",
                "departments": dept_counts,
            })
            all_rows.extend(rows)

        # ── 尝试用 LLM 生成分析回答 ──
        api_key = os.environ.get("DEEPSEEK_API_KEY") or os.environ.get("OPENAI_API_KEY") or os.environ.get("ANTHROPIC_API_KEY")
        if api_key:
            answer = await self._query_llm_analysis(message, file_stats, api_key)
            if answer:
                yield {"type": "text", "message": answer, "timestamp": self._timestamp()}
                end_span(span_analysis)
                end_run(run_id)
                yield {"type": "done", "message": "数据分析完毕", "timestamp": self._timestamp()}
                return

        # ── 回退：本地统计分析 ──
        answer = self._local_analysis(message, file_stats)
        yield {"type": "text", "message": answer, "timestamp": self._timestamp()}
        yield {"type": "done", "message": "数据分析完毕", "timestamp": self._timestamp()}
        end_span(span_analysis)
        end_run(run_id)

    async def _query_llm_analysis(
        self, question: str, file_stats: list[dict], api_key: str
    ) -> str | None:
        """调用 LLM 分析文件数据并回答用户问题。"""
        try:
            import json as _json
            from openai import AsyncOpenAI

            if os.environ.get("DEEPSEEK_API_KEY"):
                client = AsyncOpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")
                model = "deepseek-chat"
            elif os.environ.get("OPENAI_API_KEY"):
                client = AsyncOpenAI(
                    api_key=api_key,
                    base_url=os.environ.get("OPENAI_BASE_URL") or None,
                )
                model = os.environ.get("OPENAI_API_MODEL") or "gpt-4o-mini"
            else:
                # Anthropic - use httpx directly to avoid dependency
                import httpx
                stats_summary = _json.dumps(file_stats, ensure_ascii=False, indent=2)
                headers = {
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                }
                data = {
                    "model": os.environ.get("ANTHROPIC_API_MODEL") or "claude-sonnet-4-6",
                    "max_tokens": 800,
                    "messages": [
                        {"role": "user", "content": f"根据以下文件统计数据回答用户问题。\n\n## 统计数据\n{stats_summary}\n\n## 用户问题\n{question}\n\n请用中文简洁回答，使用 Markdown 格式。直接给出分析结论，不要开场白。"},
                    ],
                }
                base_url = os.environ.get("ANTHROPIC_BASE_URL") or "https://api.anthropic.com"
                async with httpx.AsyncClient(timeout=30) as c:
                    r = await c.post(f"{base_url.rstrip('/')}/v1/messages", headers=headers, json=data)
                    if r.status_code == 200:
                        return r.json()["content"][0]["text"]
                return None

            stats_summary = _json.dumps(file_stats, ensure_ascii=False, indent=2)
            response = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "你是一个数据分析助手。根据文件统计数据回答用户问题。用中文简洁回答，使用 Markdown。直接给出分析结论。"},
                    {"role": "user", "content": f"## 当前任务文件统计数据\n{stats_summary}\n\n## 用户问题\n{question}"},
                ],
                max_tokens=800,
                temperature=0.3,
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            logger.warning(f"LLM 分析失败，回退本地统计: {e}")
            return None

    def _local_analysis(self, question: str, file_stats: list[dict]) -> str:
        """本地统计分析（无 LLM 时的兜底）。"""
        lines = ["## 📊 当前任务数据统计\n"]
        for fs in file_stats:
            lines.append(f"### {fs['file']}")
            lines.append(f"- 总记录数: **{fs['total']}**")
            lines.append(f"- 有邮箱: **{fs['with_email']}**（{fs['email_rate']}）")
            lines.append(f"- 无邮箱: **{fs['total'] - fs['with_email']}**")
            lines.append("")
            lines.append(f"#### 各学院分布")
            lines.append("| 学院 | 人数 |")
            lines.append("|------|------|")
            for dept, count in sorted(fs["departments"].items(), key=lambda x: -x[1]):
                lines.append(f"| {dept} | {count} |")
            lines.append("")

        lines.append("> 💡 配置 `DEEPSEEK_API_KEY` 可启用 AI 智能分析回答。")
        return "\n".join(lines)

    async def _is_crawl_task(self, message: str) -> bool:
        """使用 DeepSeek 大模型对用户意图进行精准分类，判定是否为爬取/数据采集类任务。
        
        若无 API 密钥，回退到本地关键词硬匹配规则，保障系统鲁棒性。
        """
        import os
        import json
        
        # 1. 尝试调用 DeepSeek API 进行意图识别
        api_key = os.environ.get("DEEPSEEK_API_KEY") or os.environ.get("OPENAI_API_KEY")
        if api_key:
            try:
                import openai
                from openai import AsyncOpenAI
                
                # 配置 Client，支持 DeepSeek 官方和 OpenAI 兼容代理
                if os.environ.get("DEEPSEEK_API_KEY"):
                    base_url = "https://api.deepseek.com/v1"
                    model = "deepseek-chat"
                else:
                    base_url = os.environ.get("OPENAI_BASE_URL") or None
                    model = os.environ.get("OPENAI_API_MODEL") or ("deepseek-chat" if base_url and "deepseek" in base_url else "gpt-4o-mini")
                
                client = AsyncOpenAI(api_key=api_key, base_url=base_url)
                
                system_prompt = (
                    "你是一个精准的任务意图分类器。请判断用户输入的请求是否是关于高校教师邮箱抓取、"
                    "师资队伍信息采集、网页邮箱抓取、爬虫提取等数据采集或批量清洗任务。\n"
                    "请仅回复 CRAWL（若是爬虫/抓取/采集/提取任务）或 CHAT（若是日常闲聊、通用问答、代码解释或系统配置类普通会话）。\n"
                    "绝对不能回复除 CRAWL 和 CHAT 以外的任何其他内容或标点符号！"
                )
                
                response = await client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": message}
                    ],
                    max_tokens=5,
                    temperature=0.0
                )
                
                result = response.choices[0].message.content.strip().upper()
                logger.info(f"[Intent Classifier] Model {model} classified message intent: {result}")
                if "CRAWL" in result:
                    return True
                if "CHAT" in result:
                    return False
            except Exception as e:
                logger.warning(f"[Intent Classifier] DeepSeek 意图识别异常（将回退本地规则）: {e}")
                
        # 2. 兜底策略：本地关键词硬匹配规则
        keywords = [
            "抓取", "爬取", "爬虫", "邮箱", "教师",
            "crawl", "scrape", "email", "faculty", "teacher", "学院",
        ]
        lower = message.lower()
        hits = sum(1 for kw in keywords if kw in lower)
        # 强信号多词短语直接判定
        combos = ["教师邮箱", "crawl email", "scrape email"]
        if any(c in lower for c in combos):
            return True
        return hits >= 2

    async def _drain_stderr(self, stream: asyncio.StreamReader) -> None:
        """持续消费 stderr，防止管道缓冲区满导致子进程死锁。"""
        try:
            while True:
                chunk = await stream.read(65536)
                if not chunk:
                    break
                self._last_stderr = chunk.decode("utf-8", errors="replace")[-500:]
        except Exception:
            pass

    def _start_subprocess_thread(
        self, cmd: list[str], env: dict, queue: asyncio.Queue, task_id: str, prompt: str = ""
    ) -> None:
        """Windows 兼容：在后台线程启动 subprocess.Popen，stdout 行写入 asyncio.Queue。

        通过 stdin 管道传递 prompt（避免 Windows 命令行长度限制）。
        调用方通过 queue 获取已解码的 stdout 行，每行为 {"_type": "line", "data": str}。"""
        stop_event = threading.Event()

        def _run():
            try:
                si = None
                if sys.platform == "win32":
                    si = subprocess.STARTUPINFO()
                    si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                    si.wShowWindow = subprocess.SW_HIDE

                proc = subprocess.Popen(
                    cmd,
                    stdin=subprocess.PIPE,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    env=env,
                    startupinfo=si,
                )
                self.active_procs[task_id] = proc

                # 通过 stdin 发送 prompt（纯文本，避免 Windows 命令行长度限制）
                if prompt:
                    try:
                        proc.stdin.write(prompt.encode("utf-8"))
                        proc.stdin.flush()
                        proc.stdin.close()
                    except (BrokenPipeError, OSError) as e:
                        logger.warning(f"stdin 写入失败: {e}")

                def _read_stderr():
                    try:
                        while not stop_event.is_set():
                            chunk = proc.stderr.read(65536)
                            if not chunk:
                                break
                            self._last_stderr = chunk.decode("utf-8", errors="replace")[-500:]
                    except Exception:
                        pass

                stderr_thread = threading.Thread(target=_read_stderr, daemon=True)
                stderr_thread.start()

                for line in proc.stdout:
                    if stop_event.is_set():
                        break
                    decoded = line.decode("utf-8", errors="replace").strip()
                    if decoded:
                        try:
                            queue.put_nowait({"_type": "line", "data": decoded})
                        except asyncio.QueueFull:
                            pass  # 队列满了丢弃（极端情况）

                proc.wait()
                stderr_thread.join(timeout=2)

                if proc.returncode != 0:
                    err = getattr(self, "_last_stderr", "")
                    logger.warning(f"claude 进程退出码 {proc.returncode}: {err[-300:]}")
                    # 将错误信息注入队列，让 async generator 产出 error 消息
                    try:
                        queue.put_nowait({
                            "_type": "error",
                            "message": f"Claude Code 进程异常退出（退出码 {proc.returncode}）：{err[-300:]}" if err else f"Claude Code 进程异常退出（退出码 {proc.returncode}）"
                        })
                    except asyncio.QueueFull:
                        pass
            except Exception as e:
                logger.error(f"子进程线程异常: {e}")
            finally:
                self.active_procs.pop(task_id, None)
                stop_event.set()
                try:
                    queue.put_nowait({"_type": "done"})
                except asyncio.QueueFull:
                    pass

        thread = threading.Thread(target=_run, daemon=True)
        thread.start()

    async def _process_stream(
        self,
        process,
        task_id: str,
        task_start: float,
        message: str,
    ) -> AsyncGenerator[dict, None]:
        """处理子进程的 stdout 流，提取 Claude Code 输出并推送下载文件。

        process 可以是 asyncio.subprocess.Process 或 asyncio.Queue（线程方案）。"""
        import asyncio

        is_threaded = isinstance(process, asyncio.Queue)

        if not is_threaded:
            try:
                process.stdout._limit = 10 * 1024 * 1024  # 10MB
            except AttributeError:
                pass
            stderr_task = asyncio.create_task(self._drain_stderr(process.stderr))

        step_count = 0
        has_output = False
        tracked_files: list[Path] = []
        collected_text: list[str] = []

        should_stop = False

        async def _handle_line(line: str):
            nonlocal step_count, has_output, should_stop
            step_count += 1
            if step_count > MAX_STEPS:
                if not is_threaded:
                    process.kill()
                should_stop = True
                yield {
                    "type": "log",
                    "message": f"达到最大步数限制 ({MAX_STEPS})，任务已终止",
                    "timestamp": self._timestamp(),
                }
                return

            parsed = self._parse_stream_line(line)
            if parsed:
                for tu in parsed.get("_tool_uses", []):
                    if tu.get("name") == "Write":
                        fp = tu.get("input", {}).get("file_path", "")
                        if fp:
                            tracked_files.append(Path(fp))

                msg_type = parsed.get("_msg_type", "")
                if msg_type == "assistant":
                    # 文本 chunks → Agent 思考日志（type="log"），前端默认折叠在日志面板
                    for chunk in parsed.get("_chunks", []):
                        collected_text.append(chunk)
                        has_output = True
                        chunk = sanitize_output(chunk)
                        yield {
                            "type": "log",
                            "message": chunk,
                            "timestamp": self._timestamp(),
                        }
                    # 工具调用 chunks → 技术日志，作为 type="log"（前端默认隐藏）
                    for chunk in parsed.get("_tool_chunks", []):
                        yield {
                            "type": "log",
                            "message": chunk,
                            "timestamp": self._timestamp(),
                        }
                elif msg_type == "result":
                    has_output = True
                    if parsed.get("is_error"):
                        subtype = parsed.get("_subtype", "unknown")
                        err_text = parsed.get("_text", "").strip()
                        err_text = sanitize_output(err_text)
                        if not err_text:
                            err_text = self._last_stderr[-500:] if hasattr(self, "_last_stderr") and self._last_stderr else f"Claude Code 执行出错（subtype={subtype}，无详细信息）"
                        else:
                            err_text = f"{err_text}（subtype={subtype}）"
                        yield {
                            "type": "error",
                            "message": err_text,
                            "timestamp": self._timestamp(),
                        }
                    else:
                        text = parsed.get("_text", "")
                        text = sanitize_output(text)
                        # result 文本是 Claude Code 的最终回答 → 作为 type="text" 显示给用户
                        if not text:
                            # 无显式 result 文本时，生成一个默认完成提示
                            # 避免重用 collected_text[-1]（它已是 type="log"，重用会导致内容重复）
                            text = "## 任务完成\n\nAgent 已执行完所有步骤。"
                        if text:
                            yield {
                                "type": "text",
                                "message": text,
                                "timestamp": self._timestamp(),
                            }
                elif msg_type == "system":
                    pass

        try:
            if is_threaded:
                while True:
                    try:
                        msg = await asyncio.wait_for(process.get(), timeout=TIMEOUT_SECONDS)
                    except asyncio.TimeoutError:
                        yield {
                            "type": "error",
                            "message": f"任务超时 ({TIMEOUT_SECONDS}s)，已终止",
                            "timestamp": self._timestamp(),
                        }
                        return
                    logger.info("[SUBDEBUG] 收到队列消息 _type=%s", msg.get("_type", "?"))
                    if msg["_type"] == "line":
                        line_preview = msg["data"][:100] if isinstance(msg.get("data"), str) else "?"
                        logger.info("[SUBDEBUG] 处理行: %s", line_preview)
                        async for log in _handle_line(msg["data"]):
                            yield log
                        if should_stop:
                            logger.info("[SUBDEBUG] 达到最大步数，停止处理")
                            break
                    elif msg["_type"] == "error":
                        # 线程路径传递的进程异常退出错误
                        yield {
                            "type": "error",
                            "message": msg.get("message", "Claude Code 进程异常退出"),
                            "timestamp": self._timestamp(),
                        }
                    elif msg["_type"] == "done":
                        logger.info("[SUBDEBUG] 收到线程完成信号")
                        break
            else:
                async for raw_line in process.stdout:
                    line = raw_line.decode("utf-8", errors="replace").strip()
                    if not line:
                        continue
                    async for log in _handle_line(line):
                        yield log
                    if should_stop:
                        break

                await asyncio.wait_for(process.wait(), timeout=10)

        except asyncio.TimeoutError:
            if not is_threaded:
                process.kill()
            yield {
                "type": "error",
                "message": f"任务超时 ({TIMEOUT_SECONDS}s)，已终止",
                "timestamp": self._timestamp(),
            }
            return
        finally:
            self.active_procs.pop(task_id, None)
            if not is_threaded:
                stderr_task.cancel()
                try:
                    await stderr_task
                except (asyncio.CancelledError, Exception):
                    pass

        if not has_output:
            err_text = getattr(self, "_last_stderr", "无 stderr 输出")
            raise RuntimeError(f"Claude Code 无输出: {err_text}")

        # ── 推送下载文件 ──
        from agent.exporter import _BASE_OUTPUT_DIR

        all_text = "\n".join(collected_text)
        declared = self._parse_file_declarations(all_text)

        pushed: set[str] = set()

        if declared:
            for filename, label in declared:
                if filename in pushed:
                    continue
                safe_name = filename.replace("\\", "/").split("/")[-1]
                if not safe_name or safe_name == ".":
                    continue
                if task_id:
                    candidate = _BASE_OUTPUT_DIR / task_id / safe_name
                else:
                    candidate = _BASE_OUTPUT_DIR / safe_name
                if not candidate.exists():
                    logger.warning(f"Agent 声明了文件但不存在: {safe_name}")
                    continue
                pushed.add(safe_name)
                yield {
                    "type": "download",
                    "message": label,
                    "filename": safe_name,
                    "url": f"/api/download/{task_id}/{safe_name}" if task_id else f"/api/download/{safe_name}",
                    "timestamp": self._timestamp(),
                }
                logger.info(f"[FILES] 声明推送: {safe_name} ({label})")

        if not pushed:
            logger.info("Agent 未使用 [FILES] 声明，回退到 Write 工具追踪")
            for fp in tracked_files:
                if fp.suffix.lower() != ".csv":
                    continue
                if fp.name.startswith(".") or fp.suffix == ".tmp":
                    continue
                try:
                    base = _BASE_OUTPUT_DIR.resolve()
                    resolved = fp.resolve()
                    if not str(resolved).startswith(str(base) + os.sep) and resolved != base:
                        continue
                    if not resolved.exists():
                        continue
                    filename = resolved.name
                    if filename in pushed:
                        continue
                    pushed.add(filename)
                    fmt_label = {"csv": "CSV 表格", "xlsx": "Excel 表格"}
                    label = fmt_label.get(fp.suffix.lower(), fp.suffix.lower())
                    yield {
                        "type": "download",
                        "message": f"{label}（Agent 已生成）",
                        "filename": filename,
                        "url": f"/api/download/{task_id}/{filename}" if task_id else f"/api/download/{filename}",
                        "timestamp": self._timestamp(),
                    }
                    logger.info(f"Agent Write 追踪推送: {filename}")
                except Exception as e:
                    logger.warning(f"文件推送失败 {fp}: {e}")

        if not pushed:
            logger.info("Write 追踪无文件，回退到时间戳扫描")
            for dl in self._detect_downloads(task_id, message, task_start):
                yield dl

        yield {
            "type": "done",
            "message": "Agent 任务执行完毕",
            "timestamp": self._timestamp(),
        }

    async def _run_claude(
        self, message: str, task_id: str = "", is_continuation: bool = False
    ) -> AsyncGenerator[dict, None]:
        """通过子进程运行 claude --print（stdin 管道传 prompt，避免命令行长度限制）。"""
        import shutil

        if not shutil.which("claude"):
            raise RuntimeError("claude CLI 未安装")

        # prompt 已由 Hermes/上层构建完毕，ClaudeAgent 不再注入任何额外内容
        prompt = message

        if sys.platform == "linux":
            # Linux: 用 su - uniemail 绕过 root bypassPermissions 限制
            cmd = [
                "su", "-", "uniemail", "-c",
                " ".join([
                    "claude",
                    "--print",
                    "--output-format", "stream-json",
                    "--verbose",
                    "--no-session-persistence",
                    "--permission-mode", "bypassPermissions",
                    "--allowedTools", json.dumps(ALLOWED_TOOLS),
                    "--max-budget-usd", "20.0",
                ])
            ]
        else:
            cmd = [
                "claude",
                "--print",
                "--output-format", "stream-json",
                "--verbose",
                "--no-session-persistence",
                "--permission-mode", "bypassPermissions",
                "--allowedTools", json.dumps(ALLOWED_TOOLS),
                "--max-budget-usd", "20.0",
            ]

        env = os.environ.copy()

        # 记录任务启动时间戳，供后续文件检测隔离旧文件
        task_start = time.time()

        # 调试：记录事件循环信息
        try:
            _loop = asyncio.get_running_loop()
            logger.info(f"事件循环类型: {type(_loop).__name__}, "
                        f"has_subprocess_exec={hasattr(_loop, 'subprocess_exec')}, "
                        f"policy={type(asyncio.get_event_loop_policy()).__name__}")
        except Exception:
            pass

        # Windows 上 SelectorEventLoop 不支持 create_subprocess_exec，
        # 使用线程 + subprocess.Popen + asyncio.Queue 桥接方案
        if sys.platform == "win32":
            queue = asyncio.Queue()
            self._start_subprocess_thread(cmd, env, queue, task_id, prompt)
            async for log in self._process_stream(queue, task_id, task_start, message):
                yield log
            return

        try:
            process = await asyncio.create_subprocess_exec(
                *cmd,
                stdin=asyncio.subprocess.PIPE,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                env=env,
            )
            self.active_procs[task_id] = process
            # 通过 stdin 发送 prompt（纯文本，避免命令行长度限制）
            try:
                process.stdin.write(prompt.encode("utf-8"))
                await process.stdin.drain()
                process.stdin.close()
            except Exception as stdin_err:
                logger.warning(f"stdin 写入异常: {stdin_err}")
        except NotImplementedError as nie:
            logger.warning(f"create_subprocess_exec 不可用，回退线程方案: {nie}")
            queue = asyncio.Queue()
            self._start_subprocess_thread(cmd, env, queue, task_id, prompt)
            async for log in self._process_stream(queue, task_id, task_start, message):
                yield log
            return
        except Exception as proc_err:
            import traceback as _tb3
            _detail2 = _tb3.format_exc()
            logger.error(f"子进程启动异常 详细: {_detail2}")
            raise RuntimeError(f"无法启动 claude 子进程: {type(proc_err).__name__}: {str(proc_err)[:200]}") from proc_err

        async for log in self._process_stream(process, task_id, task_start, message):
            yield log

    def _parse_file_declarations(self, text: str) -> list[tuple[str, str]]:
        """从 Agent 输出文本中解析 [FILES]...[/FILES] 文件声明块。

        格式：每行 `文件名 | 简短描述`
        返回 [(filename, label), ...] 列表。"""
        import re as _re

        results: list[tuple[str, str]] = []
        pattern = r"\[FILES\]\s+(.*?)\[/FILES\]"
        for match in _re.finditer(pattern, text, _re.DOTALL | _re.IGNORECASE):
            for line in match.group(1).strip().split("\n"):
                line = line.strip()
                if not line or line.startswith("#") or line.startswith("//"):
                    continue
                # 格式：filename | label
                parts = line.split("|", 1)
                filename = parts[0].strip().strip("`").strip('"').strip("'")
                label = parts[1].strip() if len(parts) > 1 else filename
                if filename:
                    results.append((filename, label))
        return results

    def _detect_downloads(
        self, task_id: str = "", message: str = "", after_timestamp: float = 0
    ) -> list[dict]:
        """检测任务执行期间新生成的 CSV 文件，推送 CSV + XLSX 下载。

        隔离策略：
        1. 递归扫描 outputs/ 下所有 CSV（不限目录，claude CLI 可能自建目录）
        2. 只认 mtime > after_timestamp 的文件（杜绝旧残留）
        3. 按大学名称匹配度 + 数据行数评分，选取最佳
        """
        from agent.exporter import _BASE_OUTPUT_DIR
        import csv as _csv
        import re as _re

        results: list[dict] = []

        # ── 递归收集 CSV：限定在 task 子目录 + root 临时文件 ──
        candidates: list[Path] = []

        # 1. 优先搜索任务专属目录
        if task_id:
            safe_tid = task_id.replace("/", "_").replace("\\", "_")
            task_dir = _BASE_OUTPUT_DIR / safe_tid
            if task_dir.exists():
                for f in task_dir.rglob("*.csv"):
                    try:
                        if f.stat().st_mtime < after_timestamp - 10 or f.stat().st_size < 200:
                            continue
                        candidates.append(f)
                    except OSError:
                        continue

        # 2. 兼容旧版：root outputs/ 下的 CSV（仅限直接子文件，不递归）
        for f in _BASE_OUTPUT_DIR.glob("*.csv"):
            try:
                if f.stat().st_mtime < after_timestamp - 10 or f.stat().st_size < 200:
                    continue
                candidates.append(f)
            except OSError:
                continue

        if not candidates:
            logger.info("任务期间未生成新的 CSV 文件")
            return results

        logger.info(
            f"发现 {len(candidates)} 个新 CSV: {[c.name for c in candidates]}"
        )

        # ── 从用户消息提取目标大学名称 ──
        target_uni = ""
        if message:
            m = _re.search(r"([一-鿿]{2,4}(?:大学|学院))", message)
            if m:
                target_uni = m.group(1)

        # ── 评分：匹配度 + 数据行数 ──
        def _score_file(f: Path) -> tuple:
            # 大学名称匹配
            uni_score = 0
            if target_uni:
                if target_uni in f.stem:
                    uni_score = 100
                else:
                    for i in range(len(target_uni), 1, -1):
                        if target_uni[:i] in f.stem:
                            uni_score = i
                            break
            # 数据行数
            rows = 0
            try:
                with open(f, "r", encoding="utf-8-sig") as fp:
                    reader = _csv.reader(fp)
                    next(reader, None)  # 跳过表头
                    for row in reader:
                        if row and any(cell.strip() for cell in row):
                            rows += 1
            except Exception:
                pass
            return (uni_score, rows, f.stat().st_size)

        scored = [(_score_file(f), f) for f in candidates]
        scored.sort(key=lambda x: x[0], reverse=True)

        # ── 推送所有有效的新 CSV 文件（不只是最佳的一个） ──
        pushed_names: set[str] = set()
        for score_val, csv_path in scored:
            if score_val[1] == 0:  # 无数据行，跳过
                continue

            csv_name = csv_path.name
            if csv_name in pushed_names:
                continue
            pushed_names.add(csv_name)

            # 根据文件实际位置生成正确的下载 URL
            rel = csv_path.relative_to(_BASE_OUTPUT_DIR)
            if len(rel.parts) == 1:
                csv_url = f"/api/download/{csv_name}"
            else:
                csv_url = f"/api/download/{rel.parts[0]}/{csv_name}"

            results.append({
                "type": "download",
                "message": f"CSV: {csv_name}",
                "filename": csv_name,
                "url": csv_url,
                "timestamp": self._timestamp(),
            })

            logger.info(f"推送下载: {csv_name} (行数={score_val[1]})")

        if not results:
            logger.warning("新 CSV 文件无有效数据行")
        return results

    def _parse_stream_line(self, line: str) -> dict | None:
        """解析 stream-json 行，提取可展示的文本片段。"""
        try:
            data = json.loads(line)
        except (json.JSONDecodeError, TypeError):
            return None

        msg_type = data.get("type", "")

        # assistant 消息
        if msg_type == "assistant":
            message = data.get("message", {})
            content_list = message.get("content", [])
            chunks = []
            tool_chunks = []  # 工具调用相关的日志，与真实 AI 回复内容分离
            tool_uses = []  # 收集本条消息中的工具调用信息
            for item in content_list:
                if item.get("type") == "text":
                    text = item.get("text", "").strip()
                    if text:
                        chunks.append(text)
                elif item.get("type") == "tool_use":
                    tool_name = item.get("name", "unknown")
                    tool_input = item.get("input", {})
                    tool_uses.append({"name": tool_name, "input": tool_input})
                    # 工具调用: 从 _chunks 中分离出来，单独用 _tool_chunks 放置
                    # text chunks 保留在 _chunks 中作为真实 AI 回复内容
                    tool_chunks.append(f"🔧 调用工具: {tool_name}")
                    input_str = json.dumps(tool_input, ensure_ascii=False)
                    if len(input_str) < 200:
                        tool_chunks.append(f"   参数: {input_str}")

            return {"_msg_type": "assistant", "_chunks": chunks, "_tool_chunks": tool_chunks, "_tool_uses": tool_uses}

        # tool_result
        if msg_type == "user" and data.get("message", {}).get("role") == "user":
            content_list = data.get("message", {}).get("content", [])
            for item in content_list:
                if item.get("type") == "tool_result":
                    result_text = item.get("content", "")
                    if isinstance(result_text, list):
                        result_text = " ".join(
                            r.get("text", "") for r in result_text if isinstance(r, dict)
                        )
                    if result_text:
                        text = str(result_text)
                        # 超大结果（如文件内容）只显示摘要
                        if len(text) > 500:
                            text = text[:200] + f"\n... (共 {len(text)} 字符，已截断)"
                        return {
                            "_msg_type": "assistant",
                            "_chunks": [f"📋 结果: {text}"],
                        }
            return None

        # 最终结果
        if msg_type == "result":
            subtype = data.get("subtype", "")
            is_error = subtype != "success"
            result_text = data.get("result", "")
            duration = data.get("duration_ms", 0)
            cost = data.get("total_cost_usd", 0)

            # 结果可能很长，截断
            if isinstance(result_text, str):
                result_text = result_text[:1000]

            return {
                "_msg_type": "result",
                "_text": result_text,
                "_subtype": subtype,
                "is_error": is_error,
                "duration_ms": duration,
                "cost_usd": cost,
            }

        # 系统消息
        if msg_type == "system":
            return {
                "_msg_type": "system",
                "subtype": data.get("subtype", ""),
                "model": data.get("model", ""),
            }

        return None

    async def _respond_conversational(self, message: str) -> AsyncGenerator[dict, None]:
        """轻量级直接 API 响应非爬取类日常会话，避免拉起 CLI 进程。
        
        如果配置了 API Key，直接请求大模型 API，否则使用本地规则响应常见问答。
        """
        import os

        api_key = os.environ.get("DEEPSEEK_API_KEY") or os.environ.get("OPENAI_API_KEY") or os.environ.get("ANTHROPIC_API_KEY")

        # 有 API Key 时优先走大模型，仅无 Key 时使用本地规则兜底
        if not api_key:
            lower_msg = message.strip().lower()
        
            # 1. 你好等问候
            if any(x in lower_msg for x in ["你好", "hello", "hi", "hey"]):
                yield {
                    "type": "text",
                    "message": "你好！我是 **UniEmail Agent**。\n\n我是一个专为学术界和高校设计的教师邮箱自动抓取与分析助手。我可以帮你自动爬取各高校官网的教师邮箱，并支持导出 CSV/XLSX 等多种格式。\n\n你可以对我说：“帮我抓取南京大学计算机学院的教师邮箱”，或者问我高校抓取的规则与支持的学校！",
                    "timestamp": self._timestamp()
                }
                yield {
                    "type": "done",
                    "message": "已完成本地秒回响应",
                    "timestamp": self._timestamp()
                }
                return
            
            # 2. 你是谁/自我介绍
            if any(x in lower_msg for x in ["你是谁", "介绍", "功能", "who are you"]):
                yield {
                    "type": "text",
                    "message": "我是 **UniEmail Agent**！\n\n### 🌟 我的主要功能包括：\n1. **多源多路径爬取**：深度挖掘高校院系「师资队伍」页面，支持个人详情页深度抓取。\n2. **智能数据清洗**：清洗各种复杂的混淆字符（如 `[at]` -> `@`），过滤无效或公共邮箱。\n3. **多格式数据导出**：支持 CSV, Excel (XLSX), Markdown, HTML, PDF, Word (DOCX) 导出。\n4. **任务与文件隔离**：每个任务专属输出目录，防止数据交叉污染。\n5. **全局技能库**：从过往成功的抓取中自动沉淀提取高校官网的最佳 DOM 选择器和反爬经验。\n\n你可以尝试输入高校抓取任务，我会立即为你启动浏览器自动化爬取进程！",
                    "timestamp": self._timestamp()
                }
                yield {
                    "type": "done",
                    "message": "已完成本地秒回响应",
                    "timestamp": self._timestamp()
                }
                return

        # 3. 尝试调用大模型 API (优先使用 DeepSeek，其次是 OpenAI/Anthropic 兼容 fallback)
        if api_key:
            try:
                # 1. 优先使用 DeepSeek 官方通道
                if os.environ.get("DEEPSEEK_API_KEY"):
                    import openai
                    from openai import AsyncOpenAI
                    
                    client = AsyncOpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")
                    
                    response = await client.chat.completions.create(
                        model="deepseek-chat",
                        messages=[
                            {"role": "system", "content": "You are UniEmail Agent, a helpful assistant specialized in scraping university faculty emails."},
                            {"role": "user", "content": message}
                        ],
                        stream=True
                    )
                    
                    async for chunk in response:
                        content = chunk.choices[0].delta.content or ""
                        if content:
                            yield {
                                "type": "text",
                                "message": content,
                                "timestamp": self._timestamp()
                            }
                    yield {
                        "type": "done",
                        "message": "DeepSeek API 响应完毕",
                        "timestamp": self._timestamp()
                    }
                    return
                # 2. 回退使用 OpenAI (支持中转)
                elif os.environ.get("OPENAI_API_KEY"):
                    import openai
                    from openai import AsyncOpenAI
                    
                    # 自动读取环境变量中的 OPENAI_BASE_URL (如配置了中转或 DeepSeek 中转)
                    base_url = os.environ.get("OPENAI_BASE_URL") or None
                    client = AsyncOpenAI(api_key=api_key, base_url=base_url)
                    
                    # 如果 base_url 含有 deepseek，使用 deepseek-chat，否则使用 gpt-4o-mini
                    model = os.environ.get("OPENAI_API_MODEL") or ("deepseek-chat" if base_url and "deepseek" in base_url else "gpt-4o-mini")
                    
                    response = await client.chat.completions.create(
                        model=model,
                        messages=[
                            {"role": "system", "content": "You are UniEmail Agent, a helpful assistant specialized in scraping university faculty emails."},
                            {"role": "user", "content": message}
                        ],
                        stream=True
                    )
                    
                    async for chunk in response:
                        content = chunk.choices[0].delta.content or ""
                        if content:
                            yield {
                                "type": "text",
                                "message": content,
                                "timestamp": self._timestamp()
                            }
                    yield {
                        "type": "done",
                        "message": "OpenAI/DeepSeek-API 响应完毕",
                        "timestamp": self._timestamp()
                    }
                    return
                # 3. 回退使用 Anthropic
                elif os.environ.get("ANTHROPIC_API_KEY"):
                    # 使用 httpx 直接调用 Anthropic API 保证轻量无依赖
                    import httpx
                    headers = {
                        "x-api-key": api_key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json"
                    }
                    data = {
                        "model": os.environ.get("ANTHROPIC_API_MODEL") or "claude-3-5-sonnet-latest",
                        "messages": [{"role": "user", "content": message}],
                        "max_tokens": 1024,
                        "stream": True
                    }
                    # 自动读取环境变量中的 ANTHROPIC_BASE_URL (如配置了中转)
                    base_url = os.environ.get("ANTHROPIC_BASE_URL") or "https://api.anthropic.com"
                    url = f"{base_url.rstrip('/')}/v1/messages"
                    
                    # 异步 HTTP 流式调用
                    async with httpx.AsyncClient(timeout=30.0) as client:
                        async with client.stream("POST", url, headers=headers, json=data) as r:
                            if r.status_code == 200:
                                async for line in r.aiter_lines():
                                    line = line.strip()
                                    if line.startswith("data:"):
                                        line_data = line[5:].strip()
                                        if line_data == "[DONE]":
                                            break
                                        try:
                                            evt = json.loads(line_data)
                                            if evt.get("type") == "content_block_delta":
                                                delta_text = evt.get("delta", {}).get("text", "")
                                                if delta_text:
                                                    yield {
                                                        "type": "text",
                                                        "message": delta_text,
                                                        "timestamp": self._timestamp()
                                                    }
                                        except Exception:
                                            pass
                                yield {
                                    "type": "done",
                                    "message": "Anthropic API 响应完毕",
                                    "timestamp": self._timestamp()
                                }
                                return
                            else:
                                logger.warning(f"Anthropic API error: {r.status_code}")
            except Exception as e:
                logger.error(f"直接调用大模型 API 异常: {e}")
                # 异常后自动向下执行进入友好提示

        # 4. 无 Key 时的默认优雅反馈
        yield {
            "type": "text",
            "message": "您好！检测到您发送的是日常闲聊/通用问答。由于您当前尚未在环境变量中配置 `DEEPSEEK_API_KEY` 或 `OPENAI_API_KEY`，系统目前仅对**高校教师邮箱爬取任务**自动启用全套 Claude Code CLI 流程。\n\n*   **若要直接爬取**：请使用类似于 **“帮我抓取南京大学计算机学院教师邮箱”** 的指令，系统将立即为您拉起自动化浏览器抓取进程。\n*   **若要启用闲聊/通用问答**：建议在运行环境（或 `.env` 文件）中配置 `DEEPSEEK_API_KEY` 环境变量，系统即可自动通过高速轻量的 DeepSeek `deepseek-chat` 大模型 API 实时响应您的普通日常问答，实现秒回体验！",
            "timestamp": self._timestamp()
        }
        yield {
            "type": "done",
            "message": "已完成本地友好提示响应",
            "timestamp": self._timestamp()
        }
