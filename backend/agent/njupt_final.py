"""
南京邮电大学教师邮箱爬虫 v5 最终版
使用从 cs.njupt.edu.cn 导航发现的正确学院URL
"""
import asyncio
import csv
import re
import sys
import logging
from datetime import datetime
from pathlib import Path
from collections import Counter

try:
    from playwright.async_api import async_playwright
except ImportError:
    print("pip install playwright")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

TASK_ID = "f8d29d14-aa64-4781-8efa-ee32cd310ec5"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / TASK_ID
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

TITLE_KW = [
    "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
    "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
    "长江学者", "杰青", "优青", "博士后", "实验师", "高级实验师",
    "青年专聘教授", "校聘副教授", "预聘副教授", "特聘教授",
    "客座教授", "名誉教授", "国家级教学名师",
]

PUBLIC_LOCAL = {"webmaster", "admin", "info", "office", "master", "president",
                "xb", "xxgk", "jwc", "yjsc", "rsc", "gjc", "tw", "xsc",
                "jsjsj", "jsjyz", "jsjxy", "sxydw", "njugcglxy", "gcglxydw",
                "njudz", "jd-iam", "iamdirector", "yb_eng", "yb_"}

NAV_KW = {"学院概况", "组织机构", "机构设置", "新闻", "通知", "公告", "招生",
          "党建", "学生工作", "团学", "工会", "校友", "实验室", "联系我们",
          "下载", "规章制度", "师德师风", "信息公开", "网站地图",
          "师资队伍", "师资概况", "教师名录", "专任教师", "导师介绍",
          "现任领导", "学院领导", "领导信箱", "诚聘英才",
          "本科生", "研究生", "专业介绍", "培养", "教学", "教师信息",
          "科研", "学科", "学术", "竞赛", "创新", "导师名录",
          "返回", "首页", "南邮主页", "校内链接", "更多", "详情",
          "下一页", "上一页", "下一页", "尾页", "末页",
          "函授", "自考", "培训", "技能", "干部", "成教",
          "网上学习", "表格下载", "学习资源", "综合管理", "教育捐赠",
          "监督方式", "培养计划", "实践教学", "专接本", "二学历",
          "社会自考", "专业计划", "教学管理", "科研管理",
          "校级", "招聘", "发表论文", "授权专利", "科研获奖",
          "成果展示", "本科生录", "研究生录", "行政工作", "党务活动",
          "工会活动", "学生社团", "学生荣誉", "风采展示", "校友动态",
          "羽毛球队", "第十三周", "拟发展", "博士后", "行政教辅",
          "学院领导", "党务动态", "关于青书", "个人信息"}

# 正确URL的学院配置 (从 cs.njupt.edu.cn 导航获得)
COLLEGES = [
    ("通信与信息工程学院", "https://scie.njupt.edu.cn"),
    ("电子与光学工程学院/微电子学院", "https://eoe.njupt.edu.cn"),
    ("计算机学院/软件学院/网络空间安全学院", "https://cs.njupt.edu.cn"),
    ("自动化学院/人工智能学院", "https://coa.njupt.edu.cn"),
    ("材料科学与工程学院", "https://iam.njupt.edu.cn"),
    ("物联网学院", "https://ciot.njupt.edu.cn"),
    ("地理与生物信息学院", "https://cgb.njupt.edu.cn"),
    ("现代邮政学院", "https://simp.njupt.edu.cn"),
    ("传媒与艺术学院", "https://cm.njupt.edu.cn"),
    ("管理学院", "https://bc.njupt.edu.cn"),
    ("经济学院", "https://jjxy.njupt.edu.cn"),
    ("马克思主义学院", "https://marxism.njupt.edu.cn"),
    ("社会与人口学院", "https://sps.njupt.edu.cn"),
    ("外国语学院", "https://fld.njupt.edu.cn"),
    ("教育科学与技术学院", "https://edu.njupt.edu.cn"),
    ("贝尔英才学院", "https://bhs.njupt.edu.cn"),
    ("海外教育学院", "https://overseas.njupt.edu.cn"),
]


def is_chinese_name(text: str) -> bool:
    text = text.strip()
    if not re.fullmatch(r"[一-鿿]{2,4}", text):
        return False
    if text in NAV_KW:
        return False
    if text[-1] in "报组室部处委会局办系院所馆站网栏目页版":
        return False
    if any(text.endswith(s) for s in ["学院", "大学", "中心", "研究所", "实验室"]):
        return False
    return True


def is_public_email(email: str) -> bool:
    if not email:
        return True
    local = email.lower().split("@")[0]
    if any(kw in local for kw in PUBLIC_LOCAL):
        return True
    # 排除明显非教师邮箱
    if "@" in email.lower():
        domain = email.lower().split("@")[1]
        if domain not in ("njupt.edu.cn", "njupt.edu", "nju.edu.cn",
                           "126.com", "163.com", "qq.com", "gmail.com",
                           "hotmail.com", "outlook.com", "foxmail.com",
                           "aliyun.com", "sina.com", "yeah.net"):
            return True
    return False


def parse_at_sign(text: str) -> str:
    text = re.sub(r"\s*\[at\]\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*\(at\)\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*#@\s*", "@", text)
    text = re.sub(r"\s*\[@\]\s*", "@", text)
    text = re.sub(r"\s*\[dot\]\s*", ".", text, flags=re.IGNORECASE)
    return text


def extract_titles(text: str) -> str:
    found = []
    for kw in TITLE_KW:
        if kw in text:
            found.append(kw)
    return "、".join(found[:5])


async def find_and_scrape_college(context, name: str, site_url: str) -> list[dict]:
    """一站式处理一个学院：找教师列表页 → 收集链接 → 爬取详情"""
    results = []
    page = await context.new_page()

    teacher_list_urls = []

    try:
        # 访问学院首页
        resp = await page.goto(site_url, wait_until="domcontentloaded", timeout=20000)
        if resp and resp.status >= 400:
            await page.close()
            logger.warning(f"  {name}: HTTP {resp.status}")
            return results
        await asyncio.sleep(2)

        # 找教师列表入口
        links = await page.evaluate("""() => {
            const ls = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const t = (a.textContent || '').trim();
                const h = a.href;
                if (!t || !h || h.startsWith('javascript:') || h === '#' || h.includes('mailto:')) return;
                ls.push({t: t.substring(0, 60), h: h});
            });
            return ls;
        }""")

        for l in links:
            if any(kw in l["t"] for kw in ["师资队伍", "师资概况", "教师名录", "专任教师",
                                              "导师介绍", "教师队伍", "导师队伍", "队伍建设"]):
                if len(l["t"]) < 15:
                    teacher_list_urls.append(l["h"])

        # 如果没找到标准入口，尝试常见路径
        if not teacher_list_urls:
            common_paths = [
                f"{site_url.rstrip('/')}/szdw/list.htm",
                f"{site_url.rstrip('/')}/szdw.htm",
                f"{site_url.rstrip('/')}/list.htm",
            ]
            teacher_list_urls = common_paths

        logger.info(f"  {name}: {len(teacher_list_urls)} 个教师入口")

        # 访问每个教师列表页
        all_entries = []
        seen_urls = set()

        for turl in teacher_list_urls[:3]:  # 最多3个入口
            try:
                await page.goto(turl, wait_until="domcontentloaded", timeout=15000)
                await asyncio.sleep(2)

                t_links = await page.evaluate("""() => {
                    const ls = [];
                    document.querySelectorAll('a[href]').forEach(a => {
                        const t = (a.textContent || '').trim();
                        const h = a.href;
                        if (!t || !h || h.startsWith('javascript:') || h === '#' || h.includes('mailto:')) return;
                        ls.push({t: t.substring(0, 60), h: h});
                    });
                    return ls;
                }""")

                for l in t_links:
                    text = l["t"]
                    url = l["h"]
                    # 中文姓名
                    if is_chinese_name(text):
                        if url not in seen_urls:
                            seen_urls.add(url)
                            all_entries.append({"name": text, "url": url, "college": name})
                    # yjs导师系统详情页
                    elif "yjs.njupt.edu.cn" in url and "dsfcxq" in url:
                        if url not in seen_urls:
                            seen_urls.add(url)
                            all_entries.append({"name": text, "url": url, "college": name})

                logger.info(f"    {turl}: {len(all_entries)} 教师")
            except Exception:
                pass

        await page.close()

        if not all_entries:
            return results

        # 批量爬取详情页
        sem = asyncio.Semaphore(4)

        async def process(entry):
            async with sem:
                p2 = await context.new_page()
                try:
                    await p2.goto(entry["url"], wait_until="domcontentloaded", timeout=15000)
                    await asyncio.sleep(0.6)

                    # 提取姓名
                    fname = entry["name"]
                    if not is_chinese_name(fname):
                        title_t = await p2.evaluate("() => document.title || ''")
                        m = re.match(r"^([一-鿿]{2,4})\s*[-–—|]", title_t)
                        if m and is_chinese_name(m.group(1)):
                            fname = m.group(1)
                        else:
                            bt = await p2.evaluate("() => document.body ? document.body.innerText : ''")
                            m2 = re.search(r"姓\s*名\s*[：:]\s*([一-鿿]{2,4})", bt)
                            if m2:
                                fname = m2.group(1)

                    if not is_chinese_name(fname):
                        return None

                    # 邮箱
                    bt = await p2.evaluate("() => document.body ? document.body.innerText : ''")
                    bt = parse_at_sign(bt)
                    html = await p2.evaluate("() => document.body ? document.body.innerHTML : ''")
                    html = parse_at_sign(html)

                    email = ""
                    em = re.search(r"电子邮箱\s*[：:]\s*(\S+@\S+)", bt)
                    if em:
                        email = em.group(1).strip()
                    else:
                        emails = EMAIL_RE.findall(bt + " " + html)
                        email = next((e for e in emails if not is_public_email(e)), "")

                    if is_public_email(email):
                        email = ""

                    title = extract_titles(bt[:3000])

                    if email:
                        logger.info(f"    ✅ {fname} → {email} [{name}]")

                    return {
                        "姓名": fname,
                        "邮箱": email,
                        "学院": name,
                        "职称": title,
                        "主页链接": entry["url"],
                    }
                except Exception:
                    return None
                finally:
                    await p2.close()

        for start in range(0, len(all_entries), 15):
            end = min(start + 15, len(all_entries))
            batch = all_entries[start:end]
            tasks = [process(e) for e in batch]
            res = await asyncio.gather(*tasks)
            results.extend([r for r in res if r])

        logger.info(f"  {name}: 获得 {len(results)} 条 ({sum(1 for r in results if r['邮箱'])} 邮箱)")

    except Exception as e:
        logger.warning(f"  {name} 失败: {str(e)[:80]}")
    finally:
        try:
            await page.close()
        except Exception:
            pass

    return results


async def main():
    logger.info("=" * 60)
    logger.info("🎓 南京邮电大学 全学院教师邮箱爬虫 v5")
    logger.info(f"   共 {len(COLLEGES)} 个学院")
    logger.info("=" * 60)

    all_results = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36"

        # 逐个学院处理
        for col_name, col_url in COLLEGES:
            logger.info(f"\n🏫 {col_name}")
            try:
                ctx = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
                results = await find_and_scrape_college(ctx, col_name, col_url)
                all_results.extend(results)
                await ctx.close()
            except Exception as e:
                logger.warning(f"  ❌ {col_name}: {str(e)[:60]}")

        await browser.close()

    # === 清洗 ===
    logger.info(f"\n{'='*60}")
    logger.info("📊 数据清洗")
    logger.info(f"  原始: {len(all_results)}")

    # 去重
    seen = set()
    dedup = []
    for r in all_results:
        key = (r["姓名"], r["邮箱"])
        if key not in seen:
            seen.add(key)
            dedup.append(r)

    # 过滤
    final = [r for r in dedup if is_chinese_name(r["姓名"])]

    with_email = [r for r in final if r["邮箱"]]
    no_email = [r for r in final if not r["邮箱"]]
    logger.info(f"  去重: {len(dedup)} → 过滤: {len(final)}")
    logger.info(f"  有邮箱: {len(with_email)}, 无邮箱: {len(no_email)}")

    # 学院统计
    cc = Counter(r["学院"] for r in final)
    logger.info(f"\n📊 各学院统计:")
    for c, cnt in cc.most_common():
        e = len([r for r in final if r["学院"] == c and r["邮箱"]])
        logger.info(f"   {c}: {cnt} (有邮箱: {e})")

    if with_email:
        logger.info(f"\n📋 有邮箱教师 (前30条):")
        for r in with_email[:30]:
            logger.info(f"   {r['姓名']} <{r['邮箱']}> [{r['学院']}]")

    # === XLSX ===
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"南京邮电大学_教师邮箱_{ts}.xlsx"

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教师邮箱"
    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]

    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    cf = Font(name="微软雅黑", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = tb

    for i, r in enumerate(final, 1):
        row = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.font = cf; cell.alignment = Alignment(vertical="center"); cell.border = tb

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 60
    wb.save(xlsx_path)
    logger.info(f"\n💾 XLSX: {xlsx_path}")

    # 无邮箱
    if no_email:
        ne = OUTPUT_DIR / f"南京邮电大学_无邮箱教师_{ts}.csv"
        with open(ne, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["姓名", "邮箱", "学院", "职称", "主页链接"])
            w.writeheader()
            w.writerows(no_email)
        logger.info(f"💾 无邮箱: {ne}")

    logger.info(f"\n✅ 完成! {len(final)} 人, {len(with_email)} 邮箱, {len(cc)} 学院")

    print(f"\n[FILES]")
    print(f"{xlsx_path.name} | 南京邮电大学教师邮箱完整数据 ({len(final)}条, {len(with_email)}个邮箱, {len(cc)}个学院)")
    print(f"[/FILES]")


if __name__ == "__main__":
    asyncio.run(main())
