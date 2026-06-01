"""University catalog and harvested table indexing."""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import time
import urllib.request
from pathlib import Path
from typing import Any

from agent.exporter import _BASE_OUTPUT_DIR

logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent.parent / "data"
CATALOG_CACHE = DATA_DIR / "universities_catalog.json"
OFFICIAL_UNDERGRADUATE_XLS_URL = (
    "http://www.moe.gov.cn/jyb_xxgk/s5743/s5744/A03/202506/"
    "W020250729615142156867.xls"
)
EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

UNIVERSITY_985 = {
    "北京大学", "中国人民大学", "清华大学", "北京航空航天大学", "北京理工大学", "中国农业大学", "北京师范大学", "中央民族大学",
    "南开大学", "天津大学", "大连理工大学", "东北大学", "吉林大学", "哈尔滨工业大学", "复旦大学", "同济大学",
    "上海交通大学", "华东师范大学", "南京大学", "东南大学", "浙江大学", "中国科学技术大学", "厦门大学", "山东大学",
    "中国海洋大学", "武汉大学", "华中科技大学", "湖南大学", "中南大学", "国防科技大学", "中山大学", "华南理工大学",
    "四川大学", "电子科技大学", "重庆大学", "西安交通大学", "西北工业大学", "西北农林科技大学", "兰州大学",
}

UNIVERSITY_211_EXTRA = {
    "北京交通大学", "北京工业大学", "北京科技大学", "北京化工大学", "北京邮电大学", "北京林业大学", "北京中医药大学",
    "北京外国语大学", "中国传媒大学", "中央财经大学", "对外经济贸易大学", "北京体育大学", "中央音乐学院", "中国政法大学",
    "华北电力大学", "天津医科大学", "河北工业大学", "太原理工大学", "内蒙古大学", "辽宁大学", "大连海事大学",
    "延边大学", "东北师范大学", "哈尔滨工程大学", "东北农业大学", "东北林业大学", "华东理工大学", "东华大学",
    "上海外国语大学", "上海财经大学", "上海大学", "海军军医大学", "第二军医大学", "苏州大学", "南京航空航天大学",
    "南京理工大学", "中国矿业大学", "河海大学", "江南大学", "南京农业大学", "中国药科大学", "南京师范大学",
    "安徽大学", "合肥工业大学", "福州大学", "南昌大学", "中国石油大学（华东）", "中国石油大学(华东)", "郑州大学",
    "中国地质大学（武汉）", "中国地质大学(武汉)", "武汉理工大学", "华中农业大学", "华中师范大学", "中南财经政法大学",
    "湖南师范大学", "暨南大学", "华南师范大学", "海南大学", "广西大学", "西南交通大学", "西南财经大学", "四川农业大学",
    "贵州大学", "云南大学", "西藏大学", "西北大学", "西安电子科技大学", "长安大学", "陕西师范大学", "空军军医大学",
    "第四军医大学", "青海大学", "宁夏大学", "新疆大学", "石河子大学", "中国矿业大学（北京）", "中国矿业大学(北京)",
    "中国石油大学（北京）", "中国石油大学(北京)", "中国地质大学（北京）", "中国地质大学(北京)",
}

UNIVERSITY_211 = UNIVERSITY_985 | UNIVERSITY_211_EXTRA

DOUBLE_FIRST_CLASS_EXTRA = {
    "北京协和医学院", "首都师范大学", "外交学院", "中国人民公安大学", "中国音乐学院", "中央美术学院", "中央戏剧学院",
    "中国科学院大学", "天津工业大学", "天津中医药大学", "山西大学", "上海海洋大学", "上海中医药大学", "上海体育大学",
    "上海体育学院", "上海音乐学院", "上海科技大学", "南京邮电大学", "南京林业大学", "南京信息工程大学", "南京中医药大学",
    "南京医科大学", "中国美术学院", "宁波大学", "河南大学", "湘潭大学", "华南农业大学", "广州中医药大学", "南方科技大学",
    "成都理工大学", "西南石油大学", "成都中医药大学",
}
UNIVERSITY_DOUBLE_FIRST_CLASS = UNIVERSITY_211 | DOUBLE_FIRST_CLASS_EXTRA

DOMAIN_MAP = {
    "北京大学": "pku.edu.cn", "清华大学": "tsinghua.edu.cn", "南京大学": "nju.edu.cn", "东南大学": "seu.edu.cn",
    "南京邮电大学": "njupt.edu.cn", "南京理工大学": "njust.edu.cn", "南京航空航天大学": "nuaa.edu.cn",
    "北京邮电大学": "bupt.edu.cn", "复旦大学": "fudan.edu.cn", "上海交通大学": "sjtu.edu.cn", "浙江大学": "zju.edu.cn",
    "中国科学技术大学": "ustc.edu.cn", "武汉大学": "whu.edu.cn", "华中科技大学": "hust.edu.cn", "中山大学": "sysu.edu.cn",
    "西安交通大学": "xjtu.edu.cn", "哈尔滨工业大学": "hit.edu.cn", "四川大学": "scu.edu.cn", "电子科技大学": "uestc.edu.cn",
}

FALLBACK_UNIVERSITIES = [
    {"name": "北京大学", "province": "北京市", "city": "北京市", "type": "本科"},
    {"name": "清华大学", "province": "北京市", "city": "北京市", "type": "本科"},
    {"name": "南京大学", "province": "江苏省", "city": "南京市", "type": "本科"},
    {"name": "东南大学", "province": "江苏省", "city": "南京市", "type": "本科"},
    {"name": "南京邮电大学", "province": "江苏省", "city": "南京市", "type": "本科"},
    {"name": "南京理工大学", "province": "江苏省", "city": "南京市", "type": "本科"},
    {"name": "北京邮电大学", "province": "北京市", "city": "北京市", "type": "本科"},
    {"name": "浙江大学", "province": "浙江省", "city": "杭州市", "type": "本科"},
]


def _normalize_school_name(name: str) -> str:
    return str(name or "").replace("（", "(").replace("）", ")").strip()


def _decorate_university(item: dict[str, Any]) -> dict[str, Any]:
    name = item["name"]
    domain = DOMAIN_MAP.get(name, item.get("domain", ""))
    tags = []
    if name in UNIVERSITY_985:
        tags.append("985")
    if name in UNIVERSITY_211:
        tags.append("211")
    if name in UNIVERSITY_DOUBLE_FIRST_CLASS:
        tags.append("双一流")
    if not tags:
        tags.append("普通本科")
    return {
        **item,
        "domain": domain,
        "website": f"https://www.{domain}" if domain else item.get("website", ""),
        "is_985": name in UNIVERSITY_985,
        "is_211": name in UNIVERSITY_211,
        "is_double_first_class": name in UNIVERSITY_DOUBLE_FIRST_CLASS,
        "tags": tags,
    }


def _load_cache() -> list[dict[str, Any]] | None:
    try:
        if not CATALOG_CACHE.exists():
            return None
        with open(CATALOG_CACHE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list) or not data:
            return None
        return data
    except (OSError, json.JSONDecodeError):
        return None


def _save_cache(items: list[dict[str, Any]]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    payload = [_decorate_university(i) for i in items]
    tmp = CATALOG_CACHE.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, CATALOG_CACHE)


def _download_official_catalog() -> list[dict[str, Any]]:
    import pandas as pd

    with urllib.request.urlopen(OFFICIAL_UNDERGRADUATE_XLS_URL, timeout=30) as resp:
        raw = resp.read()
    df = pd.read_excel(io.BytesIO(raw), header=None)

    items: list[dict[str, Any]] = []
    province = ""
    for row in df.itertuples(index=False):
        cells = ["" if v is None else str(v).strip() for v in row]
        first = cells[0]
        province_match = re.match(r"^(.+?)[（(]\d+所[）)]$", first)
        if province_match:
            province = province_match.group(1)
            continue
        if len(cells) < 6 or cells[1] in ("", "学校名称"):
            continue
        if cells[5] != "本科":
            continue
        name = _normalize_school_name(cells[1])
        items.append({
            "name": name,
            "province": province,
            "city": cells[4] or province,
            "type": "本科",
        })
    if not items:
        raise RuntimeError("official university catalog is empty")
    return items


def load_universities(refresh: bool = False) -> list[dict[str, Any]]:
    if not refresh:
        cached = _load_cache()
        if cached:
            return cached
    try:
        items = _download_official_catalog()
        _save_cache(items)
        return _load_cache() or [_decorate_university(i) for i in items]
    except Exception as exc:
        logger.warning("failed to load official university catalog: %s", exc)
        return [_decorate_university(i) for i in FALLBACK_UNIVERSITIES]


def _tier_match(item: dict[str, Any], tier: str) -> bool:
    if not tier or tier == "全部":
        return True
    if tier == "985":
        return bool(item.get("is_985"))
    if tier == "211":
        return bool(item.get("is_211"))
    if tier == "双一流":
        return bool(item.get("is_double_first_class"))
    if tier == "普通本科":
        return not item.get("is_985") and not item.get("is_211") and not item.get("is_double_first_class")
    return True


def _candidate_output_files() -> list[Path]:
    if not _BASE_OUTPUT_DIR.exists():
        return []
    allowed = {".csv", ".xlsx", ".html", ".pdf", ".docx", ".md"}
    return [
        p for p in _BASE_OUTPUT_DIR.rglob("*")
        if p.is_file() and p.suffix.lower() in allowed and p.stat().st_size > 0
    ]


def _file_download_url(path: Path) -> str:
    rel = path.resolve().relative_to(_BASE_OUTPUT_DIR.resolve())
    if len(rel.parts) == 1:
        return f"/api/download/{rel.parts[0]}"
    return f"/api/download/{rel.parts[0]}/{'/'.join(rel.parts[1:])}"


def _file_task_id(path: Path) -> str:
    rel = path.resolve().relative_to(_BASE_OUTPUT_DIR.resolve())
    return rel.parts[0] if len(rel.parts) > 1 else ""


def output_stats_for_school(name: str, files: list[Path] | None = None) -> dict[str, Any]:
    files = files if files is not None else _candidate_output_files()
    matched = [p for p in files if name in p.name]
    table_files = [p for p in matched if p.suffix.lower() in (".csv", ".xlsx")]
    # 快速统计：不逐文件解析，直接取文件数量，行数/邮箱用最匹配的一个文件
    if table_files:
        # 取文件大小最大的一个表文件
        best = max(table_files, key=lambda p: p.stat().st_size)
        try:
            table = parse_table_file(best, limit=1)
            rows = int(table.get("raw_total", 0))
            emails = int(table.get("valid_email_count", 0))
        except Exception:
            rows, emails = 0, 0
    else:
        rows, emails = 0, 0
    return {"file_count": len(matched), "table_count": len(table_files), "row_count": rows, "valid_email_count": emails}


def build_university_response(province: str = "", tier: str = "", q: str = "") -> dict[str, Any]:
    q = q.strip()
    files = _candidate_output_files()
    all_schools = load_universities()
    school_names = {u["name"] for u in all_schools}

    # 快速统计文件数量（不逐文件解析表格）
    school_stats: dict[str, dict[str, int]] = {}
    # 同时记住每个学校最大的csv路径，供后续行数/邮箱数快速查询
    best_table: dict[str, Path] = {}
    for f in files:
        for name in school_names:
            if name in f.name:
                s = school_stats.setdefault(name, {"file_count": 0, "table_count": 0})
                s["file_count"] += 1
                if f.suffix.lower() in (".csv", ".xlsx"):
                    s["table_count"] += 1
                    if name not in best_table or f.stat().st_size > best_table[name].stat().st_size:
                        best_table[name] = f
                break

    items = []
    for item in all_schools:
        name = item["name"]
        if province and item.get("province") != province:
            continue
        if q and q not in name and q not in item.get("city", ""):
            continue
        if not _tier_match(item, tier):
            continue
        stats = school_stats.get(name, {"file_count": 0, "table_count": 0})
        rows, emails = 0, 0
        # 只在有表文件时解析一次最大文件
        if name in best_table:
            try:
                table = parse_table_file(best_table[name], limit=1)
                rows = int(table.get("raw_total", 0))
                emails = int(table.get("valid_email_count", 0))
            except Exception:
                pass
        items.append({**item, "records": {
            "file_count": stats["file_count"],
            "table_count": stats["table_count"],
            "row_count": rows,
            "valid_email_count": emails,
        }})

    provinces = sorted({i["province"] for i in load_universities()})
    grouped: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for item in sorted(items, key=lambda x: (x["province"], x["city"], x["name"])):
        grouped.setdefault(item["province"], {}).setdefault(item["city"], []).append(item)

    groups = []
    for prov, cities in grouped.items():
        city_groups = [
            {"city": city, "count": len(universities), "universities": universities}
            for city, universities in cities.items()
        ]
        groups.append({"province": prov, "count": sum(c["count"] for c in city_groups), "cities": city_groups})

    all_items = load_universities()
    return {
        "total": len(items),
        "groups": groups,
        "provinces": provinces,
        "tier_counts": {
            "985": sum(1 for i in all_items if i.get("is_985")),
            "211": sum(1 for i in all_items if i.get("is_211")),
            "双一流": sum(1 for i in all_items if i.get("is_double_first_class")),
            "普通本科": sum(1 for i in all_items if _tier_match(i, "普通本科")),
        },
    }


def get_university_records(name: str) -> dict[str, Any]:
    files = [p for p in _candidate_output_files() if name in p.name]
    table_files = [p for p in files if p.suffix.lower() in (".csv", ".xlsx")]
    records = []
    for path in sorted(files, key=lambda p: p.stat().st_mtime, reverse=True):
        row_count = 0
        valid_email_count = 0
        if path.suffix.lower() in (".csv", ".xlsx"):
            try:
                table = parse_table_file(path, limit=1)
                row_count = table["total"]
                valid_email_count = table["valid_email_count"]
            except Exception:
                pass
        records.append({
            "task_id": _file_task_id(path),
            "filename": path.name,
            "ext": path.suffix.lower().lstrip("."),
            "url": _file_download_url(path),
            "size": path.stat().st_size,
            "updated_at": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(path.stat().st_mtime)),
            "row_count": row_count,
            "valid_email_count": valid_email_count,
            "previewable": path.suffix.lower() in (".csv", ".xlsx"),
        })
    return {"name": name, "summary": output_stats_for_school(name, files), "records": records, "table_files": len(table_files)}


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                columns = [str(c or "").strip() for c in (reader.fieldnames or [])]
                rows = [{c: str(row.get(c, "") or "").strip() for c in columns} for row in reader]
            return columns, rows
        except UnicodeDecodeError as exc:
            last_error = exc
    raise last_error or ValueError("unable to read csv")


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for xlsx reading")
    import zipfile as _zipfile
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (_zipfile.BadZipFile, Exception):
        # 文件损坏或不是合法xlsx，返回空
        return [], []
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    headers = next(row_iter, None) or []
    columns = [str(c or "").strip() for c in headers]
    rows = []
    for values in row_iter:
        row = {}
        empty = True
        for idx, col in enumerate(columns):
            val = values[idx] if idx < len(values) else ""
            text = str(val or "").strip()
            if text:
                empty = False
            row[col] = text
        if not empty:
            rows.append(row)
    return columns, rows


def _email_value(row: dict[str, str]) -> str:
    for key in ("邮箱", "email", "Email", "EMAIL", "电子邮箱", "邮件"):
        if key in row:
            return str(row.get(key, "")).strip()
    for key, value in row.items():
        if "邮箱" in key or "email" in key.lower():
            return str(value).strip()
    return ""


def parse_table_file(path: Path, limit: int = 200, offset: int = 0, q: str = "", department: str = "", valid_only: bool = False) -> dict[str, Any]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        columns, rows = _read_csv(path)
    elif path.suffix.lower() == ".xlsx":
        columns, rows = _read_xlsx(path)
    else:
        raise ValueError("unsupported table file")

    def row_matches(row: dict[str, str]) -> bool:
        if q and q not in " ".join(row.values()):
            return False
        if department:
            dept = row.get("学院") or row.get("院系") or row.get("department") or row.get("Department") or ""
            if department not in dept:
                return False
        if valid_only and not EMAIL_RE.match(_email_value(row)):
            return False
        return True

    # 为每行添加原始文件索引（用于前端 CRUD）
    for idx, row in enumerate(rows):
        row["_index"] = str(idx)

    filtered = [row for row in rows if row_matches(row)]
    valid_email_count = sum(1 for row in rows if EMAIL_RE.match(_email_value(row)))
    departments = sorted({
        (row.get("学院") or row.get("院系") or row.get("department") or "").strip()
        for row in rows
        if (row.get("学院") or row.get("院系") or row.get("department") or "").strip()
    })
    page_rows = filtered[offset: offset + limit] if limit > 0 else filtered[offset:]
    return {
        "columns": columns,
        "rows": page_rows,
        "total": len(filtered),
        "raw_total": len(rows),
        "offset": offset,
        "limit": limit,
        "valid_email_count": valid_email_count,
        "departments": departments,
    }


def resolve_table_path(task_id: str, filename: str) -> Path | None:
    if not filename or ".." in filename or "/" in filename or "\\" in filename:
        return None
    base = _BASE_OUTPUT_DIR.resolve()
    if task_id:
        safe_tid = task_id.replace("/", "_").replace("\\", "_").replace("..", "_")
        candidate = (base / safe_tid / filename).resolve()
    else:
        candidate = (base / filename).resolve()
    if not str(candidate).startswith(str(base) + os.sep) or not candidate.exists():
        return None
    if candidate.suffix.lower() not in (".csv", ".xlsx"):
        return None
    return candidate


# ── CSV/XLSX 写操作 ──────────────────────────────────────────

def _write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    """将行数据写回 CSV 文件（UTF-8 BOM）。"""
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    """将行数据写回 XLSX 文件。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required for xlsx writing")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    wb.save(path)


def _write_table(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    """根据扩展名将行数据写回文件。"""
    if path.suffix.lower() == ".csv":
        _write_csv(path, columns, rows)
    elif path.suffix.lower() == ".xlsx":
        _write_xlsx(path, columns, rows)
    else:
        raise ValueError("不支持的表格文件格式")


# ── 表格行 CRUD ──────────────────────────────────────────────

def _read_all_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """读取表格全部行（内部辅助）。"""
    if path.suffix.lower() == ".csv":
        return _read_csv(path)
    elif path.suffix.lower() == ".xlsx":
        return _read_xlsx(path)
    else:
        raise ValueError("不支持的表格文件格式")


def add_table_row(path: Path, row_data: dict[str, str]) -> dict[str, Any]:
    """在表格末尾添加新行。"""
    columns, rows = _read_all_rows(path)
    new_row = {col: row_data.get(col, "") for col in columns}
    rows.append(new_row)
    _write_table(path, columns, rows)
    return {"ok": True, "row_index": len(rows) - 1}


def update_table_row(path: Path, row_index: int, row_data: dict[str, str]) -> dict[str, Any]:
    """更新指定索引的行。"""
    columns, rows = _read_all_rows(path)
    if row_index < 0 or row_index >= len(rows):
        raise ValueError(f"行索引 {row_index} 超出范围（共 {len(rows)} 行）")
    for col, val in row_data.items():
        if col in columns:
            rows[row_index][col] = val
    _write_table(path, columns, rows)
    return {"ok": True}


def delete_table_row(path: Path, row_index: int) -> dict[str, Any]:
    """删除指定索引的行。"""
    columns, rows = _read_all_rows(path)
    if row_index < 0 or row_index >= len(rows):
        raise ValueError(f"行索引 {row_index} 超出范围（共 {len(rows)} 行）")
    rows.pop(row_index)
    _write_table(path, columns, rows)
    return {"ok": True}


# ── 文件级 CRUD ──────────────────────────────────────────────

def upload_university_file(name: str, file_bytes: bytes, filename: str) -> dict[str, Any]:
    """上传文件到高校专属上传目录。"""
    safe_name = name.replace("/", "_").replace("\\", "_").replace("..", "_")
    upload_dir = _BASE_OUTPUT_DIR / "__uploads__" / safe_name
    upload_dir.mkdir(parents=True, exist_ok=True)
    dest = upload_dir / filename
    if dest.exists():
        raise FileExistsError(f"文件 {filename} 已存在")
    dest.write_bytes(file_bytes)
    return {
        "task_id": f"__uploads__/{safe_name}",
        "filename": filename,
        "url": _file_download_url(dest),
    }


def delete_university_file(task_id: str, filename: str) -> None:
    """删除输出目录中的文件。"""
    base = _BASE_OUTPUT_DIR.resolve()
    safe_tid = task_id.replace("/", "_").replace("\\", "_").replace("..", "_") if task_id else ""
    path = (base / safe_tid / filename).resolve() if safe_tid else (base / filename).resolve()
    if not str(path).startswith(str(base) + os.sep):
        raise ValueError("非法文件路径")
    if not path.exists():
        raise FileNotFoundError(f"文件 {filename} 不存在")
    path.unlink()


def rename_university_file(task_id: str, old_filename: str, new_filename: str) -> None:
    """重命名输出目录中的文件。"""
    if ".." in new_filename or "/" in new_filename or "\\" in new_filename:
        raise ValueError("新文件名包含非法字符")
    base = _BASE_OUTPUT_DIR.resolve()
    safe_tid = task_id.replace("/", "_").replace("\\", "_").replace("..", "_") if task_id else ""
    old_path = (base / safe_tid / old_filename).resolve() if safe_tid else (base / old_filename).resolve()
    new_path = (base / safe_tid / new_filename).resolve() if safe_tid else (base / new_filename).resolve()
    if not str(old_path).startswith(str(base) + os.sep) or not str(new_path).startswith(str(base) + os.sep):
        raise ValueError("非法文件路径")
    if not old_path.exists():
        raise FileNotFoundError(f"文件 {old_filename} 不存在")
    if new_path.exists():
        raise FileExistsError(f"文件 {new_filename} 已存在")
    old_path.rename(new_path)


# ── 一键清洗 ──────────────────────────────────────────────

def _is_valid_person_name(name: str) -> bool:
    """检测文本是否为合理的教师姓名（非导航文字）。"""
    name = name.strip()
    if len(name) < 2 or len(name) > 6:
        return False
    name_blacklist = {
        "书记信箱", "院长信箱", "联系我们", "网站地图", "版权声明",
        "师德师", "师资队", "现任教", "学院概",
    }
    if name in name_blacklist:
        return False
    nav_keywords = [
        "概况", "简介", "新闻", "通知", "公告", "招生", "培养", "就业",
        "学位", "学科", "科研", "学术", "党建", "工会", "校友", "捐赠",
        "师资", "教师", "教授", "博士", "硕士", "本科", "研究", "行政",
        "管理", "教职", "荣休", "访问", "系科", "教研", "诚聘", "联系",
        "欢迎", "首页", "返回", "更多", "详情", "查看", "下载", "友情",
        "版权", "网站", "地图", "登录", "邮箱",
    ]
    if any(kw in name for kw in nav_keywords):
        return False
    if re.match(r"^[A-Za-z\s]{1,15}$", name):
        return False
    if re.search(r"[0-9@#￥%&*()（）《》【】\[\]]", name):
        return False
    if not re.search(r"[一-鿿]{2,}", name):
        return False
    return True


def _is_valid_email(email: str) -> bool:
    """校验邮箱基本格式。"""
    return bool(re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email.strip()))


def _admin_email_prefixes() -> list[str]:
    return [
        "webmaster", "admin", "office", "info", "master", "root",
        "postmaster", "wxyxz", "xwcb", "bgs", "dangzheng", "yuanban",
    ]


def clean_university_tables(name: str) -> dict[str, Any]:
    """扫描高校所有表格文件，执行一键清洗：去重（按邮箱）、过滤非法姓名、排除公共邮箱。

    清洗流程：
    1. 合并高校所有 CSV/XLSX 文件中的行
    2. 按邮箱去重（保留信息最全的），无邮箱的按姓名去重保留
    3. 过滤姓名不合法的条目
    4. 排除公共邮箱
    5. 生成一个新的汇总文件（学校名+清洗后汇总.xlsx），同时导出 CSV、HTML 等多格式
    """
    files = [p for p in _candidate_output_files() if name in p.name and p.suffix.lower() in (".csv", ".xlsx")]
    if not files:
        return {"ok": False, "error": f"未找到 {name} 的表格文件", "files_cleaned": 0}

    all_rows: list[dict[str, str]] = []
    file_info = []
    for path in files:
        try:
            cols, rows = _read_all_rows(path)
            file_info.append({"filename": path.name, "before": len(rows), "path": str(path.resolve())})
            all_rows.extend(rows)
        except Exception as e:
            logger.warning(f"读取文件失败 {path.name}: {e}")

    if not all_rows:
        return {"ok": False, "error": "所有文件均为空", "files_cleaned": 0}

    total_before = len(all_rows)

    # 第1步：格式规范化
    for r in all_rows:
        for k in r:
            r[k] = r[k].strip()
        if "邮箱" in r and r["邮箱"]:
            r["邮箱"] = r["邮箱"].strip().lower()

    # 第2步：按邮箱去重（保留信息最全的），无邮箱保留
    email_groups: dict[str, list[dict]] = {}
    no_email_rows: list[dict] = []
    for r in all_rows:
        email_key = ""
        for ek in ("邮箱", "email", "Email"):
            if ek in r and r[ek]:
                email_key = r[ek].strip().lower()
                break
        if email_key and _is_valid_email(email_key):
            email_groups.setdefault(email_key, []).append(r)
        else:
            no_email_rows.append(r)

    deduped = []
    for email, group in email_groups.items():
        def score(item):
            s = 0
            if item.get("姓名"): s += 2
            if item.get("职称"): s += 1
            if item.get("学院"): s += 1
            return s
        group.sort(key=score, reverse=True)
        deduped.append(group[0])

    # 无邮箱按姓名去重
    seen_names = set()
    for r in no_email_rows:
        name_val = ""
        for nk in ("姓名", "name", "Name"):
            if nk in r and r[nk]:
                name_val = r[nk].strip()
                break
        if name_val and name_val not in seen_names:
            seen_names.add(name_val)
            deduped.append(r)

    # 第3步：过滤非法姓名
    after_dedup = len(deduped)
    deduped = [r for r in deduped if _is_valid_person_name(r.get("姓名", "") or r.get("name", "") or "")]
    bad_name_count = after_dedup - len(deduped)

    # 第4步：排除公共邮箱
    admin_prefixes = _admin_email_prefixes()
    before_admin = len(deduped)
    deduped = [
        r for r in deduped
        if not any(
            r.get(ek, "").strip().lower().startswith(p + "@")
            for ek in ("邮箱", "email", "Email")
            for p in admin_prefixes
        )
    ]
    admin_count = before_admin - len(deduped)

    # 第5步：生成一个全新的汇总文件（不修改原文件）
    from agent.exporter import export_csv, export_xlsx
    from agent.exporter import get_task_dir
    from datetime import datetime

    _output_dir = get_task_dir("__clean__")
    _output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    data = []
    for r in deduped:
        data.append({
            "name": r.get("姓名", "") or r.get("name", ""),
            "email": r.get("邮箱", "") or r.get("email", ""),
            "department": r.get("学院", "") or r.get("department", ""),
            "title": r.get("职称", "") or r.get("title", ""),
            "url": r.get("主页链接", "") or r.get("url", ""),
        })

    # 导出 CSV 和 XLSX 到一个新建的 task-like 目录中，以便下载端点能直接访问
    clean_task_id = f"__clean__/{name}_{ts}"
    get_task_dir(clean_task_id)
    summary_filename = f"{name}_清洗后汇总_{ts}"
    result_files = {}
    try:
        fp_csv = export_csv(data, name, clean_task_id)
        result_files["csv"] = f"/api/download/{clean_task_id}/{fp_csv.name}"
    except Exception as e:
        logger.warning(f"CSV 导出失败: {e}")
    try:
        fp_xlsx = export_xlsx(data, name, clean_task_id)
        result_files["xlsx"] = f"/api/download/{clean_task_id}/{fp_xlsx.name}"
    except Exception as e:
        logger.warning(f"XLSX 导出失败: {e}")

    stats = {
        "total_before": total_before,
        "total_after": len(deduped),
        "deduped": total_before - after_dedup,
        "bad_name": bad_name_count,
        "admin_removed": admin_count,
    }
    logger.info(f"高校清洗完成 {name}: {stats}")
    return {
        "ok": True,
        "name": name,
        "stats": stats,
        "files": result_files,
        "summary_filename": summary_filename,
        "source_file_count": len(file_info),
    }
