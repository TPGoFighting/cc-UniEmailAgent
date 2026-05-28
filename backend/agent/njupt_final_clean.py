"""
最终强力清洗 — 移除所有非人名条目
"""
import openpyxl, re
from datetime import datetime
from pathlib import Path
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / "f8d29d14-aa64-4781-8efa-ee32cd310ec5"

# 非人名词（真正的姓名不可能包含这些）
NON_NAME_WORDS = {
    '团队','之窗','工作','风采','活动','队伍','总览','服务',
    '概况','简介','动态','制度','法规','招聘','公告','通知',
    '指南','中心','平台','系统','网站','下载','方式',
    '领导','分工','联系','全体','师资','建设','历史',
    '实践','基地','学术','天地','学子','大事','荣誉',
    '之家','基金','产教','融合','校企','双高','协同',
    '政协','人大','代表','委员','校园','思政','学习',
    '廉政','监督','宣传','竞赛','科学','方向','交流',
    '专区','院徽','院训','学位','培养','流程','职责',
    '智慧','党建','工会','招生','创新','科研','学科',
    '校友','导师','党群','办公',
    # 本轮新增
    '教育','本科','助学','奖金','校历','查询','邮箱','寄语',
    '事务','课程','政策','规章','链接','组织','菁英','先锋',
    '会议','教务','就业','指导','文件','党团','发展','评奖',
    '评优','日常','战邮','红邮','资库','贤才','实验','旧版',
    '入口','部门','后台','社团','资源','改革','在线','院长',
    '信箱','党员','党政','学生','教学','检索','教授','副教授',
    '讲师','研究员','副','首页','返回','学院','电子邮箱',
    '邮件','正高','名单','获奖','情况','纳才','招贤','活力',
    '教科','审核','评估','人才','引进',
}

PUBLIC_EMAIL_KW = {
    'ici-yb','jksdsf','xcb','yuanzhang','shuji','bangongshi',
    'dangwei','tuanwei','xuesheng','jiaowu','keyan','renshi',
    'caiwu','houqin','tushuguan','xinxi','bangong',
    'gyyz','baojie','jd-iam','iamdirector','jsjxy',
    'jsjsj','jsjyz','njugcglxy','gcglxydw','sxydw',
}

def is_person_name(text):
    text = text.strip()
    if not re.fullmatch(r'[一-鿿]{2,4}', text):
        return False
    if text[-1] in '报组室部处委会局办系院所馆站网栏目页版':
        return False
    if text.endswith(('学院','大学','中心','研究所','实验室')):
        return False
    for w in NON_NAME_WORDS:
        if w in text:
            return False
    return True

def is_public_email(email):
    if not email:
        return False
    el = email.lower()
    for kw in PUBLIC_EMAIL_KW:
        if kw in el:
            return True
    if '@' in el:
        domain = el.split('@')[1]
        allowed = {'njupt.edu.cn','njupt.edu','nju.edu.cn',
                   '126.com','163.com','qq.com','gmail.com',
                   'hotmail.com','outlook.com','foxmail.com',
                   'aliyun.com','sina.com','yeah.net','139.com',
                   '189.cn','live.cn','msn.com','vip.qq.com',
                   'vip.163.com','sohu.com','tom.com','21cn.com',
                   'yahoo.com','yahoo.com.cn','wo.cn'}
        if domain not in allowed and not domain.endswith('.edu.cn'):
            return True
    return False

def main():
    src = OUTPUT_DIR / "南京邮电大学_教师邮箱_20260527_211815.xlsx"
    print(f"源文件: {src}")

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            data.append({
                '姓名': str(row[1]).strip(),
                '邮箱': str(row[2]).strip() if row[2] else '',
                '学院': str(row[3]).strip() if row[3] else '',
                '职称': str(row[4]).strip() if row[4] else '',
                '主页链接': str(row[5]).strip() if row[5] else '',
            })

    print(f"原始: {len(data)}")

    # 高频邮箱检测
    email_counts = Counter(r['邮箱'] for r in data if r['邮箱'])
    freq_public = {e for e, c in email_counts.items() if c >= 3}

    # 清洗
    clean = []
    removed_name = 0
    removed_email = 0
    for r in data:
        if not is_person_name(r['姓名']):
            removed_name += 1
            continue
        if r['邮箱']:
            if is_public_email(r['邮箱']) or r['邮箱'].lower() in freq_public:
                r['邮箱'] = ''
                removed_email += 1
        clean.append(r)

    print(f"非人名: {removed_name}, 公共邮箱: {removed_email}")
    print(f"清洗后: {len(clean)}")

    with_email = [r for r in clean if r['邮箱']]

    # 低质量学院检测
    cc = Counter(r['学院'] for r in clean)
    cc_email = Counter(r['学院'] for r in with_email)
    low_quality = set()
    for c, total in cc.items():
        emails = cc_email.get(c, 0)
        rate = emails/total*100 if total>0 else 0
        # 大样本<10%邮箱率 OR 小样本0邮箱 → 低质量
        if (total >= 6 and rate < 10) or (total < 6 and emails == 0):
            low_quality.add(c)
            print(f"  ⚠️ {c}: {emails}/{total} ({rate:.1f}%) → 低质量")

    # 移除低质量学院无邮箱条目
    final = []
    removed_lq = 0
    for r in clean:
        if r['学院'] in low_quality and not r['邮箱']:
            removed_lq += 1
            continue
        final.append(r)

    print(f"移除低质量无邮箱: {removed_lq}")
    print(f"最终: {len(final)}")

    with_email_f = [r for r in final if r['邮箱']]
    no_email_f = [r for r in final if not r['邮箱']]

    cc_final = Counter(r['学院'] for r in final)
    cc_fe = Counter(r['学院'] for r in with_email_f)
    print(f"\n最终各学院:")
    for c, total in cc_final.most_common():
        e = cc_fe.get(c, 0)
        print(f"  {c}: {total} ({e}邮箱)")

    # 导出 XLSX
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    hdrs = ['序号','姓名','邮箱','学院','职称','主页链接']
    hf = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    hfl = PatternFill(start_color='10A37F', end_color='10A37F', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center')
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),
                top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    cf = Font(name='微软雅黑', size=10)
    ca = Alignment(vertical='center')

    for records, suffix in [(final, f'南京邮电大学_教师邮箱_{ts}'),
                              (with_email_f, f'南京邮电大学_有邮箱教师_{ts}')]:
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = '教师邮箱'
        for col, h in enumerate(hdrs, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfl; c.alignment = ha; c.border = tb
        for i, r in enumerate(records, 1):
            row = [i, r['姓名'], r['邮箱'], r['学院'], r['职称'], r['主页链接']]
            for col, val in enumerate(row, 1):
                c = ws2.cell(row=i+1, column=col, value=val)
                c.font = cf; c.alignment = ca; c.border = tb
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 35
        ws2.column_dimensions['D'].width = 35
        ws2.column_dimensions['E'].width = 20
        ws2.column_dimensions['F'].width = 60
        path = OUTPUT_DIR / f'{suffix}.xlsx'
        wb2.save(path)
        print(f"💾 {path.name}")

    print(f"\n✅ {len(final)}条, {len(with_email_f)}邮箱, {len(cc_final)}学院")

    print(f"\n[FILES]")
    print(f"{suffix}.xlsx | 南京邮电大学教师邮箱完整数据 ({len(final)}条, {len(with_email_f)}个邮箱, {len(cc_final)}个学院)")
    print(f"南京邮电大学_有邮箱教师_{ts}.xlsx | 南京邮电大学教师邮箱-仅有邮箱 ({len(with_email_f)}条)")
    print(f"[/FILES]")


if __name__ == "__main__":
    main()
