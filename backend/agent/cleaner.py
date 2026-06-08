"""数据清洗模块 — 清洗爬取产出的教师邮箱数据。

清洗规则：
  1. 姓名验证 — 排除导航文字、碎片文本
  2. 邮箱验证 — 格式校验 + 排除学院公共邮箱
  3. 去重 — 按邮箱去重，保留信息最完整的记录
  4. 职称清洗 — 检测并清除被污染的职称字段
  5. 格式规范化 — 统一大小写、去除首尾空格

可被导入使用，也可作为脚本直接运行清洗已有文件。
"""

import re
import csv
import json
import logging
from pathlib import Path
from collections import defaultdict

from constants import is_valid_email_format, is_admin_or_public_email, EMAIL_RE as STD_EMAIL_RE

logger = logging.getLogger(__name__)

# 导航关键词 — 匹配到说明姓名是导航文字碎片
NAV_KEYWORDS = [
    "概况", "简介", "新闻", "通知", "公告", "招生", "培养", "就业",
    "学位", "学科", "科研", "学术", "党建", "工会", "校友", "捐赠",
    "师资", "教师", "教授", "博士", "硕士", "本科", "研究", "行政",
    "管理", "教职", "荣休", "访问", "系科", "教研", "诚聘", "联系",
    "欢迎", "首页", "返回", "更多", "详情", "查看", "下载", "友情",
    "版权", "网站", "地图", "登录", "邮箱",
]

# 学院级公共邮箱特征
ADMIN_EMAIL_PREFIXES = [
    "webmaster", "admin", "office", "info", "master", "root",
    "postmaster", "wxyxz", "xwcb", "bgs", "dangzheng", "yuanban",
    "dangban", "renshi", "jiaowu", "xuegong", "tuanwei", "yanjiusheng",
    "gysj", "gyyz", "glxb",  # 书记信箱、院长信箱、联系我们
]

# 姓名黑名单 — 精确匹配，这些肯定不是人名
NAME_BLACKLIST = {
    # 导航/页面UI文字
    "首页", "上一页", "下一页", "末页", "返回", "更多", "详情", "查看", "下载",
    "搜索", "登录", "注册", "退出", "提交", "重置", "保存", "取消", "关闭",
    "网站首页", "网站地图", "站点地图", "友情链接", "版权信息", "版权所有",
    "隐私政策", "法律声明", "网站声明", "访问统计", "技术支持", "网站管理",
    "English", "English Version", "中文版", "EN", "CN",

    # 学院子页标题
    "学院概", "师资队", "现任教", "管理架", "系科设", "教研机",
    "新闻公", "通知公", "各类公", "招生培", "本科生", "研究生", "社会培",
    "留学生", "教学科", "实验教", "图书资", "博士后", "访问学",
    "学院概况", "学院简介", "学院新闻", "通知公告", "学院通知",
    "学科建设", "学科介绍", "学科方向", "重点学科", "特色学科",
    "人才培养", "培养方案", "教学管理", "课程建设", "实践教学",
    "科学研究", "科研平台", "科研成果", "科研项目", "学术活动",
    "招生就业", "招生信息", "就业信息", "学生工作", "党团建设",
    "党建工作", "党群工作", "工会工作", "校友工作", "校友风采",
    "国际合作", "交流合作", "对外交流", "社会服务",
    "规章制度", "政策文件", "办事指南", "下载中心",
    "学院领导", "机构设置", "组织机构", "行政机构", "教辅机构",
    "专业设置", "专业介绍", "课程设置", "实验中心",
    "教师名录", "全体教师", "在职教师", "退休教师", "教师风采",
    "教师简介", "教师队伍", "教师信息", "教师列表",
    "客座教授", "兼职教授", "校外导师", "企业导师", "行业导师",
    "讲座教授", "名誉教授", "特聘教授", "讲座预告",
    # 职称级导航词（非人名，在导航页被误当教师姓名）
    "教授", "副教授", "讲师", "助教", "实验师",
    "高级实验师", "工程师", "高级工程师", "研究员", "副研究员",
    "助理研究员", "见习", "实习",
    "学术报告", "学术讲座", "科研团队", "创新团队",
    "教学团队", "精品课程", "一流课程", "课程大纲",

    # 常见非人名字段
    "南京大", "北京大", "清华大", "复旦大", "浙江大", "武汉大",
    "书记信箱", "院长信箱", "联系我们", "关于我们",
    "师德师", "党员风", "创先争", "十九大", "二十",
    "德育工", "心理健", "资助体", "学风建", "安全稳",
    "年度考", "职称评", "岗位聘", "考核办", "人事处",
    "教务处", "科研处", "研究生", "学生处", "财务处",
    "图书馆", "档案馆", "校医院", "后勤处", "保卫处",
    "实验室", "会议室", "报告厅", "体育馆", "运动场",
    "校历", "班车", "电话", "地址", "邮编", "邮箱",

    # 英文导航/UI
    "Home", "About", "Contact", "Faculty", "Staff", "Research",
    "Teaching", "Admissions", "News", "Events", "Calendar",
    "Students", "Alumni", "Giving", "Jobs", "Careers", "Links",
    "Sitemap", "Privacy", "Terms", "Help", "FAQ", "Support",
    "Login", "Register", "Search", "Site Map", "Campus Map",
    "Directory", "People", "Department", "Programs", "Degrees",
    "Courses", "Academics", "Administration", "Services",
    "Resources", "Downloads", "Gallery", "Video", "Media",
    "Quick Links", "Useful Links", "Related Links",

    # 碎片文字
    "您现在", "当前位", "您的位", "位置：", "您现",
    "共", "条记", "每页", "显示", "总访", "今日",
    "您是第", "已有", "人访", "更新", "发布", "日期",
    "时间", "作者", "来源", "点击", "次数", "人气",
    "字号", "大中", "小", "打印", "分享", "推荐",
    "相关", "标签", "分类", "归档", "热文",
    "公告", "公示", "通知", "动态", "快讯", "简报",

    # 页面底部/页眉
    "设为首页", "加入收藏", "网站导航", "帮助信息",
    "管理入口", "信息门户", "办公系统", "教务系统",
    "邮箱系统", "VPN", "WebVPN", "信息办",
    "建议", "意见", "反馈", "投诉", "咨询",
}

# 导航列表页 URL 模式 — 匹配到的页面即使有邮箱也不属于教师数据
NAV_LIST_URL_PATTERNS = [
    r'/list\.(htm|html|shtml)$',
    r'/index\.(htm|html|shtml)$',
    r'/main\.htm$',
    r'/news/', r'/notice/', r'/announcement/',
    r'/contact/', r'/links/', r'/map/',
]

# 学术邮箱域名白名单
ACADEMIC_DOMAIN_WHITELIST = {
    ".edu.cn", ".edu", ".ac.cn", ".ac.",
    # 部分高校使用 .org / .com / .net 但属于学术用途
    ".org.cn", ".org",
}


def is_admin_email(email: str) -> bool:
    """检测是否为学院公共邮箱。"""
    email_lower = email.lower().strip()
    for prefix in ADMIN_EMAIL_PREFIXES:
        if email_lower.startswith(prefix + "@"):
            return True
    return False


def is_nav_list_url(url: str) -> bool:
    """检测 URL 是否为导航列表页（而非教师详情页）。"""
    if not url:
        return False
    url_lower = url.lower().strip()
    for pattern in NAV_LIST_URL_PATTERNS:
        if re.search(pattern, url_lower):
            return True
    return False


def is_academic_domain(email: str) -> bool:
    """检测邮箱域名是否在教育/学术白名单内。"""
    if not email or "@" not in email:
        return True  # 空邮箱跳过检查
    domain = email.strip().lower().split("@")[-1]
    for allowed in ACADEMIC_DOMAIN_WHITELIST:
        if domain.endswith(allowed) or domain == allowed:
            return True
    return False


def is_valid_person_name(name: str) -> bool:
    """检测文本是否为合理的教师姓名（非导航文字）。"""
    name = name.strip()
    # 长度检查
    if len(name) < 2 or len(name) > 6:
        return False
    # 在黑名单中
    if name in NAME_BLACKLIST:
        return False
    # 含导航关键词
    if any(kw in name for kw in NAV_KEYWORDS):
        return False
    # 纯英文且很短（导航链接如"Home", "About"）
    if re.match(r"^[A-Za-z\s]{1,15}$", name):
        return False
    # 含数字或特殊符号
    if re.search(r"[0-9@#￥%&*()（）《》【】\[\]]", name):
        return False
    # 中文姓名：至少包含2个汉字
    if not re.search(r"[一-鿿]{2,}", name):
        return False
    return True


def is_valid_email_format(email: str) -> bool:
    """校验邮箱格式（统一从 constants 导入的正则）。"""
    return bool(STD_EMAIL_RE.match(email.strip()))


def is_clean_title(title: str) -> bool:
    """检测职称是否被污染。"""
    title = title.strip()
    if not title:
        return True  # 空值视为干净
    # 过长（项目经历或论文标题）
    if len(title) > 30:
        return False
    # 含可疑关键词（排除合法职称中的字）
    pollution_keywords = [
        "基金", "项目", "论文", "国家", "年度", "重点",
        "编号", "发表", "课题", "获奖", "专利", "出版",
        "通知", "公告", "公示", "招聘", "计划", "著",
    ]
    if any(kw in title for kw in pollution_keywords):
        return False
    return True


def clean_title(title: str) -> str:
    """清洗职称字段，返回干净的职称或空字符串。"""
    title = title.strip()
    if not title:
        return ""
    if is_clean_title(title):
        return title
    # 尝试从脏数据中提取真实的职称关键词
    valid_titles = [
        "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
        "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
        "长江学者", "杰青", "优青", "院长", "副院长", "系主任", "副主任",
        "所长", "副所长", "博士后", "高级实验师", "实验师",
    ]
    found = [t for t in valid_titles if t in title]
    if found:
        return max(found, key=len)  # 返回最长的匹配（更具体）
    return ""


def normalize_email(email: str) -> str:
    """规范化邮箱：转小写、去空格。"""
    return email.strip().lower()


def clean_records(records: list[dict]) -> list[dict]:
    """对记录列表执行全部清洗规则，返回干净的记录列表。

    清洗流程：
    1. 格式规范化
    2. 姓名验证（过滤导航文字）
    3. 邮箱格式验证
    4. 排除学院公共邮箱
    5. 域名白名单校验（仅保留教育/学术邮箱）
    6. 无邮箱 + 列表页 URL → 跳过
    7. 按邮箱去重（保留信息最全的）
    8. 职称清洗
    """
    stats = {
        "total": len(records),
        "bad_name": 0,
        "bad_email_format": 0,
        "admin_email": 0,
        "bad_domain": 0,
        "nav_list_no_email": 0,
        "deduped": 0,
        "bad_title": 0,
        "cleaned": 0,
    }

    # 第1步：格式规范化
    for r in records:
        r["name"] = r.get("name", "").strip()
        r["email"] = normalize_email(r.get("email", ""))
        r["title"] = r.get("title", "").strip()
        r["department"] = r.get("department", "").strip()
        r["url"] = r.get("url", "").strip()

    # 第2步：姓名验证
    before = len(records)
    records = [r for r in records if not r["name"] or is_valid_person_name(r["name"])]
    stats["bad_name"] = before - len(records)

    # 第3步：邮箱格式验证（空邮箱跳过不验证，保留为"无邮箱"）
    before_email = len(records)
    records = [r for r in records if not r["email"] or is_valid_email_format(r["email"])]
    stats["bad_email_format"] = before_email - len(records)

    # 第4步：排除学院公共邮箱
    before_admin = len(records)
    records = [r for r in records if not is_admin_email(r["email"])]
    stats["admin_email"] = before_admin - len(records)

    # 第5步：邮箱域名白名单校验（仅保留教育/学术域名邮箱，空邮箱跳过）
    before_domain = len(records)
    records = [r for r in records if not r["email"] or is_academic_domain(r["email"])]
    stats["bad_domain"] = before_domain - len(records)

    # 第6步：无邮箱 + 列表页 URL → 跳过（导航页不可能有教师数据）
    before_nav = len(records)
    records = [
        r for r in records
        if r["email"] or not is_nav_list_url(r.get("url", ""))
    ]
    stats["nav_list_no_email"] = before_nav - len(records)

    # 第5步：按邮箱去重（保留信息最全的 — 有姓名+职称优先）
    # 同时记录无邮箱的条目（空邮箱也保留，但只保留最新的一个）
    email_groups: dict[str, list[dict]] = defaultdict(list)
    no_email_entries = []  # 无邮箱的单独收集
    for r in records:
        if r["email"]:
            email_groups[r["email"]].append(r)
        else:
            no_email_entries.append(r)

    deduped = []
    for email, group in email_groups.items():
        # 按信息完整度排序：有姓名+职称 > 有姓名 > 其他
        def score(item):
            s = 0
            if item.get("name"):
                s += 2
            if item.get("title"):
                s += 1
            if item.get("department"):
                s += 1
            if item.get("url"):
                s += 1
            return s

        group.sort(key=score, reverse=True)
        deduped.append(group[0])

    # 无邮箱条目去重：按姓名去重
    seen_no_email = set()
    for r in no_email_entries:
        name = r.get("name", "")
        if name and name not in seen_no_email:
            seen_no_email.add(name)
            deduped.append(r)

    stats["deduped"] = len(records) - len(deduped)
    records = deduped

    # 第8步：职称清洗
    for r in records:
        if r["title"] and not is_clean_title(r["title"]):
            stats["bad_title"] += 1
            r["title"] = clean_title(r["title"])

    stats["cleaned"] = len(records)
    logger.info(f"数据清洗完成: {stats}")
    return records


def clean_csv(input_path: str, output_path: str | None = None) -> list[dict]:
    """清洗 CSV 文件，返回清洗后的记录列表。"""
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"文件不存在: {input_path}")

    records = []
    with open(input_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({
                "name": row.get("姓名", ""),
                "email": row.get("邮箱", ""),
                "department": row.get("学院", ""),
                "title": row.get("职称", ""),
                "url": row.get("主页链接", ""),
            })

    cleaned = clean_records(records)

    if output_path:
        output_path = Path(output_path)
        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
            for i, r in enumerate(cleaned, 1):
                writer.writerow([i, r["name"], r["email"], r["department"], r["title"], r["url"]])
        logger.info(f"清洗后 CSV 已保存: {output_path}")

    return cleaned


def clean_xlsx(input_path: str, output_path: str | None = None) -> list[dict]:
    """清洗 XLSX 文件，返回清洗后的记录列表。"""
    import openpyxl

    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"文件不存在: {input_path}")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        if h == "姓名":
            col_map["name"] = i
        elif h == "邮箱":
            col_map["email"] = i
        elif "学院" in str(h):
            col_map["department"] = i
        elif h == "职称":
            col_map["title"] = i
        elif "链接" in str(h) or "url" in str(h).lower():
            col_map["url"] = i

    records = []
    for row_idx in range(2, ws.max_row + 1):
        r = {}
        for field, col_idx in col_map.items():
            val = ws.cell(row_idx, col_idx).value
            r[field] = str(val).strip() if val else ""
        records.append(r)

    cleaned = clean_records(records)

    if output_path:
        from agent.exporter import export_xlsx
        export_xlsx(cleaned, Path(output_path).stem)

    return cleaned


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)

    if len(sys.argv) < 2:
        print("用法: python cleaner.py <input.csv|input.xlsx> [output.csv|output.xlsx]")
        print("示例: python cleaner.py outputs/南京大学_20260526.csv outputs/南京大学_cleaned.csv")
        sys.exit(1)

    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None

    if inp.endswith(".csv"):
        result = clean_csv(inp, out)
    elif inp.endswith(".xlsx"):
        result = clean_xlsx(inp, out)
    else:
        print("错误: 仅支持 .csv 或 .xlsx 文件")
        sys.exit(1)

    print(f"\n清洗完成: {len(result)} 条有效记录")
    if result:
        print(f"\n前10条预览:")
        for i, r in enumerate(result[:10], 1):
            print(f"  {i}. {r['name']} | {r['email']} | {r['department']} | {r['title']}")
