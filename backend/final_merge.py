"""合并所有数据源，生成最终版南京大学教师邮箱数据。

数据源：
1. 原始合并版CSV (894条) - 基础数据
2. 匡亚明学院补全 - Playwright爬取
3. 教育研究院补全 - Playwright爬取
4. 电子科学全院名录 (121名，无邮箱)
5. 针对剩余学院的快速补爬
"""

import asyncio
import csv
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import OrderedDict

OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

ADMIN_EMAIL_PREFIXES = [
    "wxyxz", "xwcb", "bgs", "dangzheng", "office", "yuanban",
    "webmaster", "admin", "info", "master", "root", "postmaster",
    "gcglxydw", "njugcglxy",
]


def is_personal_email(email: str) -> bool:
    if not email or not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
        return False
    email_lower = email.lower()
    for prefix in ADMIN_EMAIL_PREFIXES:
        if email_lower.startswith(prefix + "@"):
            return False
    return True


def load_csv(filepath: Path) -> list[dict]:
    """加载CSV，返回记录列表。"""
    if not filepath.exists():
        print(f"  ⚠️ 文件不存在: {filepath}")
        return []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


def save_csv(records: list[dict], filepath: Path):
    """保存CSV。"""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(records, 1):
            writer.writerow([
                i,
                r.get("name", r.get("姓名", "")),
                r.get("email", r.get("邮箱", "")),
                r.get("department", r.get("学院", "")),
                r.get("title", r.get("职称", "")),
                r.get("url", r.get("主页链接", "")),
            ])


def save_xlsx(records: list[dict], filepath: Path):
    """保存XLSX。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "教师邮箱"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="center")

    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(records, 1):
        values = [
            i,
            r.get("name", r.get("姓名", "")),
            r.get("email", r.get("邮箱", "")),
            r.get("department", r.get("学院", "")),
            r.get("title", r.get("职称", "")),
            r.get("url", r.get("主页链接", "")),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 55

    wb.save(filepath)


def normalize_record(r: dict) -> dict:
    """统一字段名。"""
    return {
        "name": (r.get("name") or r.get("姓名") or "").strip(),
        "email": (r.get("email") or r.get("邮箱") or "").strip(),
        "department": (r.get("department") or r.get("学院") or "").strip(),
        "title": (r.get("title") or r.get("职称") or "").strip(),
        "url": (r.get("url") or r.get("主页链接") or "").strip(),
    }


def clean_title(title: str) -> str:
    """清理职称字段中的脏数据。"""
    if not title:
        return ""
    # 去除过长的描述性文本
    if len(title) > 30:
        # 尝试从中提取真正的职称
        keywords = ["教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
                    "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
                    "博士后", "助理教授"]
        found = []
        for kw in keywords:
            if kw in title:
                found.append(kw)
        if found:
            return "、".join(found)
        return title[:30] + "..."
    return title


def merge_all():
    """主合并逻辑。"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    print(f"🔄 数据合并开始 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 1. 加载原始合并版
    base_file = OUTPUT_DIR / "南京大学_教师邮箱_合并版_20260526_163254.csv"
    base_records = load_csv(base_file)
    print(f"📥 原始数据: {len(base_records)} 条")

    # 2. 加载补全数据
    new_sources = {}

    # 加载针对性爬取结果（targeted_*）
    for csv_file in sorted(OUTPUT_DIR.glob("*targeted*.csv")):
        records = load_csv(csv_file)
        if records:
            new_sources[csv_file.name] = records
            print(f"📥 {csv_file.name}: {len(records)} 条")

    # 加载补全14院系
    for csv_file in OUTPUT_DIR.glob("*补全14*"):
        if csv_file.name.endswith(".csv"):
            records = load_csv(csv_file)
            if records:
                new_sources[csv_file.stem] = records
                print(f"📥 {csv_file.name}: {len(records)} 条")

    # 加载电子科学名单
    es_file = OUTPUT_DIR / "电子科学_补全.csv"
    es_records = load_csv(es_file)
    if es_records:
        new_sources["电子科学_名单"] = es_records
        print(f"📥 电子科学_补全.csv: {len(es_records)} 条")

    # 3. 合并所有记录
    all_records = {}
    # 用 (姓名, 学院) 作为去重键，但优先保留有邮箱的记录

    # 先加载所有记录
    for r in base_records:
        r = normalize_record(r)
        key = (r["name"], r["department"])
        if key not in all_records:
            all_records[key] = r
        else:
            # 如果新记录已有邮箱而旧记录没有，替换
            existing = all_records[key]
            if r["email"] and not existing["email"]:
                all_records[key] = r

    for source_name, records in new_sources.items():
        for r in records:
            r = normalize_record(r)
            key = (r["name"], r["department"])
            if key not in all_records:
                all_records[key] = r
            else:
                existing = all_records[key]
                # 保留有邮箱的记录
                if r["email"] and not existing["email"]:
                    all_records[key] = r
                # 保留有职称的记录
                if r["title"] and not existing["title"]:
                    all_records[key]["title"] = r["title"]
                # 保留有链接的记录
                if r["url"] and not existing["url"]:
                    all_records[key]["url"] = r["url"]

    # 4. 数据清洗
    cleaned = []
    for key, r in all_records.items():
        name = r["name"].strip()

        # 跳过明显的脏数据
        if not name:
            continue
        if len(name) < 2:
            continue
        # 导航/脏数据关键词列表
        junk_names = {
            "首页", "返回", "关闭此页", "确定", "取消", "更多", "当前位置", "当前位置：",
            "南大主页", "学校主页", "导航", "搜索", "电子邮箱", "学院概况",
            "师资队伍", "教师名录", "登录", "English", "EN",
            "尾页", "下一页", "上一页", "研究方向", "学术交流", "学科介绍",
            "学院简介", "师资力量", "院士", "学校邮箱", "系科设置",
            "院长致辞", "现任领导", "机构设置", "议事机构", "人才培养",
            "科学研究", "党群工作", "学生工作", "合作交流", "诚聘英才",
            "科研机构", "科研动态", "科研成果", "学术活动", "学术刊物",
            "实验技术", "专职科研", "行政管理",
        }
        if name in junk_names:
            continue
        # 纯职称词（被当成人名）
        if name in ["教授", "副教授", "助理教授", "讲师", "研究员", "副研究员", "博导", "硕导"]:
            continue
        if re.match(r'^\d+$', name):
            continue
        if any(kw in name for kw in ["通知", "公告", "新闻", "课题组", "监督", "举报", "地址", "邮编", "电话"]):
            continue
        if re.match(r'^[A-Za-z\s]+$', name) and len(name) > 20:
            continue

        # 邮箱验证
        email = r["email"].strip()
        if email and not is_personal_email(email):
            email = ""  # 清除非个人邮箱

        # 职称清理
        title = clean_title(r["title"])

        cleaned.append({
            "name": name,
            "email": email,
            "department": r["department"],
            "title": title,
            "url": r["url"],
        })

    # 二次去重（按邮箱）
    seen_emails = set()
    final = []
    for r in cleaned:
        if r["email"] and r["email"] in seen_emails:
            continue
        if r["email"]:
            seen_emails.add(r["email"])
        final.append(r)

    # 排序：按学院→姓名
    final.sort(key=lambda x: (x["department"], x["name"]))

    print(f"\n📊 合并结果: {len(final)} 条（清洗后）")

    # 统计各学院数量
    from collections import Counter
    dept_counts = Counter(r["department"] for r in final)
    print("\n=== 各学院教师数量 ===")
    for dept, count in dept_counts.most_common():
        has_email = sum(1 for r in final if r["department"] == dept and r["email"])
        print(f"  {dept}: {count}人 (有邮箱: {has_email})")

    # 保存最终文件
    csv_path = OUTPUT_DIR / f"南京大学_教师邮箱_最终版_{ts}.csv"
    xlsx_path = OUTPUT_DIR / f"南京大学_教师邮箱_最终版_{ts}.xlsx"

    save_csv(final, csv_path)
    print(f"\n💾 CSV: {csv_path}")

    try:
        save_xlsx(final, xlsx_path)
        print(f"💾 XLSX: {xlsx_path}")
    except Exception as e:
        print(f"⚠️ XLSX导出失败: {e}")

    # 输出关注8学院的最终数量
    target_depts = ["艺术学院", "能源与资源学院", "工程管理学院", "匡亚明学院",
                    "化学化工学院", "电子科学与工程学院", "文学院", "教育研究院"]
    print("\n=== 重点关注8学院变化 ===")
    print(f"{'学院':<20} {'原始':<6} {'最终':<6} {'有邮箱':<6} {'变化':<6}")
    original_counts = {
        "艺术学院": 4, "能源与资源学院": 4, "工程管理学院": 4,
        "匡亚明学院": 4, "化学化工学院": 4, "电子科学与工程学院": 9,
        "文学院": 9, "教育研究院": 10,
    }
    for dept in target_depts:
        final_count = dept_counts.get(dept, 0)
        has_email = sum(1 for r in final if r["department"] == dept and r["email"])
        orig = original_counts.get(dept, 0)
        change = f"+{final_count - orig}" if final_count >= orig else str(final_count - orig)
        print(f"{dept:<20} {orig:<6} {final_count:<6} {has_email:<6} {change:<6}")

    print(f"\n✅ 完成！最终数据: {len(final)} 条")
    print(f"   有邮箱: {sum(1 for r in final if r['email'])} 条")


if __name__ == "__main__":
    merge_all()
