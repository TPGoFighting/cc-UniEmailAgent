"""合并东南大学新旧数据，清洗并导出最终版。"""
import csv
import json
import re
import glob
from datetime import datetime
from pathlib import Path
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TASK_ID = "bda95480-ec96-4bd7-bc18-05f797e28dd4"
OUTPUT_DIR = Path(__file__).parent / "outputs" / TASK_ID

NAV_WORDS = {
    '首页','概况','新闻','通知','公告','招生','培养','就业','学位','学科',
    '科研','学术','党建','工会','校友','捐赠','图书馆','校园','地图','网站',
    '登录','邮箱','联系我们','欢迎','返回','更多','详情','查看','下载',
    '学院','大学','管理','后台','English','日本語',
    '人才引进','人才招聘','院长书记','信箱','相关链接','联系方式',
    '学校首页','学校主页','收藏本站','旧版入口','暑期学校','平湖芳草',
    '下载专区','捐赠通道','院长邮箱','院内文档','标识系统',
    '院系设置','教师教学','技术转移','海外教育','仪器设备','化工时刊',
    '尾页','网站首页','招生信息','教师登录','现任领导','历任领导','办公电话',
    '院长寄语','组织机构','学科建设','历史沿革','学院简介','学院介绍',
    '组织框架','系所设置','学科组织','本院概况','本院简介','学院简况',
    '学院概述','学院架构','快捷入口','学术论文','专利成果','获奖成果',
    '课程改革','牵头学科','学位管理','出国交流','答辩公示',
    '本科生','研究生','学生工作','党群工作','人才培养','科学研究',
    '人才引进','校友天地','合作交流','诚聘英才',
    '拔尖基地','教学管理','本科生培','研究生培','规章制度',
    '学生自我','教职工','教师教学','发展中心','教师查询',
    '教师风采','专任教师','客座教授','教师简介','兼职教授',
    '离退休','荣休','知名专家','知名学者','全体教师','硕博导师',
    '各系名单','国家高层次','人才','参观预约','博士后','博士','硕士',
    '助理','秘书','处长','科长','主任','书记','馆长',
    '师资修改','管理入口','个人中心','师资维护','师资概况','师资队伍',
    '师资力量','教师名录','教授风采','杰出人才','杰出教师',
    '声影机械','行政服务','资料下载','常用下载','培养动态',
    '招生动态','行政机构','系所介绍','学院治理','学院新闻','学术信息',
    '重要通知','教务信息','图片新闻','党群组织','关工委工作',
    '学术委员会','教学委员会','学位评定',
    '本科生培养','研究生培养','重点科研平台','广纳英才',
    '两院院士','人才职称','博导','硕导',
    '实验中心','院机关','硕士导师','博士导师','博士生导师','硕士生导师',
    '在站博士后','专职科研岗','职业规划导师','学系','附属医院',
    '研究进展','科研平台','重点实验室','留学生培养','本科生事务',
    '研究生事务','理论学习指南','学院宣传片','学系部门名单',
    '实验教学中心','医学教育发展中心','全员合影','参观活动',
    '会议活动','院徽','百年党史','教师活动照片','大合唱照片',
    '东大土木历史人物','教职工名录','建筑工程系','建设与房地产系',
    '工程力学系','桥隧与地下工程系','市政工程系','智慧建造与运维系',
    '全职博士后','学术团体任职','退休教职工','本科生教育',
    '本科生教学','研究生培养','院庆专栏','国家级人才','科技成果',
    '学术年专栏','电子信息','院长信箱','书记院长',
    '组织架构','发展历程','学科设置','专业介绍','基地建设',
    '学术动态','科研概况','科研通知','科研基地','党群工作',
    '学术交流','科研机构','培养动态','招生动态',
    '科研成果','科研项目','科研奖励','科研团队',
    '学院刊物','特色专业','重点专业','高端培训','历史人物',
    '大事记','英文版','登录','账号','密码','注册','忘记密码',
    '首页','上一页','下一页','尾页','第','页','共',
    '校友会专题','同济大学','东南大学','江苏省特色','江苏省重点',
    '党史上的','能源热转换','江苏省能源','学院主页','回到旧版',
    '国家信息','互联网治理','师资概览','江苏省应用',
    '江苏省高等','中国数学','中国科学','数字医学','生物材料',
    '儿童发展','教学实验','生物科学','校友之家',
    '学科发展','日本語','具体','查看','全部','点击','链接',
    '学校主页','按系所查',
}

ADMIN_EMAILS = {
    "webmaster", "admin", "office", "info", "master", "root", "postmaster",
    "wxyxz", "xwcb", "bgs", "dangzheng", "yuanban", "radio", "seuradiojob",
    "dzxy", "seuem", "seutc_official", "ysxy", "slst", "deanoffice_seuarch",
    "jupiter", "seueexb", "cyber", "zscq", "seuyzb", "gsyz", "yauc",
    "jy", "zhaoban", "sunway", "dean_seuem", "waxy", "stephenxu",
    "jiashun", "clchu", "phyljp", "gongfeng", "ddfzyjy", "songjiangang",
    "wgu", "ygwang", "zhangli", "weixy", "ic-seu", "math",
    "ykxy", "hytong", "101006856", "101011416",
}


def is_bad_entry(name: str, email: str) -> bool:
    """检查是否是非教师条目。"""
    # 检查名字是否是导航词
    for w in NAV_WORDS:
        if name == w or w in name:
            return True
    # 检查名字长度
    if len(name) < 2 or len(name) > 4:
        return True
    # 名字必须是纯中文
    if not re.match(r'^[一-鿿·]+$', name):
        return True
    return False


def clean_name(name: str) -> str:
    """清理名字：去掉教授/院士等职称后缀。"""
    # 去掉括号中的内容
    name = re.sub(r'[（(][^)）]*[)）]', '', name).strip()
    # 去掉职称后缀
    for suffix in ['教授', '院士', '研究员', '高级工程师',
                   '副教授', '副研究员', '助理教授', '助理研究员',
                   '讲师', '工程师', '博导', '硕导', '博士', '硕士',
                   '院长', '主任', '书记', '所长', '处长', '科长',
                   '秘书', '助理', '助手', '教']:
        if name.endswith(suffix) and len(name.replace(suffix, '')) >= 2:
            name = name.replace(suffix, '')
            break
    return name.strip()


def dedup_teachers(teachers: list[dict]) -> list[dict]:
    """按邮箱去重，同一邮箱保留 title 更详细者。"""
    email_map = {}
    for t in teachers:
        email = t.get("email", "").lower().strip()
        if not email:
            continue
        if email not in email_map:
            email_map[email] = t
        else:
            # 保留 title 更详细的
            existing = email_map[email]
            if len(t.get("title", "")) > len(existing.get("title", "")):
                email_map[email] = t
            # 保留 name 更准确的（不含教授等词的）
            if any(w in existing["name"] for w in ['教授','院士','教']):
                email_map[email] = t

    return list(email_map.values())


def main():
    all_teachers = []

    # === 1. 加载旧数据 (CSV) ===
    old_csv = list(OUTPUT_DIR.glob("东南大学_教师邮箱_20260*.csv"))
    if old_csv:
        latest = max(old_csv, key=lambda x: x.stat().st_mtime)
        print(f"加载旧数据: {latest.name}")
        with open(latest, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("姓名", "").strip()
                email = row.get("邮箱", "").strip()
                dept = row.get("学院", "").strip()
                title = row.get("职称", "").strip()
                url = row.get("主页链接", "").strip()

                if is_bad_entry(name, email):
                    continue
                # 跳过admin邮箱
                low = email.lower()
                if any(p in low for p in ADMIN_EMAILS):
                    continue

                name = clean_name(name)
                if len(name) < 2:
                    continue

                all_teachers.append({
                    "name": name,
                    "email": email,
                    "department": dept,
                    "title": title,
                    "url": url,
                })

    print(f"  旧数据清洗后: {len(all_teachers)} 位")

    # === 2. 加载修复数据 ===
    fix_files = sorted(OUTPUT_DIR.glob("seu_fix_*.json"), key=lambda x: x.stat().st_mtime)
    if fix_files:
        latest_fix = fix_files[-1]
        print(f"加载修复数据: {latest_fix.name}")
        data = json.loads(latest_fix.read_text(encoding="utf-8"))
        for t in data.get("teachers", []):
            name = clean_name(t.get("name", "").strip())
            email = t.get("email", "").strip().lower()
            dept = t.get("department", "").strip()
            title = t.get("title", "").strip()
            url = t.get("url", "").strip()

            if len(name) < 2:
                continue
            # 修复名字中的"教"尾
            if name.endswith("教") and len(name) > 2:
                name = name[:-1]

            all_teachers.append({
                "name": name,
                "email": email,
                "department": dept,
                "title": title,
                "url": url,
            })
        print(f"  修复数据: {len(data.get('teachers', []))} 位")

    # === 3. 去重 ===
    unique = dedup_teachers(all_teachers)
    print(f"  去重后: {len(unique)} 位")

    # === 4. 按学院+姓名排序 ===
    unique.sort(key=lambda x: (x["department"], x["name"]))

    # === 5. 统计 ===
    dept_counts = Counter(t["department"] for t in unique)
    print(f"\n各学院统计 (总计 {len(unique)} 位):")
    for dept, cnt in dept_counts.most_common():
        print(f"  {dept}: {cnt} 位")

    # === 6. 导出 ===
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"东南大学_教师邮箱_最终版_{ts}.xlsx"
    csv_path = OUTPUT_DIR / f"东南大学_教师邮箱_最终版_{ts}.csv"

    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]

    # --- XLSX ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "东南大学教师邮箱"

    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cf = Font(name="微软雅黑", size=10)
    ca = Alignment(vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = border

    for i, r in enumerate(unique, 1):
        for col, v in enumerate([i, r["name"], r["email"], r["department"],
                                  r["title"], r["url"]], 1):
            cell = ws.cell(row=i+1, column=col, value=v)
            cell.font = cf; cell.alignment = ca; cell.border = border

    for col, w in zip("ABCDEF", [8, 16, 38, 22, 16, 55]):
        ws.column_dimensions[col].width = w

    ws.auto_filter.ref = ws.dimensions
    wb.save(xlsx_path)
    print(f"\nXLSX: {xlsx_path}")

    # --- CSV ---
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i, r in enumerate(unique, 1):
            w.writerow([i, r["name"], r["email"], r["department"],
                       r["title"], r["url"]])
    print(f"CSV: {csv_path}")

    # --- JSON ---
    final_json = OUTPUT_DIR / "seu_final_merged.json"
    final_json.write_text(json.dumps({
        "total": len(unique),
        "xlsx": str(xlsx_path),
        "csv": str(csv_path),
        "dept_counts": dict(dept_counts),
        "teachers": unique,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n[FILES]")
    print(f"{xlsx_path.name}|东南大学教师邮箱XLSX（最终版）")
    print(f"{csv_path.name}|东南大学教师邮箱CSV（最终版）")
    print(f"[/FILES]")


if __name__ == "__main__":
    main()
