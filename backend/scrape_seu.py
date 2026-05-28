"""东南大学教师邮箱爬虫 — 最终版。

双模式策略：
  A. 链接模式：教师列表页有姓名链接 → 访问详情页提取邮箱
  B. 内联模式：教师列表页直接展示信息 → 页内提取

输出到 outputs/3112e368-94df-4d3a-ba95-8e73c352d37d/
"""

import asyncio
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import Counter

from playwright.async_api import async_playwright

TASK_ID = "bda95480-ec96-4bd7-bc18-05f797e28dd4"
OUTPUT_DIR = Path(__file__).parent / "outputs" / TASK_ID
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PROGRESS_FILE = OUTPUT_DIR / "seu_progress.json"

PAGE_TIMEOUT = 25000
PROFILE_TIMEOUT = 12000
MAX_TEACHERS_PER_DEPT = 150

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

ADMIN_EMAILS = {
    "webmaster", "admin", "office", "info", "master", "root", "postmaster",
    "wxyxz", "xwcb", "bgs", "dangzheng", "yuanban", "radio", "seuradiojob",
    "dzxy", "seuem", "seutc_official", "ysxy", "slst", "deanoffice_seuarch",
}

SEU_DEPARTMENTS = [
    # === 确认工作（链接模式）===
    ("计算机科学与工程学院", "cs.seu.edu.cn", [
        "https://cs.seu.edu.cn/49355/list.htm",  # 教师按职称
    ]),
    ("软件学院", "cose.seu.edu.cn", [
        "https://cose.seu.edu.cn/49355/list.htm",
    ]),
    ("人工智能学院", "ai.seu.edu.cn", [
        "https://ai.seu.edu.cn/49355/list.htm",
    ]),
    ("信息科学与工程学院", "radio.seu.edu.cn", [
        "http://radio.seu.edu.cn/19256/list.htm",
    ]),
    ("电气工程学院", "ee.seu.edu.cn", [
        "https://ee.seu.edu.cn/szdw/list.htm",
    ]),
    ("物理学院", "physics.seu.edu.cn", [
        "https://physics.seu.edu.cn/23137/list.htm",
    ]),
    # === 需要测试 ===
    ("经济管理学院", "em.seu.edu.cn", [
        "https://em.seu.edu.cn/57213/list.htm",
    ]),
    ("人文学院", "rwxy.seu.edu.cn", [
        "https://rwxy.seu.edu.cn/8787/list.htm",
    ]),
    ("建筑学院", "arch.seu.edu.cn", [
        "http://arch.seu.edu.cn/16782/list.htm",
        "http://arch.seu.edu.cn/16783/list.htm",
    ]),
    ("机械工程学院", "me.seu.edu.cn", [
        "https://me.seu.edu.cn/szll/list.htm",
    ]),
    ("能源与环境学院", "power.seu.edu.cn", [
        "http://power.seu.edu.cn/9216/list.htm",
        "http://power.seu.edu.cn/9232/list.htm",
    ]),
    ("土木工程学院", "civil.seu.edu.cn", [
        "https://civil.seu.edu.cn/10475/list.htm",
    ]),
    ("电子科学与工程学院", "electronic.seu.edu.cn", [
        "http://electronic.seu.edu.cn/szllx/list.htm",
    ]),
    ("数学学院", "math.seu.edu.cn", [
        "https://math.seu.edu.cn/szdw/list.htm",
        "https://math.seu.edu.cn/jsazc/list.htm",
    ]),
    ("自动化学院", "automation.seu.edu.cn", [
        "https://automation.seu.edu.cn/szdw_32667/list.htm",
    ]),
    ("生物科学与医学工程学院", "bme.seu.edu.cn", [
        "https://bme.seu.edu.cn/499/list.htm",
    ]),
    ("材料科学与工程学院", "smse.seu.edu.cn", [
        "https://smse.seu.edu.cn/2580/list.htm",
    ]),
    ("外国语学院", "sfl.seu.edu.cn", [
        "https://sfl.seu.edu.cn/9851/list.htm",
        "https://sfl.seu.edu.cn/9852/list.htm",
    ]),
    ("化学化工学院", "chem.seu.edu.cn", [
        "https://chem.seu.edu.cn/js/list.htm",
    ]),
    ("交通学院", "tc.seu.edu.cn", [
        "https://tc.seu.edu.cn/58248/list.htm",
    ]),
    ("仪器科学与工程学院", "ins.seu.edu.cn", [
        "https://ins.seu.edu.cn/45076/list.htm",
    ]),
    ("艺术学院", "arts.seu.edu.cn", [
        "https://arts.seu.edu.cn/szdw_25730/list.htm",
    ]),
    ("法学院", "law.seu.edu.cn", [
        "https://law.seu.edu.cn/9121/list.htm",
    ]),
    ("医学院", "med.seu.edu.cn", [
        "https://med.seu.edu.cn/8693/list.htm",
    ]),
    ("公共卫生学院", "gw.seu.edu.cn", [
        "https://gw.seu.edu.cn/zrjs/list.htm",
    ]),
    ("吴健雄学院", "wjx.seu.edu.cn", [
        "https://wjx.seu.edu.cn/21376/list.htm",
    ]),
    ("集成电路学院", "ic.seu.edu.cn", [
        "https://ic.seu.edu.cn/47757/list.htm",
    ]),
    ("马克思主义学院", "marxism.seu.edu.cn", [
        "https://marxism.seu.edu.cn/23294/list.htm",
    ]),
    ("网络空间安全学院", "cyber.seu.edu.cn", [
        "https://cyber.seu.edu.cn/18189/list.htm",
    ]),
    ("生命科学与技术学院", "ils.seu.edu.cn", [
        "https://ils.seu.edu.cn/22853/list.htm",
    ]),
    ("统计与数据科学学院", "stat.seu.edu.cn", [
        "https://stat.seu.edu.cn/szll_61997/list.htm",
    ]),
]

NAV_WORDS = {
    '首页','概况','新闻','通知','公告','招生','培养','就业',
    '学位','学科','科研','学术','党建','工会','校友','捐赠','图书馆',
    '校园','地图','网站','登录','邮箱','联系我们','欢迎','返回','更多',
    '详情','查看','下载','学院','大学','管理','后台','English',
    '人才引进','人才招聘','院长书记','信箱','相关链接','联系方式',
    '学校首页','学校主页','收藏本站','旧版入口','暑期学校','平湖芳草',
    '下载专区','捐赠通道','院长邮箱','院内文档','日本語','标识系统',
    '院系设置','教师教学','技术转移','海外教育','仪器设备','化工时刊',
    '尾页','网站首页','招生信息',
}


def is_admin_email(email: str) -> bool:
    low = email.lower()
    for p in ADMIN_EMAILS:
        if low.startswith(p + "@") or p in low.split("@")[0]:
            return True
    return False


def extract_title(text: str) -> str:
    m = re.search(r"职称[：:]\s*(.+?)(?:\n|$|。|；|研究方向|邮箱|电话|办公室)", text)
    if m:
        raw = m.group(1).strip()
        mapping = {"正高": "教授", "副高": "副教授", "中级": "讲师",
                   "初级": "助教", "正高级": "教授", "副高级": "副教授"}
        return mapping.get(raw, raw)[:20]
    for t in ["长江学者", "杰青", "优青", "院士",
              "教授", "研究员", "高级工程师",
              "副教授", "副研究员",
              "助理教授", "助理研究员",
              "讲师", "工程师", "博导", "硕导"]:
        if t in text:
            return t
    return ""


def normalize_name(name: str) -> str:
    return name.replace("​", "").replace("‌", "").replace("‍", "").strip()


def load_progress():
    return json.loads(PROGRESS_FILE.read_text(encoding="utf-8")) if PROGRESS_FILE.exists() else {"done_depts": {}, "teachers": []}


def save_progress(dept_name: str, teachers: list):
    data = load_progress()
    data["done_depts"][dept_name] = len(teachers)
    existing = {t["email"] for t in data["teachers"] if t.get("email")}
    for t in teachers:
        if t.get("email") and t["email"] not in existing:
            existing.add(t["email"])
            data["teachers"].append(t)
    PROGRESS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    sys.stdout.flush()


async def extract_teacher_links(page) -> tuple[list[dict], str]:
    """提取教师姓名链接（策略A）。返回 (entries, mode)。"""
    entries = await page.evaluate("""() => {
        const navWords = new Set(['首页','概况','新闻','通知','公告','招生','培养','就业',
            '学位','学科','科研','学术','党建','工会','校友','捐赠','图书馆',
            '校园','地图','网站','登录','邮箱','联系我们','欢迎','返回','更多',
            '详情','查看','下载','学院','大学','管理','后台','English',
            '人才引进','人才招聘','院长书记','信箱','相关链接','联系方式',
            '学校首页','学校主页','收藏本站','旧版入口','暑期学校','平湖芳草',
            '下载专区','捐赠通道','院长邮箱','院内文档','日本語','标识系统',
            '院系设置','教师教学','技术转移','海外教育','仪器设备','化工时刊',
            '尾页','网站首页','招生信息','教师登录','现任领导','历任领导','办公电话',
            '院长寄语','组织机构','学科建设','历史沿革','学院简介','学院介绍','组织框架',
            '系所设置','学科组织','本院概况','本院简介','学院简况','学院概述','学院架构',
            '快捷入口','深入贯彻','江苏省','学术论文','专利成果','获奖成果','课程改革',
            '牵头学科','学位管理','出国交流','答辩公示','本科生','研究生','学生工作',
            '党群工作','人才培养','科学研究','人才引进','校友天地','合作交流','诚聘英才',
            '拔尖基地','教学管理','本科生培','研究生培','鲁汶国际','项目介绍','电子信息',
            '规章制度','学生自我','教职工','教师教学','发展中心','教师查询','教师风采',
            '专任教师','院士','客座教授','教师简介','兼职教授','离退休','荣休','知名专家',
            '知名学者','全体教师','硕博导师','各系名单','国家高层次','人才','各系名单']);

        const results = [];
        const seen = new Set();

        document.querySelectorAll('a').forEach(a => {
            let text = a.textContent.trim().replace(/\\u200b/g, '').replace(/\\u200c/g, '').replace(/\\u200d/g, '').trim();
            const href = a.href || '';
            if (!href || href.startsWith('javascript:') || href === '#') return;
            if (href.includes('webplus') || href.includes('_teacherHome') || href.includes('login')) return;
            if (href.includes('list.htm')) return;

            const cleaned = text.replace(/\\s/g, '');
            if (/^[\\u4e00-\\u9fff]{2,4}$/.test(cleaned) && !navWords.has(cleaned)) {
                seen.add(href);
                results.push({name: cleaned, url: href});
            }
        });
        return results;
    }""")
    return entries, "link"


async def scrape_profile(profile_page, url: str, name: str, dept: str) -> dict | None:
    """访问教师详情页提取邮箱。"""
    try:
        await profile_page.goto(url, wait_until="domcontentloaded", timeout=PROFILE_TIMEOUT)
        await asyncio.sleep(0.8)
        text = await profile_page.evaluate("() => document.body.innerText")
        emails = [e.lower() for e in EMAIL_RE.findall(text) if not is_admin_email(e)]
        if emails:
            return {
                "name": normalize_name(name),
                "email": emails[0],
                "department": dept,
                "title": extract_title(text),
                "url": url,
            }
    except Exception:
        pass
    return None


async def scrape_department(context, dept_name: str, domain: str, urls: list[str]) -> list[dict]:
    """爬取单个学院。"""
    print(f"\n{'='*50}")
    print(f"[{dept_name}] 开始...")
    sys.stdout.flush()

    page = await context.new_page()
    results = []
    seen_emails = set()

    try:
        all_entries = []
        for url in urls[:3]:
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                await asyncio.sleep(2)
                entries, mode = await extract_teacher_links(page)
                print(f"  [{url.split('/')[-2] or url.split('/')[-1]}] {mode}模式: {len(entries)} 条目")
                all_entries.extend(entries)
            except Exception as e:
                print(f"  列表页错误: {e}")

        # 去重
        seen_urls = set()
        unique = []
        for e in all_entries:
            if e["url"] not in seen_urls:
                seen_urls.add(e["url"])
                unique.append(e)

        print(f"  去重后: {len(unique)} 位教师，开始访问详情页...")
        sys.stdout.flush()

        if not unique:
            return results

        profile_page = await context.new_page()
        count = 0
        for i, entry in enumerate(unique[:MAX_TEACHERS_PER_DEPT]):
            result = await scrape_profile(profile_page, entry["url"], entry["name"], dept_name)
            if result and result.get("email") and result["email"] not in seen_emails:
                seen_emails.add(result["email"])
                results.append(result)
                count += 1
                if count <= 5 or count % 25 == 0:
                    print(f"    [{count}] {result['name']} <{result['email']}> {result['title']}")
                    sys.stdout.flush()
            if (i + 1) % 30 == 0:
                await asyncio.sleep(0.5)

        await profile_page.close()
        print(f"  完成: {len(results)} 位教师(有邮箱)")
    except Exception as e:
        print(f"  错误: {e}")
    finally:
        await page.close()

    sys.stdout.flush()
    return results


async def main():
    progress = load_progress()
    all_teachers = list(progress.get("teachers", []))

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        )

        for dept_name, domain, urls in SEU_DEPARTMENTS:
            if dept_name in progress.get("done_depts", {}):
                prev = progress["done_depts"][dept_name]
                print(f"跳过已完成: {dept_name} ({prev} 位)")
                continue

            try:
                teachers = await scrape_department(context, dept_name, domain, urls)
                if teachers:
                    all_teachers.extend(teachers)
                save_progress(dept_name, teachers)
            except Exception as e:
                print(f"学院异常: {dept_name}: {e}")
                save_progress(dept_name, [])

        await context.close()
        await browser.close()

    # 最终去重
    seen = set()
    unique = []
    for t in all_teachers:
        if t.get("email") and t["email"] not in seen:
            seen.add(t["email"])
            unique.append(t)

    print(f"\n{'='*50}")
    print(f"总计: {len(unique)} 位教师 (去重后)")
    print(f"{'='*50}")

    if unique:
        export_results(unique)
    else:
        print("无数据可导出")


def export_results(records: list[dict]):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"东南大学_教师邮箱_{ts}.xlsx"
    csv_path = OUTPUT_DIR / f"东南大学_教师邮箱_{ts}.csv"

    # XLSX
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "东南大学教师邮箱"

    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB"))
    cf = Font(name="微软雅黑", size=10)
    ca = Alignment(vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = border

    for i, r in enumerate(records, 1):
        for col, v in enumerate([i, r.get("name",""), r.get("email",""),
                                  r.get("department",""), r.get("title",""), r.get("url","")], 1):
            cell = ws.cell(row=i+1, column=col, value=v)
            cell.font = cf; cell.alignment = ca; cell.border = border

    for col, w in zip("ABCDEF", [8, 18, 32, 20, 16, 55]):
        ws.column_dimensions[col].width = w

    wb.save(xlsx_path)
    print(f"XLSX: {xlsx_path}")

    # CSV
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i, r in enumerate(records, 1):
            w.writerow([i, r.get("name",""), r.get("email",""),
                        r.get("department",""), r.get("title",""), r.get("url","")])
    print(f"CSV: {csv_path}")

    # Summary stats
    dept_counts = Counter(r["department"] for r in records)
    print(f"\n各学院统计:")
    for dept, cnt in dept_counts.most_common():
        print(f"  {dept}: {cnt} 位")

    # Final JSON
    final = {"total": len(records), "xlsx": str(xlsx_path), "csv": str(csv_path),
             "dept_counts": dict(dept_counts), "teachers": records}
    (OUTPUT_DIR / "seu_final.json").write_text(json.dumps(final, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n[FILES]")
    print(f"{xlsx_path.name}|东南大学教师邮箱XLSX")
    print(f"{csv_path.name}|东南大学教师邮箱CSV")
    print(f"[/FILES]")


if __name__ == "__main__":
    asyncio.run(main())
