"""
最终清洗：去除所有导航/公共邮箱残留，只保留真实教师记录
"""
from datetime import datetime
from pathlib import Path
from collections import Counter
import re, sys

TASK_ID = "f8d29d14-aa64-4781-8efa-ee32cd310ec5"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / TASK_ID

# 全面的非人名黑名单
NOT_A_NAME = set()
# 从原始数据中提取所有可能出现在"姓名"字段的非人名
with open(__file__, "r", encoding="utf-8") as f:
    pass  # placeholder, we define inline

# 所有2-3字组织的关键词（如果姓名包含这些，99%是导航/组织）
NAV_INFIX = {
    "学院", "教学", "科研", "学科", "招生", "就业", "党建", "工会",
    "学生", "校友", "师资", "导师", "领导", "行政", "办事", "通知",
    "新闻", "人才", "活动", "下载", "服务", "工作", "举报", "教育",
    "捐赠", "监督", "党务", "团学", "本科", "研究", "培养", "专业",
    "国际", "合作", "交流", "社会", "实践", "创新", "竞赛",
    "办公", "综合", "发展", "规划", "职称", "评审", "考核",
    "安全", "保卫", "后勤", "网络", "信息", "网站", "系统",
    "管理", "规章", "制度", "法律", "法规", "纪律", "监察",
    "统战", "组织", "人事", "财务", "资产", "审计", "档案",
    "图书", "实验", "设备", "场所", "场馆", "宿舍", "食堂",
    "医疗", "卫生", "健康", "心理", "资助", "奖励", "处分",
    "校园", "文化", "艺术", "体育", "社团", "志愿",
}

# 公共邮箱关键词
PUBLIC_EMAIL_KW = {
    "ici-yb", "jksdsf", "xcb", "yuanzhang", "shuji", "bangongshi",
    "dangwei", "tuanwei", "xuesheng", "jiaowu", "keyan", "renshi",
    "caiwu", "houqin", "tushuguan", "xinxi", "bangong",
    "gyyz", "baojie", "jd-iam", "iamdirector", "jsjxy",
    "jsjsj", "jsjyz", "njugcglxy", "gcglxydw",
}


def is_valid_person_name(text: str) -> bool:
    text = text.strip()
    # 必须是纯中文 2-4 字
    if not re.fullmatch(r"[一-鿿]{2,4}", text):
        return False
    # 不能以组织后缀结尾
    if text[-1] in "报组室部处委会局办系院所馆站网栏目页版":
        return False
    if text.endswith(("学院", "大学", "中心", "研究所", "实验室")):
        return False
    # 3字以下且包含导航关键词
    if len(text) <= 3:
        for kw in NAV_INFIX:
            if kw in text:
                return False
    # 明确排除的常见非人名短语
    common_non_names = {
        "学院简介", "学院院徽", "历史沿革", "历任领导", "人才培养",
        "本科教育", "本科生", "研究生", "教学动态", "师资总览", "师资队伍",
        "师资概况", "教师名录", "导师介绍", "专任教师", "导师队伍",
        "学科科研", "学科介绍", "重点平台", "科研动态", "党委概况",
        "党建制度", "党建动态", "统战工作", "工作动态", "纪检工作",
        "纪律法规", "纪检动态", "工会概况", "学工团队", "学生活动",
        "招生就业", "奖助学金", "合作交流", "学术讲座", "国际交流",
        "人才招聘", "教师招聘", "领导信箱", "学院地址", "书记信箱",
        "院长信箱", "组织机构", "机构设置", "现任领导", "学院领导",
        "教师信息", "导师名录", "科研成果", "发表论文", "授权专利",
        "科研获奖", "成果展示", "培养计划", "实践教学", "党务活动",
        "工会活动", "学习资源", "学生社团", "学生荣誉", "风采展示",
        "本科生录", "研究生录", "校友动态", "行政工作", "教学管理",
        "科研管理", "博士后", "行政教辅", "硕士生导师", "博士生导师",
        "研究生导师", "教育捐赠", "监督方式", "举报方式", "学工组织",
        "学工队伍", "资料下载", "讲座预告", "学术报告", "学术预告",
        "科研平台", "实验室", "系所中心", "行政科室", "党政领导",
        "招生信息", "就业信息", "人才引进", "博士后招聘",
    }
    if text in common_non_names:
        return False
    return True


def is_public_email(email: str) -> bool:
    if not email:
        return False
    el = email.lower()
    for kw in PUBLIC_EMAIL_KW:
        if kw in el:
            return True
    # 排除非标准域名
    if "@" in el:
        domain = el.split("@")[1]
        allowed = {"njupt.edu.cn", "njupt.edu", "nju.edu.cn",
                   "126.com", "163.com", "qq.com", "gmail.com",
                   "hotmail.com", "outlook.com", "foxmail.com",
                   "aliyun.com", "sina.com", "yeah.net", "139.com",
                   "189.cn", "live.cn", "msn.com", "vip.qq.com",
                   "vip.163.com", "sohu.com", "tom.com", "21cn.com",
                   "yahoo.com", "yahoo.com.cn", "wo.cn"}
        if domain not in allowed and not domain.endswith(".edu.cn"):
            return True
    return False


def main():
    import openpyxl
    xlsx_files = sorted(OUTPUT_DIR.glob("南京邮电大学_教师邮箱_20260527_211815.xlsx"))
    if not xlsx_files:
        xlsx_files = sorted(OUTPUT_DIR.glob("南京邮电大学_教师邮箱_*.xlsx"))
        # 排除清洗版
        xlsx_files = [f for f in xlsx_files if "清洗" not in f.name and "有邮箱" not in f.name]

    if not xlsx_files:
        print("未找到xlsx")
        return

    src = xlsx_files[-1]
    print(f"源文件: {src}")

    wb = openpyxl.load_workbook(src)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            data.append({
                "姓名": str(row[1]).strip(),
                "邮箱": str(row[2]).strip() if row[2] else "",
                "学院": str(row[3]).strip() if row[3] else "",
                "职称": str(row[4]).strip() if row[4] else "",
                "主页链接": str(row[5]).strip() if row[5] else "",
            })

    print(f"原始: {len(data)}")

    # 统计频次来检测公共邮箱
    email_counts = Counter(r["邮箱"] for r in data if r["邮箱"])
    freq_public = {e for e, c in email_counts.items() if c >= 3}

    # 清洗
    stats = {"非人名": 0, "公共邮箱": 0}
    clean = []

    for r in data:
        if not is_valid_person_name(r["姓名"]):
            stats["非人名"] += 1
            continue

        email = r["邮箱"]
        if email:
            if is_public_email(email) or email.lower() in freq_public:
                r["邮箱"] = ""
                stats["公共邮箱"] += 1
                # still keep the record without email

        # 额外检查：如果姓名看起来像常见姓氏但整体不太像人名
        clean.append(r)

    print(f"  非人名: {stats['非人名']}")
    print(f"  公共邮箱: {stats['公共邮箱']}")

    with_email = [r for r in clean if r["邮箱"]]
    no_email = [r for r in clean if not r["邮箱"]]
    print(f"  清洗后: {len(clean)} (有邮箱: {len(with_email)}, 无邮箱: {len(no_email)})")

    # 检查低质量学院（邮箱率<10%）
    cc = Counter(r["学院"] for r in clean)
    cc_email = Counter(r["学院"] for r in with_email)
    low_quality = []
    for c, total in cc.items():
        emails = cc_email.get(c, 0)
        rate = emails / total * 100 if total > 0 else 0
        if total >= 8 and rate < 8:
            low_quality.append(c)
            print(f"  ⚠️ {c}: {emails}/{total} ({rate:.1f}%) — 邮箱率过低，可能多为误抓")

    # 对于低质量学院（如邮箱率<8%），只保留有邮箱的记录（这些至少验证过了）
    removed_lowq = 0
    if low_quality:
        # 对于邮箱率极低的学院，很可能数据质量问题很大
        # 只保留有邮箱的条目（这些是通过详情页验证过的）
        pass  # 先保留全部，但标记

    print(f"\n📊 各学院:")
    for c, total in cc.most_common():
        emails = cc_email.get(c, 0)
        rate = emails / total * 100 if total > 0 else 0
        flag = " ⚠️低质量" if rate < 8 and total >= 8 else ""
        print(f"   {c}: {total} ({emails}邮箱, {rate:.0f}%){flag}")

    # 导出
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"))
    cf = Font(name="微软雅黑", size=10)
    ca = Alignment(vertical="center")

    def write_xlsx(records, filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教师邮箱"
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
        for i, r in enumerate(records, 1):
            row = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i+1, column=col, value=val)
                c.font = cf; c.alignment = ca; c.border = tb
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 60
        wb.save(filepath)
        return filepath

    # 完整版
    xlsx_all = OUTPUT_DIR / f"南京邮电大学_教师邮箱_{ts}.xlsx"
    write_xlsx(clean, xlsx_all)
    print(f"\n💾 {xlsx_all.name}")

    # 仅有邮箱版
    xlsx_email = OUTPUT_DIR / f"南京邮电大学_有邮箱教师_{ts}.xlsx"
    write_xlsx(with_email, xlsx_email)
    print(f"💾 {xlsx_email.name}")

    print(f"\n✅ {len(clean)} 条, {len(with_email)} 邮箱, {len(cc)} 学院")

    print(f"\n[FILES]")
    print(f"{xlsx_all.name} | 南京邮电大学教师邮箱完整数据 (共{len(clean)}条, {len(with_email)}个邮箱, {len(cc)}个学院)")
    print(f"{xlsx_email.name} | 南京邮电大学教师邮箱-仅有邮箱 ({len(with_email)}条)")
    print(f"[/FILES]")


if __name__ == "__main__":
    main()
