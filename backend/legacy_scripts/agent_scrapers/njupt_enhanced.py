"""
南京邮电大学教师邮箱爬虫 v4 — 增强版
改进：
1. 从主站获取正确的学院URL
2. 更严格的教师链接过滤
3. 更优的数据清洗
"""

import asyncio
import csv
import re
import sys
import logging
from datetime import datetime
from pathlib import Path
from collections import Counter

try:
    from playwright.async_api import async_playwright
except ImportError:
    print("pip install playwright && playwright install chromium")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

TASK_ID = "f8d29d14-aa64-4781-8efa-ee32cd310ec5"
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "outputs" / TASK_ID
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

TITLE_KW = [
    "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
    "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
    "长江学者", "杰青", "优青", "博士后", "实验师", "高级实验师",
    "青年专聘教授", "校聘副教授", "预聘副教授", "特聘教授",
    "客座教授", "名誉教授", "国家级教学名师",
]

PUBLIC_LOCAL = {"webmaster", "admin", "info", "office", "master", "president",
                "xb", "xxgk", "jwc", "yjsc", "rsc", "gjc", "tw", "xsc",
                "wxyxz", "sxy", "dw", "yb_eng", "yb_", "oice", "xwzx", "jcjs",
                "jsjsj", "jsjyz", "jsjxy", "sxydw", "njugcglxy", "gcglxydw", "njudz",
                "jd-iam", "iamdirector", "Jjy_sj", "zdh"}

NAV_KW = ["学院概况", "组织机构", "机构设置", "新闻", "通知", "公告", "招生",
          "党建", "学生工作", "团学", "工会", "校友", "实验室", "联系我们",
          "下载", "规章制度", "师德师风", "信息公开", "网站地图",
          "师资队伍", "师资概况", "教师名录", "专任教师", "导师介绍",
          "现任领导", "学院领导", "领导信箱", "诚聘英才",
          "本科生", "研究生", "专业介绍", "培养", "教学",
          "科研", "学科", "学术", "竞赛", "创新",
          "返回", "首页", "南邮主页", "校内链接",
          "更多", "详情", "下一页", "上一页",
          "函授", "自考", "培训", "技能", "干部", "成教",
          "网上学习", "表格下载", "学习资源", "综合管理",
          "行政", "党务", "工会活动", "学生社团", "学生荣誉",
          "风采展示", "校友动态", "教育捐赠", "监督方式",
          "培养计划", "实践教学", "专接本", "二学历",
          "社会自考", "专业计划", "教学管理", "科研管理",
          "校级", "招聘", "发表论文", "授权专利", "科研获奖",
          "成果展示", "教师信息", "导师名录", "本科生录", "研究生录",
          "羽毛球队", "第十三周", "拟发展"]


def is_chinese_name(text: str) -> bool:
    text = text.strip()
    if not re.fullmatch(r"[一-鿿]{2,4}", text):
        return False
    if text in NAV_KW:
        return False
    if text[-1] in "报组室部处委会局办系院所馆站网栏目页版":
        return False
    # 组织/机构名后缀
    if any(text.endswith(s) for s in ["学院", "大学", "中心", "研究所"]):
        return False
    return True


def is_public_email(email: str) -> bool:
    if not email:
        return True
    email_lower = email.lower()
    local = email_lower.split("@")[0]
    if any(kw in local for kw in PUBLIC_LOCAL):
        return True
    # 排除非edu邮箱 (来自导航页面)
    if "@" in email_lower and not any(domain in email_lower for domain in
                                       ["njupt.edu.cn", "njupt.edu", "nju.edu.cn"]):
        return True
    return False


def parse_at_sign(text: str) -> str:
    text = re.sub(r"\s*\[at\]\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*\(at\)\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*#@\s*", "@", text)
    text = re.sub(r"\s*\[@\]\s*", "@", text)
    text = re.sub(r"\s*\(@\)\s*", "@", text)
    text = re.sub(r"\s*\[dot\]\s*", ".", text, flags=re.IGNORECASE)
    return text


def extract_titles(text: str) -> str:
    found = []
    for kw in TITLE_KW:
        if kw in text:
            found.append(kw)
    return "、".join(found[:5])


async def discover_colleges_from_main(context) -> list[tuple]:
    """从南邮主站获取学院URL"""
    page = await context.new_page()
    colleges = []

    try:
        # 尝试多个可能的教学单位页面
        urls_to_try = [
            "https://www.njupt.edu.cn",
            "https://www.njupt.edu.cn/xxgk/jxjj.htm",
            "https://www.njupt.edu.cn/1/list.htm",
        ]

        all_found_links = []
        for url in urls_to_try:
            try:
                resp = await page.goto(url, wait_until="domcontentloaded", timeout=15000)
                if resp and resp.status >= 400:
                    continue
                await asyncio.sleep(2)

                links = await page.evaluate("""() => {
                    const ls = [];
                    document.querySelectorAll('a[href]').forEach(a => {
                        const t = (a.textContent || '').trim();
                        const h = a.href;
                        if (!t || !h || h.startsWith('javascript:') || h === '#') return;
                        if (h === 'https://www.njupt.edu.cn' || h === 'https://www.njupt.edu.cn/') return;
                        ls.push({t: t.substring(0, 60), h: h});
                    });
                    return ls;
                }""")

                # 找含"学院"的链接
                for l in links:
                    t = l["t"].strip()
                    h = l["h"]
                    if re.search(r"学院|体育部|研究所", t) and len(t) < 25:
                        if "njupt.edu.cn" in h:
                            all_found_links.append((t, h))
            except Exception:
                continue

        # 去重
        seen = set()
        for name, url in all_found_links:
            if url not in seen:
                seen.add(url)
                colleges.append((name, url))
                logger.info(f"  🏫 发现学院: {name} → {url}")

        await page.close()
        return colleges

    except Exception:
        await page.close()
        return []


async def find_teacher_list_pages(context, college_url: str) -> list[str]:
    """访问学院首页，找教师列表页"""
    page = await context.new_page()
    teacher_urls = []

    try:
        await page.goto(college_url, wait_until="domcontentloaded", timeout=15000)
        await asyncio.sleep(2)

        links = await page.evaluate("""() => {
            const ls = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const t = (a.textContent || '').trim();
                const h = a.href;
                if (!t || !h || h.startsWith('javascript:') || h === '#') return;
                if (h.includes('mailto:')) return;
                ls.push({t: t.substring(0, 60), h: h});
            });
            return ls;
        }""")

        for l in links:
            if any(kw in l["t"] for kw in ["师资", "教师", "导师", "队伍建设"]):
                if len(l["t"]) < 15:
                    teacher_urls.append(l["h"])

        logger.info(f"    找到 {len(teacher_urls)} 个教师列表入口")
    except Exception:
        pass
    finally:
        await page.close()

    return teacher_urls


async def extract_teacher_links(context, list_url: str, college_name: str) -> list[dict]:
    """从教师列表页提取教师详情链接"""
    page = await context.new_page()
    entries = []

    try:
        resp = await page.goto(list_url, wait_until="domcontentloaded", timeout=20000)
        if resp and resp.status >= 400:
            await page.close()
            return entries
        await asyncio.sleep(2)

        all_links = await page.evaluate("""() => {
            const ls = [];
            document.querySelectorAll('a[href]').forEach(a => {
                const t = (a.textContent || '').trim();
                const h = a.href;
                if (!t || !h || h.startsWith('javascript:') || h === '#') return;
                if (h.includes('mailto:')) return;
                ls.push({t: t.substring(0, 80), h: h});
            });
            return ls;
        }""")

        for link in all_links:
            text = link["t"].strip()
            url = link["h"]

            # 链接文本是中文姓名
            if is_chinese_name(text):
                entries.append({"name": text, "url": url, "college": college_name})
                continue

            # yjs 导师系统详情页（最可靠来源）
            if "yjs.njupt.edu.cn" in url and "dsfcxq" in url:
                entries.append({"name": text, "url": url, "college": college_name})
                continue

            # CS学院page.htm模式
            if "/page.htm" in url and college_name.startswith("计算机"):
                entries.append({"name": text, "url": url, "college": college_name})
                continue

        logger.info(f"  {list_url}: {len(entries)} 个教师链接")
        return entries

    except Exception as e:
        logger.warning(f"  {list_url}: {str(e)[:60]}")
        return []
    finally:
        await page.close()


async def scrape_profile(context, entry: dict) -> dict | None:
    """访问详情页提取信息"""
    page = await context.new_page()
    try:
        url = entry["url"]
        await page.goto(url, wait_until="domcontentloaded", timeout=15000)
        await asyncio.sleep(0.8)

        # 提取姓名
        name = entry["name"]
        if not is_chinese_name(name):
            title_text = await page.evaluate("() => document.title || ''")
            m = re.match(r"^([一-鿿]{2,4})\s*[-–—|]", title_text)
            if m and is_chinese_name(m.group(1)):
                name = m.group(1)

        if not is_chinese_name(name):
            body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
            m2 = re.search(r"姓\s*名\s*[：:]\s*([一-鿿]{2,4})", body_text)
            if m2:
                name = m2.group(1)

        if not is_chinese_name(name):
            body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
            headings = await page.evaluate("""() => {
                const hs = [];
                document.querySelectorAll('h1,h2,h3').forEach(h => hs.push(h.textContent.trim()));
                return hs;
            }""")
            for h in headings:
                m3 = re.match(r"^([一-鿿]{2,4})", h)
                if m3 and is_chinese_name(m3.group(1)):
                    name = m3.group(1)
                    break
            else:
                if not body_text:
                    body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
                for line in body_text.split("\n")[:5]:
                    line = line.strip()
                    if is_chinese_name(line):
                        name = line
                        break

        if not is_chinese_name(name):
            return None

        # 邮箱
        body_text = await page.evaluate("() => document.body ? document.body.innerText : ''")
        body_text = parse_at_sign(body_text)
        html = await page.evaluate("() => document.body ? document.body.innerHTML : ''")
        html = parse_at_sign(html)

        email = ""
        em = re.search(r"电子邮箱\s*[：:]\s*(\S+@\S+)", body_text)
        if em:
            email = em.group(1).strip()
        else:
            all_text = body_text + " " + html
            emails = EMAIL_RE.findall(all_text)
            email = next((e for e in emails if not is_public_email(e)), "")

        if is_public_email(email):
            email = ""

        # 只保留 @njupt.edu.cn 或合理的外部邮箱（比如老师用的126/gmail等）
        # 但排除明显不是教师邮箱的（如 pku.edu.cn）
        if email and "@" in email:
            domain = email.split("@")[1].lower()
            if domain in ("pku.edu.cn", "tsinghua.edu.cn", "fudan.edu.cn"):
                email = ""

        title = extract_titles(body_text[:3000])

        if email:
            logger.info(f"  ✅ {name} → {email} | {title[:30]} [{entry['college']}]")

        return {
            "姓名": name,
            "邮箱": email,
            "学院": entry["college"],
            "职称": title,
            "主页链接": url,
        }

    except Exception:
        return None
    finally:
        await page.close()


async def main():
    logger.info("=" * 60)
    logger.info("🎓 南京邮电大学 全学院教师邮箱爬虫 v4")
    logger.info("=" * 60)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36"

        # === 阶段1: 从主站发现学院 ===
        logger.info("\n📌 阶段1: 从主站发现学院URL")
        ctx = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
        discovered = await discover_colleges_from_main(ctx)
        await ctx.close()

        # 合并手动配置
        manual_colleges = {
            "计算机学院/软件学院/网络空间安全学院": ["https://cs.njupt.edu.cn/18762/list.htm"],
        }
        college_teacher_urls = {}

        # 从发现的学院中找教师列表页
        logger.info(f"\n📌 阶段2: 查找各学院教师列表页")
        for name, url in discovered:
            logger.info(f"  🔍 {name}: {url}")
            ctx2 = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
            teacher_pages = await find_teacher_list_pages(ctx2, url)
            await ctx2.close()
            if teacher_pages:
                college_teacher_urls[name] = teacher_pages

        # 合并手动配置
        for name, urls in manual_colleges.items():
            if name in college_teacher_urls:
                college_teacher_urls[name].extend(urls)
            else:
                college_teacher_urls[name] = urls

        # 加上直接已知有效的URL
        known_urls = {
            "计算机学院/软件学院/网络空间安全学院": [
                "https://cs.njupt.edu.cn/18762/list.htm",
            ],
        }
        for name, urls in known_urls.items():
            if name not in college_teacher_urls:
                college_teacher_urls[name] = urls

        logger.info(f"\n📊 共 {len(college_teacher_urls)} 个学院有教师页面")

        # === 阶段3: 收集教师详情链接 ===
        logger.info(f"\n📌 阶段3: 收集教师详情链接")
        all_entries = []
        seen_urls = set()

        for college_name, urls in college_teacher_urls.items():
            logger.info(f"\n🏫 {college_name}")
            ctx3 = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
            try:
                for url in urls:
                    entries = await extract_teacher_links(ctx3, url, college_name)
                    for e in entries:
                        if e["url"] not in seen_urls:
                            seen_urls.add(e["url"])
                            all_entries.append(e)
                logger.info(f"  累计: {len([e for e in all_entries if e['college'] == college_name])} 个")
            except Exception as e:
                logger.warning(f"  ❌ {college_name}: {str(e)[:60]}")
            finally:
                await ctx3.close()

        logger.info(f"\n📊 共 {len(all_entries)} 个教师详情链接")

        # === 阶段4: 批量爬取 ===
        logger.info(f"\n📌 阶段4: 批量爬取详情页")
        all_results = []
        sem = asyncio.Semaphore(5)

        async def process(entry):
            async with sem:
                ctx = await browser.new_context(viewport={"width": 1920, "height": 1080}, user_agent=ua)
                try:
                    return await scrape_profile(ctx, entry)
                finally:
                    await ctx.close()

        batch_size = 20
        for start in range(0, len(all_entries), batch_size):
            end = min(start + batch_size, len(all_entries))
            batch = all_entries[start:end]
            logger.info(f"\n  [{start+1}-{end}/{len(all_entries)}]")
            tasks = [process(e) for e in batch]
            results = await asyncio.gather(*tasks)
            batch_count = 0
            for r in results:
                if r:
                    all_results.append(r)
                    batch_count += 1
            logger.info(f"    本批: {batch_count}, 累计: {len(all_results)}")

        await browser.close()

    # === 阶段5: 清洗 ===
    logger.info(f"\n{'='*60}")
    logger.info("📊 数据清洗")

    # 去重
    seen_key = set()
    clean = []
    for r in all_results:
        key = (r["姓名"], r["邮箱"])
        if key not in seen_key:
            seen_key.add(key)
            clean.append(r)

    # 严格过滤
    final = []
    for r in clean:
        name = r["姓名"]
        # 必须是中文姓名
        if not is_chinese_name(name):
            continue
        # 必须是有邮箱或合理的链接
        if not r["邮箱"] and not r["主页链接"]:
            continue
        # 邮箱不能是公共的
        if r["邮箱"] and is_public_email(r["邮箱"]):
            # 把邮箱去掉但保留记录
            r["邮箱"] = ""
            final.append(r)
        else:
            final.append(r)

    logger.info(f"  原始={len(all_results)} → 去重={len(clean)} → 过滤={len(final)}")

    with_email = [r for r in final if r["邮箱"]]
    no_email = [r for r in final if not r["邮箱"]]
    logger.info(f"  有邮箱: {len(with_email)}, 无邮箱: {len(no_email)}")

    if with_email:
        logger.info(f"\n📋 有邮箱教师 (前30条):")
        for r in with_email[:30]:
            logger.info(f"   {r['姓名']} <{r['邮箱']}> [{r['学院']}]")

    # === 导出 XLSX ===
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"南京邮电大学_教师邮箱_{timestamp}.xlsx"

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教师邮箱"

    HEADERS_CN = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    cell_font = Font(name="微软雅黑", size=10)

    for col, h in enumerate(HEADERS_CN, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(final, 1):
        row_data = [i, r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = cell_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 60

    wb.save(xlsx_path)
    logger.info(f"\n💾 XLSX: {xlsx_path}")

    # 无邮箱名单
    if no_email:
        ne_path = OUTPUT_DIR / f"南京邮电大学_无邮箱教师_{timestamp}.csv"
        with open(ne_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=["姓名", "邮箱", "学院", "职称", "主页链接"])
            w.writeheader()
            w.writerows(no_email)
        logger.info(f"💾 无邮箱: {ne_path}")

    college_counts = Counter(r["学院"] for r in final)
    logger.info(f"\n📊 各学院:")
    for c, cnt in college_counts.most_common():
        e_cnt = len([r for r in final if r["学院"] == c and r["邮箱"]])
        logger.info(f"   {c}: {cnt} (有邮箱: {e_cnt})")

    logger.info(f"\n✅ 完成! {len(final)} 条, {len(with_email)} 个邮箱")

    print(f"\n[FILES]")
    print(f"{xlsx_path.name} | 南京邮电大学教师邮箱完整数据 ({len(final)}条, {len(with_email)}个邮箱)")
    print(f"[/FILES]")

    return xlsx_path


if __name__ == "__main__":
    asyncio.run(main())
