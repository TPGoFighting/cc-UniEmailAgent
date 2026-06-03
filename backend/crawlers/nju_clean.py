"""
南京大学爬取数据清洗 + 整合脚本
"""

import csv, os, re
from datetime import datetime
from collections import Counter, defaultdict

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "outputs")
TASK_DIR = os.path.join(OUTPUT_DIR, "nju_final")
os.makedirs(TASK_DIR, exist_ok=True)

# 公共邮箱前缀
PUBLIC = ['webmaster', 'admin', 'office', 'info', 'master', 'root',
          'postmaster', 'bgs', 'dangzheng', 'yuanban', 'wxyxz', 'xwcb', 'yanju',
          'yzyx', 'hoffice', 'hydw', 'njudz']

def is_valid_email(e):
    if not e: return False
    e = e.strip()
    if not e: return False
    if e.endswith('.css') or e.endswith('.js') or e.endswith('.min.js'):
        return False
    if '@' not in e: return False
    prefix = e.split('@')[0].lower()
    for p in PUBLIC:
        if p in prefix: return False
    if not e.endswith('.edu.cn') and not e.endswith('.edu') and not any(e.endswith(d) for d in ['.com', '.cn', '.org', '.net', '.io']):
        if '@' in e:
            domain = e.split('@')[1]
            if '.' not in domain: return False
    return True

def is_valid_name(n):
    if not n: return False
    n = n.strip()
    # 纯中文2-4字姓名
    if re.match(r'^[一-鿿]{2,4}$', n): return True
    return False

# 已知导航/非教师词汇
NAV_WORDS = set("首页 学院概况 师资队伍 科学研究 人才培养 新闻动态 通知公告 联系我们 联系 联系方式 "
    "学院简介 学院领导 历史渊源 院系领导 组织机构 专业设置 规章制度 诚聘英才 "
    "行政管理 退休教师 现任教师 师资 在职教师 人才招聘 招生就业 教育教学 "
    "南大概况 南大新闻 学院部门 校园服务 国际合作 English 本科生 研究生 "
    "留学生 党建 工会 团委 返回 更多 友情链接 网站地图 搜索 办公系统 "
    "书记信箱 院长信箱 师德监督 学校主页 南大主页 南大邮箱 内网入口 "
    "网站首页 设为首页 加入收藏 版权 联系地址 邮政编码 电话 传真 "
    "友情链接 常用链接 快速通道 信息门户 网上办事 服务大厅 校园卡 "
    "教务 科研 学工 研究生院 加入我们 English 用户登录 学生登录 教师登录".split())

def is_nav_word(n):
    return n in NAV_WORDS

# 有职称评定的排除
TITLE_WORDS = {'教授', '副教授', '讲师', '助教', '研究员', '副研究员', '助理研究员',
               '博导', '硕导', '博士后', '助理教授', '工程师', '高级工程师',
               '主任医师', '副主任医师', '实验师', '高级实验师'}

# ===== 读取v3数据 =====
v3_path = None
for f in os.listdir(os.path.join(OUTPUT_DIR, "nju_v3")):
    if f.endswith('.csv'):
        v3_path = os.path.join(OUTPUT_DIR, "nju_v3", f)
        break

if not v3_path:
    print("未找到v3 CS数据！")
    exit(1)

print(f"读取: {v3_path}")

rows = []
with open(v3_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

print(f"原始记录: {len(rows)}")

# ===== 清洗 =====
clean = []
seen = set()

for r in rows:
    name = r.get("姓名", "").strip()
    email = r.get("邮箱", "").strip()
    dept = r.get("学院", "").strip()
    title = r.get("职称", "").strip()
    url = r.get("主页链接", "").strip()

    # 过滤无效姓名
    if not is_valid_name(name):
        # 有些记录只有邮箱没有姓名
        if is_valid_email(email):
            name = email.split('@')[0]
        else:
            continue

    # 过滤导航词汇
    if is_nav_word(name):
        continue

    # 无效邮箱置空
    if not is_valid_email(email):
        email = ""

    # 去重
    key = f"{name}|{email}|{dept}"
    if key in seen:
        continue
    seen.add(key)

    clean.append({
        "姓名": name,
        "邮箱": email,
        "学院": dept,
        "职称": title if title else "",
        "主页链接": url if (url and url.startswith('http')) else ""
    })

print(f"清洗后: {len(clean)}")

# ===== 统计 =====
dept_count = Counter()
dept_email = Counter()
dept_data = defaultdict(list)

for r in clean:
    d = r["学院"]
    dept_count[d] += 1
    dept_data[d].append(r)
    if r["邮箱"]:
        dept_email[d] += 1

# ===== 合并计算机学院已有数据 =====
# 从之前的精确爬取中加载
cs_data_path = os.path.join(OUTPUT_DIR, "a468ea4c-1347-4a08-adc8-696eb43df27c", "南京大学_计算机学院_教师邮箱_20260528_091423.xlsx")
# 尝试读取已有CSV
cs_csv = None
for root, dirs, files in os.walk(OUTPUT_DIR):
    for f in files:
        if '南京大学_计算机学院' in f and f.endswith('.csv'):
            cs_csv = os.path.join(root, f)
            break

if cs_csv:
    print(f"\n找到计算机学院已有CSV: {cs_csv}")
    cs_count = 0
    with open(cs_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r.get("姓名", "").strip()
            email = r.get("邮箱", "").strip()
            title = r.get("职称", "").strip()
            # 学院字段可能有不同名称
            dept = r.get("学院", "计算机学院") or "计算机学院"

            if is_valid_name(name):
                key = f"{name}|{email}|计算机学院"
                if key not in seen:
                    seen.add(key)
                    clean.append({
                        "姓名": name,
                        "邮箱": email if is_valid_email(email) else (email if email else ""),
                        "学院": "计算机学院",
                        "职称": title,
                        "主页链接": ""
                    })
                    cs_count += 1
    print(f"从已有CSV合并: {cs_count} 条")
else:
    print("未找到计算机学院已有CSV，仅使用v3数据")
    # 尝试从excel读取
    try:
        import openpyxl
        xlsx_path = os.path.join(OUTPUT_DIR, "a468ea4c-1347-4a08-adc8-696eb43df27c", "南京大学_计算机学院_教师邮箱_20260528_091423.xlsx")
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
            for r in range(2, ws.max_row+1):
                vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
                if len(vals) >= 4:
                    name = str(vals[1] or '').strip()
                    email = str(vals[2] or '').strip()
                    title = str(vals[4] or '').strip() if len(vals) > 4 else ''
                    if is_valid_name(name):
                        key = f"{name}|{email}|计算机学院"
                        if key not in seen:
                            seen.add(key)
                            clean.append({
                                "姓名": name,
                                "邮箱": email if is_valid_email(email) else "",
                                "学院": "计算机学院",
                                "职称": title,
                                "主页链接": ""
                            })
            print(f"从Excel合并计算机学院数据")
    except Exception as e:
        print(f"读取Excel失败: {e}")

# ===== 重新统计 =====
dept_count = Counter()
dept_email = Counter()
for r in clean:
    d = r["学院"]
    dept_count[d] += 1
    if r["邮箱"]:
        dept_email[d] += 1

# ===== 输出 =====
ts = datetime.now().strftime("%Y%m%d_%H%M%S")

# 详细CSV
csv_path = os.path.join(TASK_DIR, f"南京大学_全部教师邮箱_{ts}.csv")
with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(["姓名", "邮箱", "学院", "职称", "主页链接"])
    for r in clean:
        w.writerow([r["姓名"], r["邮箱"], r["学院"], r["职称"], r["主页链接"]])

# 统计CSV
stat_path = os.path.join(TASK_DIR, f"南京大学_学院统计_{ts}.csv")
with open(stat_path, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(["学院", "教师数", "有邮箱数", "状态"])
    small_colleges = []
    for d, c in sorted(dept_count.items(), key=lambda x: -x[1]):
        ec = dept_email.get(d, 0)
        status = "⚠️ <50需补充爬取" if c < 50 else "✅"
        if c < 50:
            small_colleges.append((d, c, ec))
        w.writerow([d, c, ec, status])

# ===== 打印 =====
print(f"\n{'='*60}")
print("南京大学全学院教师数据统计（清洗后）")
print(f"{'='*60}")
print(f"{'学院':<28} {'教师':<6} {'邮箱':<6} {'状态'}")
print(f"{'-'*60}")
total_t = 0
total_e = 0
for d, c in sorted(dept_count.items(), key=lambda x: -x[1]):
    ec = dept_email.get(d, 0)
    total_t += c
    total_e += ec
    flag = "" if c >= 50 else " ⚠️"
    print(f"{d:<26} {c:<4} {ec:<4}{flag}")

print(f"{'-'*60}")
print(f"总计: {total_t} 教师, {total_e} 有邮箱")

if small_colleges:
    print(f"\n{'='*60}")
    print(f"⚠️ 教师数 < 50 的学院（需要替代策略重点爬取）：")
    for d, c, ec in small_colleges:
        print(f"  {d}: {c} 教师 ({ec} 有邮箱)")

print(f"\nCSV: {csv_path}")
print(f"统计: {stat_path}")
