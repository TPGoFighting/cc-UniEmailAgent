"""合并补全数据到现有CSV — 智能过滤非教师条目。

过滤策略：
1. 姓名在导航黑名单中的 → 移除
2. 姓名中包含导航关键词的 → 移除
3. 姓名看起来不像中文人名的 → 移除
4. 保留有邮箱或有职称信息的记录，但对明显是部门邮箱的做清空处理
"""

import csv
import re
from collections import defaultdict
from pathlib import Path
from datetime import datetime

OUTPUT_DIR = Path(__file__).parent / "outputs"

# 上次合并的最佳版本作基
EXISTING_CSV = OUTPUT_DIR / "南京大学_教师邮箱_合并版_20260526_161858.csv"

# 新爬取的14院系CSV
NEW_CSV = OUTPUT_DIR / "南京大学_补全14院系_20260526_161502.csv"
# 最后3院系CSV
LAST3_CSV = OUTPUT_DIR / "南京大学_最后3院系_20260526_163201.csv"

# ⭐ 非人名的导航/分类文本（在黑名单中的条目直接移除）
NON_NAME_BLACKLIST = {
    # 通用导航
    "首页", "首页信息", "南大主页", "学校主页", "学院首页",
    "后退", "前进", "返回", "返回首页", "设为首页",
    # 学院/机构导航
    "学院简介", "学院介绍", "学院概况", "院况概览",
    "学院领导", "院系领导", "现任领导", "机构设置",
    "学科介绍", "系科设置", "教研平台",
    "委员会", "工会", "职能部门", "群众团体",
    "学院标识", "学院标志",
    # 师资导航（这些是分类页，不是具体教师）
    "师资力量", "师资队伍", "教师名录", "教师主页",
    "全院名录", "在职教师", "专任教师", "退休教师", "荣休教师",
    "教授", "副教授", "讲师", "研究员", "两院院士",
    "研究系列", "实验技术", "行政学工", "教学科研", "工程技术",
    "专业教师", "中心教师", "导师介绍", "导师简介",
    # 教学/培养
    "本科教学", "本科生", "研究生", "硕士生", "博士生",
    "培养方案", "教学计划", "课程简介", "实践教学",
    "教学互动", "学位管理", "专业学位",
    "教学动态", "精品课程", "出版教材",
    "教学大纲", "课程设置", "招生培养",
    "规定细则", "英文硕士", "双学位",
    "考核开题", "学位工作",
    # 科研/学术
    "研究方向", "科学研究", "学术研究", "学术交流",
    "科研项目", "科研奖励", "科研成果",
    "科研动态", "科研信息", "科研平台",
    "创新团队", "科研机构", "研究机构",
    "实验室", "科技奖励", "成果转化", "研究基地",
    "学术动态", "学术讲座", "学术前沿",
    "学术活动", "学术组织", "学术成果",
    "学术科研", "讲座信息",
    # 新闻/通知
    "新闻动态", "新闻中心", "通知公告", "通知通告",
    "最新通知", "讲座信息", "学术动态",
    # 党群
    "党建工作", "党群工作", "党群建设",
    "支部建设", "支部活动", "团建工作",
    "团委工作", "群团工作", "工会工作",
    "师德师风", "党员发展", "党员管理",
    "学习教育", "党史馆",
    # 学生
    "学生天地", "学生园地", "学生工作",
    "学工园地", "学生管理", "学工布告",
    "学生活动", "学子风采",
    "就业信息", "就业创业", "升学就业",
    "活动掠影", "多彩地科",
    # 校友
    "校友天地", "校友之窗", "校友风采", "校友名录",
    "校友动态", "校友资讯", "校友捐赠", "校友活动",
    "院友风采", "院友动态", "院友之家", "院友天地",
    "校友作品", "岁月留痕", "院庆专栏", "百年院庆",
    "校友会", "发展基金", "校友登记",
    # 行政/管理
    "人事", "教务", "科研", "财务", "网络", "资产",
    "安全", "外事", "大楼", "其他",
    "本科教务", "研究生院", "教务处", "财务处",
    "科技处", "国际处", "本科生院",
    "行政部门", "办公部门",
    # 下载/资源
    "下载专区", "下载中心", "文档下载",
    "资料汇编", "学习资料",
    # 招聘/招生
    "诚聘英才", "人才招聘", "招生招聘",
    "招生信息", "招生简章", "招生报名",
    "报名方式", "学院招聘",
    # 国际交流
    "国际交流", "国际合作", "交流概况",
    "科研合作", "学生交换", "合作联盟",
    "历届论坛",
    # 安全
    "安全园地", "安全规章", "安全知识", "安全工作",
    # 财务/报销
    "财务报销", "资产建账", "综合办公", "聘用指南",
    # 其他
    "联系我们", "友情链接", "网站地图", "热点链接",
    "教师登录", "登录系统", "后台管理",
    "中心项目", "职业发展", "主题展览",
    "教学成果", "教学展览", "教学年鉴", "教学论文",
    "教学平台", "毕业设计", "优秀校友", "教师作品",
    "乡村振兴", "云上南雍", "创新项目", "奖助学金",
    "爱国卫生", "运动活动",
    # 公示
    "公示", "出国公示", "人事动态", "组织动态",
    # 经费/资金
    "办事指南", "试剂采购",
    # 英文
    "English", "中文", "英文",
    "MORE", "more", "More",
    # 联系方式
    "院长信箱", "师德监督", "微服务", "文献资源",
    "数据库", "本院著作", "院图馆藏", "院长致辞",
    # 行政部门
    "党委", "党委行政", "行政",
    # 学科/系名
    "生物系", "生化系", "生理系", "生态系",
    "水科学系", "地球化学", "地质工程",
    "无机化学", "分析化学", "有机化学", "物理化学",
    "高分子化学", "生物化学", "化学工程",
    "关于我们", "中心简介",
    # 更多导航
    "工作动态", "通知新闻", "成果推送",
    # ⭐ 截断的导航文本（原始数据中的问题项）
    "常见问", "专职科", "专职科研", "准聘助", "准聘助理",
    "在职教", "在职教师", "数字与", "数字经",
    "教学科", "跨学科", "兼职教授", "兼职教师",
    "访问教授", "讲座教授", "名誉教授",
    "李涛教", "王小教", "张三教",  # 截断的"姓名+教授"
    # ⭐ 更多漏网的非人名
    "教研室", "教职工", "党团工作", "南京大学",
    "常见问", "更多", "全部", "查看", "详情",
    "经费", "资产", "设备", "实验室安全",
}

# 部门/公共邮箱特征（这些邮箱要清空，不是个人邮箱）
DEPARTMENT_EMAIL_PATTERNS = [
    r"^nju[a-z]*@nju\.edu\.cn$",  # njudz@, njuxx@
    r"^[a-z]{2,4}@nju\.edu\.cn$",  # bnhy@ (部门缩写)
    r"^[a-z]+bgs@",  # 办公室
    r"@nbu\.edu\.cn$",  # 宁波大学？不是南大
    r"@163\.com$",  # 非学校邮箱
    r"@126\.com$",  # 非学校邮箱
    r"@hotmail\.com$",
    r"@qq\.com$",
]

def is_department_email(email: str) -> bool:
    """判断是否为部门公共邮箱（非个人邮箱）。"""
    if not email:
        return False
    for pattern in DEPARTMENT_EMAIL_PATTERNS:
        if re.match(pattern, email, re.IGNORECASE):
            return True
    return False

def is_valid_person_name(name: str) -> bool:
    """判断文本是否为有效的中文人名（而非导航文本）。"""
    name = name.strip()
    # 在黑名单中 → 否
    if name in NON_NAME_BLACKLIST:
        return False
    # 必须为2-4个汉字
    if not re.match(r'^[一-鿿]{2,4}$', name):
        return False
    # 包含导航关键词 → 否
    nav_chars = {
        '首页', '学院', '系', '院', '教学', '科研', '学术', '师资',
        '教师', '学生', '校友', '新闻', '通知', '公告', '招生', '培养',
        '学位', '管理', '就业', '党建', '党群', '党委', '工会', '团委',
        '行政', '下载', '资源', '资料', '国际', '交流', '合作', '招聘',
        '财务', '网络', '安全', '采购', '联系', '登录', '注册', '简介',
        '概况', '介绍', '设置', '导航', '地图', '链接', '友情', '版权',
        '中心', '项目', '论坛', '讲座', '动态', '公报', '布告', '公示',
        '规章', '制度', '细则', '方案', '计划', '大纲', '课程', '精品',
        '培育', '创新', '实践', '实验', '基地', '平台', '实验室', '机构',
        '专栏', '展览', '年鉴', '论文', '成果', '奖励', '基金',
        '办公', '部门', '组织', '队伍', '名录', '主页', '主题',
        '地科', '多彩', '爱国', '云上', '乡村', '试剂',
        '办事', '指南', '监督', '信箱',
    }
    for nc in nav_chars:
        if nc in name:
            return False
    return True

def is_public_email(email: str) -> bool:
    """检查是否明显是公共/部门邮箱。"""
    if not email:
        return False
    email_lower = email.lower()
    # 常见公共邮箱前缀
    public_prefixes = {
        "wxyxz", "xwcb", "bgs", "dangzheng", "office", "yuanban",
        "webmaster", "admin", "info", "master", "root", "postmaster",
        "bgsdz", "ybdz", "hdb", "jzcg", "njudz", "bnhy",
    }
    local = email_lower.split("@")[0]
    for prefix in public_prefixes:
        if local.startswith(prefix):
            return True
    # 小于3个字母的纯字母前缀通常也是部门邮箱
    if len(local) <= 3 and local.isalpha() and "@nju.edu.cn" in email_lower:
        return True
    return False


def main():
    # ——— 1. 读取原始完整版 ———
    existing_records = []
    with open(EXISTING_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            existing_records.append({
                "name": row["姓名"].strip(),
                "email": row["邮箱"].strip(),
                "department": row["学院"].strip(),
                "title": row["职称"].strip(),
                "url": row["主页链接"].strip(),
            })

    existing_departments = set(r["department"] for r in existing_records)
    print(f"原始数据: {len(existing_records)} 条, {len(existing_departments)} 个院系")

    # ——— 2. 读取新爬取数据 ———
    new_records = []
    with open(NEW_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            new_records.append({
                "name": row["姓名"].strip(),
                "email": row["邮箱"].strip(),
                "department": row["学院"].strip(),
                "title": row["职称"].strip(),
                "url": row["主页链接"].strip(),
            })

    print(f"新爬取数据: {len(new_records)} 条")

    # ——— 2b. 读取最后3院系数据 ———
    last3_records = []
    if LAST3_CSV.exists():
        with open(LAST3_CSV, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                last3_records.append({
                    "name": row["姓名"].strip(),
                    "email": row["邮箱"].strip(),
                    "department": row["学院"].strip(),
                    "title": row["职称"].strip(),
                    "url": row["主页链接"].strip(),
                })
        print(f"最后3院系数据: {len(last3_records)} 条")

    # ——— 3. 合并新数据（14院系 + 最后3院系） ———
    all_new = new_records + last3_records
    cleaned_new = []
    new_stats = {"total": len(all_new), "bad_name": 0, "kept": 0, "email_cleared": 0}

    for r in all_new:
        if not is_valid_person_name(r["name"]):
            new_stats["bad_name"] += 1
            continue

        if r["email"] and (is_public_email(r["email"]) or is_department_email(r["email"])):
            print(f"  清除公共邮箱: {r['name']} <{r['email']}> ({r['department']})")
            r["email"] = ""
            new_stats["email_cleared"] += 1

        cleaned_new.append(r)
        new_stats["kept"] += 1

    print(f"\n新数据过滤: {new_stats['total']} → {new_stats['kept']} (移除{new_stats['bad_name']}条非人名, 清除{new_stats['email_cleared']}条公共邮箱)")

    # ——— 4. 合并（先用新数据补充现有数据中无邮箱的记录） ———
    existing_urls = set(r["url"] for r in existing_records if r["url"])

    truly_new = [r for r in cleaned_new if r["url"] not in existing_urls]

    updated = 0
    for new_r in cleaned_new:
        if new_r["url"] in existing_urls and new_r["email"]:
            for existing_r in existing_records:
                if existing_r["url"] == new_r["url"] and not existing_r["email"]:
                    existing_r["email"] = new_r["email"]
                    existing_r["title"] = new_r["title"] or existing_r["title"]
                    updated += 1
                    print(f"  补充邮箱: {new_r['name']} <{new_r['email']}> ({new_r['department']})")
                    break

    print(f"\n真正新增: {len(truly_new)} 条, 补充邮箱: {updated} 条")

    # 合并
    all_records = existing_records + truly_new

    # ——— 5. ⭐ 对全部数据做最终清洗 ———
    before_all = len(all_records)
    cleaned_all = []
    all_bad = 0

    for r in all_records:
        if not is_valid_person_name(r["name"]):
            all_bad += 1
            continue
        cleaned_all.append(r)

    all_records = cleaned_all
    print(f"\n全量清洗: {before_all} → {len(all_records)} (移除{all_bad}条非人名记录)")

    # 按学院排序
    all_records.sort(key=lambda r: (r["department"], r["name"]))

    # ——— 6. 统计 ———
    dept_stats = defaultdict(lambda: {"total": 0, "with_email": 0})
    for r in all_records:
        d = r["department"]
        dept_stats[d]["total"] += 1
        if r["email"]:
            dept_stats[d]["with_email"] += 1

    print(f"\n=== 最终合并结果 ({len(all_records)}条, {len(dept_stats)}个院系) ===")
    for dept in sorted(dept_stats):
        s = dept_stats[dept]
        pct = s["with_email"] / s["total"] * 100 if s["total"] else 0
        flag = "✅" if s["with_email"] >= 5 else "⚠️"
        print(f"  {flag} {dept}: {s['total']}人, {s['with_email']}邮箱 ({pct:.0f}%)")

    # ——— 7. 导出 ———
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = OUTPUT_DIR / f"南京大学_教师邮箱_合并版_{ts}.csv"

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(all_records, 1):
            writer.writerow([
                i,
                r["name"],
                r["email"],
                r["department"],
                r["title"],
                r["url"],
            ])

    print(f"\n最终CSV: {csv_path}")
    print(f"共 {len(all_records)} 位教师, {sum(1 for r in all_records if r['email'])} 个有邮箱")

    # 同时输出XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        xlsx_path = OUTPUT_DIR / f"南京大学_教师邮箱_合并版_{ts}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "教师邮箱"

        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        cell_font = Font(name="微软雅黑", size=10)
        cell_align = Alignment(vertical="center")

        headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, r in enumerate(all_records, 1):
            values = [i, r["name"], r["email"], r["department"], r["title"], r["url"]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 55

        wb.save(xlsx_path)
        print(f"最终XLSX: {xlsx_path}")
    except Exception as e:
        print(f"XLSX导出失败: {e}")

    # 显示新增的记录样本
    print(f"\n=== 新增记录样本 ===")
    for r in truly_new[:15]:
        print(f"  {r['name']:6s} | {r['email'] or '(无)':30s} | {r['department']:20s} | {r['title']}")


if __name__ == "__main__":
    main()
