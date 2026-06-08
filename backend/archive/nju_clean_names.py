"""南京大学数据姓名清洗脚本"""
import csv
import re
import os
import shutil

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # backend/

SRC = "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.3.csv"
SRC_XLSX = "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.3.xlsx"
DST = "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.4.csv"
DST_XLSX = "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.4.xlsx"
OLD_FILES = [
    "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.3.csv",
    "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.3.xlsx",
    "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.2.csv",
    "outputs/nju_final_20260603_131944/南京大学_全部教师邮箱_V1.0.2.xlsx",
    "outputs/nju_final_20260603_131944/南京大学_仅邮箱教师_V1.0.2.csv",
    "outputs/nju_final_20260603_131944/南京大学_仅邮箱教师_V1.0.2.xlsx",
    "outputs/nju_final_20260603_131931/南京大学_全部教师邮箱_V1.0.2.csv",
    "outputs/nju_final_20260603_131931/南京大学_仅邮箱教师_V1.0.2.csv",
    "outputs/南京大学_增量爬取_20260603_131526.csv",
    "outputs/南京大学_增量爬取_20260603_131903.csv",
]

# 导航/非人名的脏关键词（出现在姓名中的）
NAV_KEYWORDS = [
    "概况", "新闻", "通知", "公告", "招生", "培养", "就业", "学位", "学科",
    "科研", "学术", "党建", "工会", "校友", "捐赠", "图书馆", "校园", "地图",
    "网站", "登录", "邮箱", "联系我们", "欢迎", "首页", "返回", "更多", "详情",
    "查看", "下载", "师资", "教师", "博士", "硕士", "本科", "研究", "行政",
    "管理", "教职", "荣休", "访问", "系科", "教研", "诚聘", "信箱",
    "师德师", "师资队", "现任教", "学院概", "学院", "管理架", "系科设", "教研机",
    "教授", "副教", "长聘", "准聘", "助理", "讲师", "秘书", "办公", "公共",
    "机关", "校内", "关于", "本系", "本院", "本学院", "集群", "总览", "一览",
    "标识", "地址", "简介", "简报", "寄语", "风采", "视点", "百年", "地址",
    "印章", "标志", "贯彻", "发展", "党政", "寄语", "公众",
    "与公共", "与学院", "关学院", "学学院", "科学院", "义学院",
    "究生", "本科生", "商学院", "数学院", "慈善", "南赫", "院内", "院系",
    "博士后", "基地", "平台", "重点", "实验", "示范", "特色", "品牌", "课程",
    "资源", "开放", "国际", "合作", "交流", "社会", "服务", "实验", "队伍",
    "建设", "机构", "设置", "学生", "创新", "创业", "竞赛", "成果", "获奖",
    "专利", "论文", "著作", "教材", "项目", "政府", "中心",
    "教务", "学工", "团委", "党总支", "党委", "院长", "书记", "主任", "科长",
    "主管", "全部", "所有", "共计", "合计", "总数", "教师信息", "人员信息",
    "现任", "兼职", "信息", "用户", "简介", "搜索", "咨询", "指南", "链接",
    "版权", "所有", "声明", "条件", "条款", "隐私", "安全", "帮助", "支持",
    "招聘", "诚聘", "招贤", "纳士", "招人", "加入", "联系",
    "专业", "方向", "领域", "前沿", "动态", "进展", "报告", "讲座", "论坛",
    "研讨会", "会议", "征集", "征稿", "投稿", "订阅", "关注", "分享",
    "友情", "相关", "常用", "快速", "便捷", "推荐",
    "自然科学", "社会科学", "人文", "理学", "工学", "医学",
    "数位", "网上", "远程", "在线", "智慧", "智能", "数字",
    "体育", "艺术", "音乐", "美术", "设计",
    "教学", "教务", "学籍", "成绩", "评估", "督导",
    "信息中心", "网络", "技术", "维护", "系统", "数据",
    "基金", "经费", "财务", "资产", "设备", "采购",
    "后勤", "保卫", "房产", "基建", "维修", "餐饮",
    "人事", "人才", "薪酬", "福利", "考核", "培训",
    "外事", "国际", "港澳", "台湾", "留学", "孔子",
    "出版", "期刊", "学报", "杂志", "编辑", "发行",
    "支部", "总支", "党组", "统战", "纪检", "监察", "审计",
    "巡视", "巡查", "督查", "督办",
    "退休", "离休", "退协", "关工委", "老龄",
    "集团", "公司", "企业", "产业", "科技园",
    "中学", "小学", "幼儿园", "附中", "附小",
    "馆藏", "流通", "借阅", "文献", "数据库", "检索",
    "公告", "公示", "通告", "通知", "启事",
    "学术报告", "学术讲座", "学术活动", "学术会议",
    "值班", "作息", "校历", "班车", "交通",
    "年鉴", "大事记", "历史", "沿革",
]

# 姓名黑名单（完整匹配的非人名）
NAME_BLACKLIST = [
    "人才培养", "专业学位", "专业教师", "专任教师", "专职教师", "专职研究", "专职科研",
    "了解详情", "交流培养", "业校友会", "中国研究", "主要研究", "产权研究", "人口研究",
    "信息管理", "体育科研", "优秀校友", "值班安排", "全部教师", "全部职位",
    "兼职教师", "兼职教授", "博士后", "学院概况", "学院介绍", "学院新闻",
    "学院通知", "学院公告", "学院简介", "学生工作", "学生管理", "学生服务",
    "学生活动", "学生风采", "学术研究", "学术成果", "学术动态", "学术交流",
    "学术活动", "学术报告", "学术讲座", "导师信息", "导师队伍", "导师名单",
    "招生信息", "招生就业", "招生通知", "招生公告", "教学管理", "教学科研",
    "教学工作", "教学成果", "教务管理", "教师队伍", "教师名录", "教师信息",
    "教师首页", "教师介绍", "教师风采", "教师招聘", "教研机构", "教研组",
    "教育科研", "教育管理", "教职员工", "教职队伍", "数据研究", "文物研究",
    "新闻公告", "新闻动态", "新闻通知", "服务指南", "服务机构", "查阅服务",
    "欢迎光临", "欢迎访问", "欢迎您", "下载专区", "下载中心", "研究中心",
    "研究团队", "研究方向", "研究领域", "研究成果", "科研项目", "科研团队",
    "科研方向", "科研成果", "科研平台", "科研基地", "科研机构",
    "管理科学", "管理系统", "管理办法", "管理条例", "管理团队",
    "管理服务", "管理学院", "组织机构", "组织架构", "党建工作",
    "党务公开", "党委工作", "党委通知", "党派团体", "退休教师",
    "退协工作", "通知公告", "人才培养", "人文研究", "人文社科",
    "优秀人才", "优质资源", "信息资源", "信息公告", "信息公开",
    "信息门户", "信息服务", "信息中心", "信息技术",
    "国际交流", "国际合作", "国际教育", "国际学院",
    "图书信息", "图书馆", "数据资源", "数据研究",
    "文化建设", "文化研究", "文学研究", "文学教育",
    "新教工", "新教师", "新进教师", "新进教工",
    "师资队伍", "师资建设", "师资介绍", "师资概况",
    "物理研究", "环境研究", "现代管理", "现任领导",
    "现任教师", "品牌专业", "品牌课程", "品牌特色",
    "博士后流动站", "学位授权", "学位点", "学科建设",
    "学科带头人", "学科方向", "学科平台", "实验教学",
    "实验中心", "实验室", "实习实践", "实习基地",
    "实训基地", "实践教学", "实践基地",
    "技术研究", "技术创新", "技术支持", "技术服务",
    "招生信息", "招生简章", "招生政策", "招生专业",
    "拔尖人才", "创新人才", "创新团队", "创新创业",
    "特色专业", "特色课程", "特色学科",
]

# 有效邮箱域名（南大相关）
NJU_DOMAINS = ["nju.edu.cn", "sina.com"]


def is_valid_name(name):
    """判断是否为有效教师姓名"""
    name = name.strip()
    if not name:
        return False
    # 完整匹配黑名单
    if name in NAME_BLACKLIST:
        return False
    # 包含脏关键词
    for kw in NAV_KEYWORDS:
        if kw in name:
            return False
    # 必须是2-6个中文汉字
    if not re.match(r'^[一-鿿]{2,6}$', name):
        return False
    return True


def main():
    # 读取
    with open(SRC, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    print(f"原始数据: {len(rows)} 条")

    # 分类
    good = []
    removed = []
    for row in rows:
        name = row["姓名"].strip()
        email = row["邮箱"].strip()
        if is_valid_name(name):
            good.append(row)
        else:
            removed.append(row)
            reason = "黑名单匹配" if name in NAME_BLACKLIST else "含脏关键词"
            print(f"  移除: 姓名=[{name}] 邮箱=[{email}] 学院=[{row['学院']}] ({reason})")

    print(f"\n保留: {len(good)} 条")
    print(f"移除: {len(removed)} 条")

    # 去重（仅移除完全重复行：同名+同邮箱）
    seen = set()
    final = []
    dup_removed = 0
    for row in good:
        name = row["姓名"].strip()
        email = row["邮箱"].strip()
        key = f"{name}|{email}"
        if key in seen:
            dup_removed += 1
        else:
            seen.add(key)
            final.append(row)

    print(f"完全重复（同名+同邮箱）移除: {dup_removed} 条")
    print(f"最终: {len(final)} 条")

    # 写入新CSV
    fieldnames = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    with open(DST, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for i, row in enumerate(final, 1):
            writer.writerow({
                "序号": i,
                "姓名": row["姓名"].strip(),
                "邮箱": row["邮箱"].strip(),
                "学院": row["学院"].strip(),
                "职称": row["职称"].strip(),
                "主页链接": "",
            })

    print(f"\n清洗完成: {DST}")

    # 生成xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "教师邮箱"

        # 表头样式
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # 写表头
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写数据
        cell_font = Font(name="微软雅黑", size=10)
        cell_align = Alignment(vertical="center")
        for i, row in enumerate(final, 2):
            for col, field in enumerate(fieldnames, 1):
                val = row[field].strip() if field in row else ""
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        # 自适应列宽
        for col in range(1, len(fieldnames) + 1):
            max_len = len(str(ws.cell(row=1, column=col).value)) * 2
            for row_idx in range(2, min(len(final) + 2, 200)):
                cell_val = str(ws.cell(row=row_idx, column=col).value or "")
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[chr(64 + col)].width = min(max_len + 4, 50)

        wb.save(DST_XLSX)
        print(f"XLSX 已生成: {DST_XLSX}")
    except ImportError:
        print("openpyxl 未安装，跳过 XLSX 生成")

    # 删除旧文件（忽略正在使用中的错误）
    for fpath in OLD_FILES:
        full = os.path.join(BASE, fpath) if not os.path.isabs(fpath) else fpath
        full = os.path.normpath(full)
        if os.path.exists(full):
            try:
                os.remove(full)
                print(f"已删除旧文件: {full}")
            except PermissionError:
                print(f"⚠️ 旧文件正在使用中，无法删除: {full}")
            except OSError as e:
                print(f"⚠️ 无法删除旧文件 {full}: {e}")
        else:
            print(f"旧文件不存在(跳过): {full}")


if __name__ == "__main__":
    main()
