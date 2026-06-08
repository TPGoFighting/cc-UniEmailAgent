"""南京理工大学计算机学院教师信息补全脚本。"""

import csv
import re
import time
import logging
from pathlib import Path
from datetime import datetime

import requests
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
}

EMAIL_PATTERN = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")


def parse_at_sign(text: str) -> str:
    """恢复反爬邮箱。"""
    text = re.sub(r"\s*\[at\]\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*\(at\)\s*", "@", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*#@\s*", "@", text)
    text = re.sub(r"\s*\[@\]\s*", "@", text)
    text = re.sub(r"\s*\(@\)\s*", "@", text)
    return text


def extract_emails(text: str) -> list[str]:
    """从文本中提取邮箱地址。"""
    return list(set(EMAIL_PATTERN.findall(text)))


def extract_title_from_profile(text: str) -> str:
    """从教师详情页文本中提取职称。"""
    title_keywords = [
        "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
        "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
        "长江学者", "杰青", "优青",
    ]
    for title in title_keywords:
        if title in text:
            return title
    return ""


def is_valid_email(email: str) -> bool:
    """校验邮箱格式。"""
    return bool(re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email))


def is_admin_email(email: str) -> bool:
    """检测是否为学院公共邮箱。"""
    email_lower = email.lower()
    admin_prefixes = [
        "webmaster", "admin", "office", "info", "master", "root",
        "postmaster", "wxyxz", "xwcb", "bgs", "dangzheng", "yuanban",
    ]
    for prefix in admin_prefixes:
        if email_lower.startswith(prefix + "@"):
            return True
    return False


def scrape_teacher_page(url: str, session: requests.Session) -> dict:
    """访问教师个人主页，提取邮箱和职称。"""
    result = {"email": "", "title": ""}
    try:
        r = session.get(url, timeout=15)
        r.encoding = "utf-8"
        text = r.text
        soup = BeautifulSoup(text, "html.parser")
        body_text = soup.get_text()
        body_text = parse_at_sign(body_text)

        # 提取邮箱
        emails = extract_emails(body_text)
        valid_emails = [e for e in emails if is_valid_email(e) and not is_admin_email(e)]
        if valid_emails:
            result["email"] = valid_emails[0]

        # 提取职称
        title = extract_title_from_profile(body_text)
        result["title"] = title

        if result["email"] or result["title"]:
            logger.info(f"  → {url.split('/')[-3]}: email={'✅' if result['email'] else '❌'} title={'✅' if result['title'] else '❌'}")
        return result
    except Exception as e:
        logger.debug(f"  ✗ {url}: {e}")
        return result


def main():
    base_dir = Path(__file__).parent.parent / "outputs"
    input_file = base_dir / "南京理工大学_计算机学院_教师邮箱_清洗版_20260527_112043.csv"

    # 读取记录
    records = []
    with open(input_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            records.append({
                "name": row.get("姓名", "").strip(),
                "email": row.get("邮箱", "").strip(),
                "department": row.get("学院", "").strip() or "计算机科学与工程学院",
                "title": row.get("职称", "").strip(),
                "url": row.get("主页链接", "").strip(),
            })

    logger.info(f"共 {len(records)} 位教师")
    logger.info(f"已有邮箱: {sum(1 for r in records if r['email'])}")
    logger.info(f"已有职称: {sum(1 for r in records if r['title'])}")
    logger.info(f"缺少邮箱: {sum(1 for r in records if not r['email'])}")
    logger.info(f"缺少职称: {sum(1 for r in records if not r['title'])}")

    session = requests.Session()
    session.headers.update(HEADERS)

    # 访问每位教师的个人主页
    updated = 0
    for i, record in enumerate(records):
        if not record["url"]:
            continue

        # 如果已有完整信息，跳过
        has_email = bool(record["email"])
        has_title = bool(record["title"])
        if has_email and has_title:
            continue

        logger.info(f"[{i+1}/{len(records)}] {record['name']}: 缺{'邮箱 ' if not has_email else ''}{'职称 ' if not has_title else ''}")

        info = scrape_teacher_page(record["url"], session)

        if info["email"] and not has_email:
            record["email"] = info["email"]
            updated += 1
            logger.info(f"  ✅ 补全邮箱: {info['email']}")

        if info["title"] and not has_title:
            record["title"] = info["title"]
            updated += 1
            logger.info(f"  ✅ 补全职称: {info['title']}")

        # 适当延时，避免被 ban
        if i % 10 == 9:
            time.sleep(1)

    logger.info(f"\n补全完成！共更新 {updated} 个字段")
    logger.info(f"最终统计: 共 {len(records)} 位教师")
    logger.info(f"有邮箱: {sum(1 for r in records if r['email'])}")
    logger.info(f"有职称: {sum(1 for r in records if r['title'])}")

    # 去重（按邮箱，空邮箱也保留但只保留一份）
    seen_emails = set()
    seen_names = set()
    deduped = []
    for r in records:
        email_key = r["email"] if r["email"] else f"__noemail__{r['name']}"
        name_key = r["name"]
        if email_key not in seen_emails:
            seen_emails.add(email_key)
            seen_names.add(name_key)
            deduped.append(r)

    logger.info(f"去重后: {len(deduped)} 位教师")

    # 生成输出文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_csv = base_dir / f"南京理工大学_计算机学院_教师邮箱_完整版_{timestamp}.csv"

    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(deduped, 1):
            writer.writerow([i, r["name"], r["email"], r["department"], r["title"], r["url"]])

    logger.info(f"CSV 已保存: {output_csv}")

    # 生成 XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        output_xlsx = base_dir / f"南京理工大学_计算机学院_教师邮箱_完整版_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "南京理工大学计算机学院教师"

        # 表头样式
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据样式
        data_font = Font(name="微软雅黑", size=10)
        data_align = Alignment(vertical="center")
        email_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

        for i, r in enumerate(deduped, 1):
            ws.cell(i + 1, 1, i).font = data_font
            ws.cell(i + 1, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(i + 1, 2, r["name"]).font = data_font
            ws.cell(i + 1, 2).alignment = data_align
            email_cell = ws.cell(i + 1, 3, r["email"])
            email_cell.font = email_font if r["email"] else data_font
            email_cell.alignment = data_align
            ws.cell(i + 1, 4, r["department"]).font = data_font
            ws.cell(i + 1, 4).alignment = data_align
            ws.cell(i + 1, 5, r["title"]).font = data_font
            ws.cell(i + 1, 5).alignment = data_align
            url_cell = ws.cell(i + 1, 6, r["url"])
            url_cell.font = email_font if r["url"] else data_font
            url_cell.alignment = data_align

            for col in range(1, 7):
                ws.cell(i + 1, col).border = thin_border

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 45

        wb.save(output_xlsx)
        logger.info(f"XLSX 已保存: {output_xlsx}")
    except ImportError:
        logger.warning("openpyxl 未安装，跳过 XLSX 导出")

    # 统计摘要
    with_email = [r for r in deduped if r["email"]]
    without_email = [r for r in deduped if not r["email"]]

    logger.info("\n" + "=" * 60)
    logger.info(f"📊 统计摘要")
    logger.info(f"   教师总数: {len(deduped)}")
    logger.info(f"   有邮箱: {len(with_email)} ({len(with_email)/len(deduped)*100:.1f}%)")
    logger.info(f"   无邮箱: {len(without_email)} ({len(without_email)/len(deduped)*100:.1f}%)")
    logger.info(f"   有职称: {sum(1 for r in deduped if r['title'])}")
    logger.info(f"   无职称: {sum(1 for r in deduped if not r['title'])}")

    if without_email:
        logger.info(f"\n   无邮箱教师:")
        for r in without_email:
            logger.info(f"     - {r['name']} ({r['title'] or '职称未知'})")


if __name__ == "__main__":
    main()
