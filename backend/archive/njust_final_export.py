"""南京理工大学计算机学院教师信息 —— 最终完整版导出脚本。

合并已有数据和搜索到的新教师信息，生成完整的 CSV/XLSX。
"""

import csv
import re
import logging
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# 新发现的教师（搜索补充）
NEW_TEACHERS = [
    {"name": "杨杨", "email": "yyang@njust.edu.cn", "title": "教授", "url": ""},
    {"name": "张姗姗", "email": "shanshan.zhang@njust.edu.cn", "title": "教授", "url": ""},
    {"name": "陈莹", "email": "ychen@njust.edu.cn", "title": "教授", "url": ""},
    {"name": "夏睿", "email": "rxia@njust.edu.cn", "title": "教授", "url": ""},
    {"name": "贾修一", "email": "jiaxy@njust.edu.cn", "title": "教授", "url": ""},
    {"name": "胡雪蕾", "email": "xlhu@njust.edu.cn", "title": "副教授", "url": ""},
    {"name": "王树梅", "email": "hwasm@mail.njust.edu.cn", "title": "教授", "url": ""},
    {"name": "孙兴华", "email": "xinghuasun@mail.njust.edu.cn", "title": "副教授", "url": ""},
    {"name": "王玲", "email": "songln@mail.njust.edu.cn", "title": "副教授", "url": ""},
    {"name": "邹建伟", "email": "Zjwmail2000@163.com", "title": "讲师", "url": ""},
    {"name": "叶庆生", "email": "yeqingsh@mail.njust.edu.cn", "title": "副教授", "url": ""},
    {"name": "杨静宇", "email": "yangjy@mail.njust.edu.cn", "title": "教授", "url": ""},
    {"name": "沈思", "email": "shensi@njust.edu.cn", "title": "副教授", "url": ""},
]


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", email.strip()))


def main():
    base_dir = Path(__file__).parent.parent / "outputs"
    input_file = base_dir / "南京理工大学_计算机学院_教师邮箱_清洗版_20260527_112043.csv"

    # 读取已有记录
    records = []
    existing_names = set()
    existing_emails = set()
    with open(input_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            r = {
                "name": row.get("姓名", "").strip(),
                "email": row.get("邮箱", "").strip(),
                "department": row.get("学院", "").strip() or "计算机科学与工程学院",
                "title": row.get("职称", "").strip(),
                "url": row.get("主页链接", "").strip(),
            }
            records.append(r)
            existing_names.add(r["name"])
            if r["email"]:
                existing_emails.add(r["email"])

    logger.info(f"已有数据: {len(records)} 位教师")

    # 添加新教师（排除已存在的）
    added = 0
    for t in NEW_TEACHERS:
        if t["name"] in existing_names:
            # Check if existing already has this email
            existing = [r for r in records if r["name"] == t["name"]]
            if existing and not existing[0]["email"] and t["email"]:
                # Update the email
                existing[0]["email"] = t["email"]
                logger.info(f"  ✏️ 补全 {t['name']} 邮箱: {t['email']}")
                added += 1
            elif existing and not existing[0]["title"] and t["title"]:
                existing[0]["title"] = t["title"]
                logger.info(f"  ✏️ 补全 {t['name']} 职称: {t['title']}")
                added += 1
            continue
        records.append({
            "name": t["name"],
            "email": t["email"],
            "department": "计算机科学与工程学院",
            "title": t["title"],
            "url": t.get("url", ""),
        })
        existing_names.add(t["name"])
        added += 1
        logger.info(f"  ➕ 添加 {t['name']}: {t['email']} ({t['title']})")

    logger.info(f"新增/补全: {added} 条")

    # 最终去重（按邮箱，空邮箱保留但去重）
    seen_emails = set()
    seen_names = set()
    deduped = []
    for r in records:
        email_key = r["email"] if r["email"] else f"__noemail__{r['name']}"
        name_key = r["name"]
        if name_key not in seen_names:
            seen_names.add(name_key)
            seen_emails.add(email_key)
            deduped.append(r)
        elif name_key not in [d["name"] for d in deduped]:
            # Shouldn't happen but just in case
            deduped.append(r)

    logger.info(f"去重后: {len(deduped)} 位教师")

    # 统计
    with_email = [r for r in deduped if r["email"] and is_valid_email(r["email"])]
    without_email = [r for r in deduped if not r["email"] or not is_valid_email(r["email"])]
    with_title = [r for r in deduped if r["title"]]

    logger.info(f"\n📊 统计摘要:")
    logger.info(f"   教师总数: {len(deduped)}")
    logger.info(f"   有邮箱: {len(with_email)} ({len(with_email)/len(deduped)*100:.1f}%)")
    logger.info(f"   无邮箱: {len(without_email)} ({len(without_email)/len(deduped)*100:.1f}%)")
    logger.info(f"   有职称: {len(with_title)} ({len(with_title)/len(deduped)*100:.1f}%)")

    # 按姓名排序
    deduped.sort(key=lambda r: r["name"])

    # 生成文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # === CSV ===
    output_csv = base_dir / f"南京理工大学_计算机学院_教师邮箱_完整版_{timestamp}.csv"
    with open(output_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(deduped, 1):
            writer.writerow([i, r["name"], r["email"], r["department"], r["title"], r["url"]])
    logger.info(f"✅ CSV: {output_csv}")

    # === 仅有邮箱的版本 ===
    output_email_csv = base_dir / f"南京理工大学_计算机学院_有邮箱教师_{timestamp}.csv"
    with open(output_email_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(with_email, 1):
            writer.writerow([i, r["name"], r["email"], r["department"], r["title"], r["url"]])
    logger.info(f"✅ 仅有邮箱 CSV: {output_email_csv} ({len(with_email)} 人)")

    # === XLSX ===
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        for suffix, data, label in [
            ("完整版", deduped, "全部教师"),
            ("有邮箱教师", with_email, "有邮箱教师"),
        ]:
            output_xlsx = base_dir / f"南京理工大学_计算机学院_教师邮箱_{suffix}_{timestamp}.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"南京理工大学计算机学院{label}"

            # 表头样式
            header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # 数据样式
            data_font = Font(name="微软雅黑", size=10)
            data_align = Alignment(vertical="center")
            link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

            for i, r in enumerate(data, 1):
                ws.cell(i + 1, 1, i).font = data_font
                ws.cell(i + 1, 1).alignment = Alignment(horizontal="center", vertical="center")

                name_cell = ws.cell(i + 1, 2, r["name"])
                name_cell.font = data_font
                name_cell.alignment = data_align

                email_cell = ws.cell(i + 1, 3, r["email"])
                email_cell.font = link_font if r["email"] else data_font
                email_cell.alignment = data_align

                ws.cell(i + 1, 4, r["department"]).font = data_font
                ws.cell(i + 1, 4).alignment = data_align

                title_cell = ws.cell(i + 1, 5, r["title"])
                title_cell.font = data_font
                title_cell.alignment = data_align

                url_cell = ws.cell(i + 1, 6, r["url"])
                url_cell.font = link_font if r["url"] else data_font
                url_cell.alignment = data_align

                for col in range(1, 7):
                    ws.cell(i + 1, col).border = thin_border

            # 列宽
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 32
            ws.column_dimensions["D"].width = 22
            ws.column_dimensions["E"].width = 14
            ws.column_dimensions["F"].width = 48

            # 冻结首行
            ws.freeze_panes = "A2"

            wb.save(output_xlsx)
            logger.info(f"✅ XLSX {suffix}: {output_xlsx}")

    except ImportError:
        logger.warning("openpyxl 未安装，跳过 XLSX 导出")

    # 列出无邮箱教师
    if without_email:
        logger.info(f"\n📋 无邮箱教师 ({len(without_email)} 人):")
        for r in without_email:
            logger.info(f"   - {r['name']} ({r['title'] or '职称未知'})")

    logger.info("\n🎉 全部完成！")


if __name__ == "__main__":
    main()
