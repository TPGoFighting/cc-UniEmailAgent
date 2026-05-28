"""东南大学教师邮箱爬虫 V2 — 支持两级分类导航。

改进：
  1. 先找子分类链接（教授/副教授/讲师等）
  2. 进入子分类获取教师姓名链接
  3. 访问教师详情页提取邮箱
  4. 支持内联模式（教师信息直接在列表页）
"""
import asyncio
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import Counter

from playwright.async_api import async_playwright

TASK_ID = "bda95480-ec96-4bd7-bc18-05f797e28dd4"
OUTPUT_DIR = Path(__file__).parent / "outputs" / TASK_ID
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
PROGRESS_FILE = OUTPUT_DIR / "seu_progress_v2.json"

PAGE_TIMEOUT = 25000
PROFILE_TIMEOUT = 12000
MAX_TEACHERS_PER_DEPT = 200

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

ADMIN_PREFIXES = {"webmaster","admin","office","info","master","root","postmaster",
    "wxyxz","xwcb","bgs","dangzheng","yuanban","radio","seuradiojob",
    "dzxy","seuem","seutc_official","ysxy","slst","deanoffice_seuarch","bgs","dzz",
    "yjsb","xgk","xsc","jwc","lxyz","bks","cwc","hq","wj","sbc","rsc"}

NAV_WORDS = {
    '首页','概况','新闻','通知','公告','招生','培养','就业','学位','学科','科研','学术',
    '党建','工会','校友','捐赠','图书馆','校园','地图','网站','登录','邮箱','联系我们','欢迎',
    '返回','更多','详情','查看','下载','学院','大学','管理','后台','English','人才引进','人才招聘',
    '院长书记','信箱','相关链接','联系方式','学校首页','学校主页','收藏本站','旧版入口','暑期学校',
    '平湖芳草','下载专区','捐赠通道','院长邮箱','院内文档','日本語','标识系统','院系设置','教师教学',
    '技术转移','海外教育','仪器设备','化工时刊','尾页','网站首页','招生信息','教师登录','现任领导',
    '历任领导','办公电话','院长寄语','组织机构','学科建设','历史沿革','学院简介','学院介绍','组织框架',
    '系所设置','学科组织','本院概况','本院简介','学院简况','学院概述','学院架构','快捷入口','深入贯彻',
    '江苏省','学术论文','专利成果','获奖成果','课程改革','牵头学科','学位管理','出国交流','答辩公示',
    '本科生','研究生','学生工作','党群工作','人才培养','科学研究','人才引进','校友天地','合作交流',
    '诚聘英才','拔尖基地','教学管理','本科生培','研究生培','鲁汶国际','项目介绍','电子信息','规章制度',
    '学生自我','教职工','教师教学','发展中心','教师查询','教师风采','专任教师','院士','客座教授',
    '教师简介','兼职教授','离退休','荣休','知名专家','知名学者','全体教师','硕博导师','各系名单',
    '国家高层次','人才','各系名单','管理入口','师资维护','师资修改','个人中心','师资概况','师资力量',
    '师资队伍','师资概览','师资骨干','教师队伍','杰出人才','人才工程','人才职称','教授风采','教授',
    '副教授','讲师','助教','研究员','副研究员','助理教授','助理研究员','高级工程师','工程师','博士后',
    '博士生导师','硕士生导师','博士导师','硕士导师','博导','硕导','在职教师','退休教师','访问学者',
    '实验室','研究中心','学院部门','联系我们','教务系统','中文','关闭此页','学院主页','艺术学院',
    '标识系统','校内办公','网络教学','校园信息','校园卡','电子邮件','图书馆主页','一卡通','财务查询',
    '校园门户','信息服务','网络服务','VPN服务','正版软件','软件下载','校园网','无线网','邮箱系统',
    '信息门户','网上办事','办事大厅','服务大厅','统一身份','认证','一网通办','系所介绍','声影机械',
    '行政服务','资料下载','培养动态','招生动态','常用下载','教师名录','学系介绍','学院治理','行政机构',
    '两院院士','教育学','走近外院','机构设置','党群组织','行政机构','学术组织','督导组','学术期刊',
    '院务公告','学院领导','学院院标','实验教师','行政人员','退休教工','招聘信息','中大学人','学术研究',
    '科研通知','学术活动','科研项目','科研成果','院长信箱','学院新闻','学术信息','重要通知','教务信息',
    '图片新闻','院训院标','学系','附属医院','信息门户','师资队伍','师资力量','学院概况',
}

# 需要重爬的问题学院，以及额外的子分类URL
PROBLEM_DEPTS = [
    # (学院名, domain, [主列表URL], {额外子分类URLs})
    ("机械工程学院", "me.seu.edu.cn",
     ["https://me.seu.edu.cn/szll/list.htm"],
     ["https://me.seu.edu.cn/xscz/list.htm"]),  # 教师名录
    ("能源与环境学院", "power.seu.edu.cn",
     ["http://power.seu.edu.cn/9216/list.htm"],
     ["http://power.seu.edu.cn/9232/list.htm"]),  # 专任教师
    ("土木工程学院", "civil.seu.edu.cn",
     ["https://civil.seu.edu.cn/10475/list.htm"],
     []),
    ("电子科学与工程学院", "electronic.seu.edu.cn",
     ["http://electronic.seu.edu.cn/szllx/list.htm"],
     ["https://electronic.seu.edu.cn/11478/list.htm"]),  # 导师简介
    ("外国语学院", "sfl.seu.edu.cn",
     ["https://sfl.seu.edu.cn/9851/list.htm"],
     []),
    ("生物科学与医学工程学院", "bme.seu.edu.cn",
     ["https://bme.seu.edu.cn/499/list.htm"],
     []),
    ("化学化工学院", "chem.seu.edu.cn",
     ["https://chem.seu.edu.cn/js/list.htm"],
     []),
    ("仪器科学与工程学院", "ins.seu.edu.cn",
     ["https://ins.seu.edu.cn/45076/list.htm"],
     ["https://ins.seu.edu.cn/zrjs/list.htm"]),  # 全体教师
    ("艺术学院", "arts.seu.edu.cn",
     ["https://arts.seu.edu.cn/szdw_25730/list.htm"],
     ["https://arts.seu.edu.cn/17122/list.htm"]),  # 师资队伍
    ("法学院", "law.seu.edu.cn",
     ["https://law.seu.edu.cn/9121/list.htm"],
     ["http://law.seu.edu.cn/9125/list.htm"]),  # 在职教师
    ("医学院", "med.seu.edu.cn",
     ["https://med.seu.edu.cn/8693/list.htm"],
     ["https://med.seu.edu.cn/8694/list.htm", "https://med.seu.edu.cn/8695/list.htm"]),  # 博导+硕导
    ("吴健雄学院", "wjx.seu.edu.cn",
     ["https://wjx.seu.edu.cn/21376/list.htm"],
     ["https://wjx.seu.edu.cn/xyds/list.htm"]),  # 学院导师
    ("马克思主义学院", "marxism.seu.edu.cn",
     ["https://marxism.seu.edu.cn/23294/list.htm"],
     []),
    ("统计与数据科学学院", "stat.seu.edu.cn",
     ["https://stat.seu.edu.cn/szll_61997/list.htm"],
     []),
]


def is_admin_email(email: str) -> bool:
    low = email.lower()
    for p in ADMIN_PREFIXES:
        if low.startswith(p + "@") or p in low.split("@")[0]:
            return True
    return False


def extract_title(text: str) -> str:
    m = re.search(r"职称[：:]\s*(.+?)(?:\n|$|。|；|研究方向|邮箱|电话|办公室)", text)
    if m:
        raw = m.group(1).strip()
        mapping = {"正高": "教授", "副高": "副教授", "中级": "讲师",
                   "初级": "助教", "正高级": "教授", "副高级": "副教授"}
        return mapping.get(raw, raw)[:20]
    for t in ["长江学者","杰青","优青","院士","教授","研究员","高级工程师",
              "副教授","副研究员","助理教授","助理研究员","讲师","工程师","博导","硕导"]:
        if t in text:
            return t
    return ""


def normalize_name(name: str) -> str:
    return name.replace("​","").replace("‌","").replace("‍","").strip()


def load_progress():
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text(encoding="utf-8"))
    return {"done_depts": {}, "teachers": []}


def save_progress(dept_name: str, teachers: list, data=None):
    if data is None:
        data = load_progress()
    data["done_depts"][dept_name] = len(teachers)
    existing = {t["email"] for t in data["teachers"] if t.get("email")}
    for t in teachers:
        if t.get("email") and t["email"] not in existing:
            existing.add(t["email"])
            data["teachers"].append(t)
    PROGRESS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    sys.stdout.flush()


async def find_subcategory_links(page) -> list[dict]:
    """从页面中查找教师子分类链接（教授/副教授/讲师等）"""
    links = await page.evaluate("""() => {
        const results = [];
        const keywords = ['教授', '副教授', '讲师', '研究员', '副研究员', '助理教授',
            '导师', '教师', '师资', '专任', '兼职', '客座', '退休'];
        document.querySelectorAll('a').forEach(a => {
            const text = a.textContent.trim();
            const href = a.href || '';
            if (!href || href.startsWith('javascript:') || href === '#') return;
            if (href.includes('webplus') || href.includes('_teacherHome') || href.includes('login')) return;
            for (const kw of keywords) {
                if (text.includes(kw) && text.length <= 15) {
                    results.push({text, href});
                    return;
                }
            }
        });
        return results;
    }""")
    # 去重
    seen = set()
    unique = []
    for l in links:
        if l["href"] not in seen:
            seen.add(l["href"])
            unique.append(l)
    return unique


async def extract_teacher_links_from_page(page) -> list[dict]:
    """从页面中提取教师姓名链接"""
    entries = await page.evaluate("""() => {
        const navWords = new Set(['首页','概况','新闻','通知','公告','招生','培养','就业',
            '学位','学科','科研','学术','党建','工会','校友','捐赠','图书馆','校园','地图',
            '网站','登录','邮箱','联系我们','欢迎','返回','更多','详情','查看','下载','学院',
            '大学','管理','后台','English','人才引进','人才招聘','院长书记','信箱','相关链接',
            '联系方式','学校首页','学校主页','收藏本站','旧版入口','暑期学校','平湖芳草',
            '下载专区','捐赠通道','院长邮箱','院内文档','日本語','标识系统','院系设置','教师教学',
            '技术转移','海外教育','仪器设备','化工时刊','尾页','网站首页','招生信息','教师登录',
            '现任领导','历任领导','办公电话','院长寄语','组织机构','学科建设','历史沿革','学院简介',
            '学院介绍','组织框架','系所设置','学科组织','本院概况','本院简介','学院简况','学院概述',
            '学院架构','快捷入口','深入贯彻','江苏省','学术论文','专利成果','获奖成果','课程改革',
            '牵头学科','学位管理','出国交流','答辩公示','本科生','研究生','学生工作','党群工作',
            '人才培养','科学研究','人才引进','校友天地','合作交流','诚聘英才','拔尖基地','教学管理',
            '本科生培','研究生培','鲁汶国际','项目介绍','电子信息','规章制度','学生自我','教职工',
            '教师教学','发展中心','教师查询','教师风采','专任教师','院士','客座教授','教师简介',
            '兼职教授','离退休','荣休','知名专家','知名学者','全体教师','硕博导师','各系名单',
            '国家高层次','人才','管理入口','师资维护','师资修改','个人中心','师资概况','师资力量',
            '师资队伍','师资概览','师资骨干','教师队伍','杰出人才','人才工程','人才职称','教授风采',
            '教授','副教授','讲师','助教','研究员','副研究员','助理教授','助理研究员','高级工程师',
            '工程师','博士后','博士生导师','硕士生导师','博士导师','硕士导师','博导','硕导','在职教师',
            '退休教师','访问学者','实验室','研究中心','学院部门','联系我们','教务系统','中文',
            '关闭此页','学院主页','艺术学院','标识系统','校内办公','网络教学','校园信息','校园卡',
            '电子邮件','图书馆主页','一卡通','财务查询','校园门户','信息服务','网络服务','VPN服务',
            '正版软件','软件下载','校园网','无线网','邮箱系统','信息门户','网上办事','办事大厅',
            '服务大厅','统一身份','认证','一网通办','系所介绍','声影机械','行政服务','资料下载',
            '培养动态','招生动态','常用下载','教师名录','学系介绍','学院治理','行政机构','两院院士',
            '教育学','走近外院','机构设置','党群组织','学术组织','督导组','学术期刊','院务公告',
            '学院领导','学院院标','实验教师','行政人员','退休教工','招聘信息','中大学人','学术研究',
            '科研通知','学术活动','科研项目','科研成果','院长信箱','学院新闻','学术信息','重要通知',
            '教务信息','图片新闻','院训院标','学系','附属医院','Version','EN','EnglishVersion']);

        const results = [];
        const seen = new Set();

        document.querySelectorAll('a').forEach(a => {
            let text = a.textContent.trim().replace(/\\u200b/g, '').replace(/\\u200c/g, '').replace(/\\u200d/g, '').trim();
            const href = a.href || '';
            if (!href || href.startsWith('javascript:') || href === '#') return;
            if (href.startsWith('mailto:')) return;
            if (href.includes('webplus') || href.includes('_teacherHome') || href.includes('login')) return;
            if (href.includes('list.htm') || href.includes('list.psp')) return;

            const cleaned = text.replace(/\\s/g, '');
            if (/^[\\u4e00-\\u9fff]{2,4}$/.test(cleaned) && !navWords.has(cleaned)) {
                if (!seen.has(href)) {
                    seen.add(href);
                    results.push({name: cleaned, url: href});
                }
            }
        });
        return results;
    }""")
    return entries


async def extract_emails_inline(page) -> list[dict]:
    """从列表页直接提取内联的教师信息（未做链接的教师列表）"""
    return await page.evaluate("""() => {
        const body = document.body.innerText;
        const emailRe = /([a-zA-Z0-9._%+\\-]+@[a-zA-Z0-9.\\-]+\\.[a-zA-Z]{2,})/g;
        const results = [];
        const lines = body.split('\\n');

        for (let i = 0; i < lines.length; i++) {
            const line = lines[i];
            const match = emailRe.exec(line);
            if (match) {
                const email = match[1].toLowerCase();
                // Look for Chinese name nearby
                let name = '';
                const nameMatch = line.match(/([\\u4e00-\\u9fff]{2,4})/);
                if (nameMatch) {
                    name = nameMatch[1];
                }
                results.push({name, email, source: 'inline'});
            }
        }
        return results;
    }""")


async def scrape_profile(profile_page, url: str, name: str, dept: str) -> dict | None:
    """访问教师详情页提取邮箱"""
    if url.startswith("mailto:"):
        return None
    try:
        await profile_page.goto(url, wait_until="domcontentloaded", timeout=PROFILE_TIMEOUT)
        await asyncio.sleep(0.8)
        text = await profile_page.evaluate("() => document.body.innerText")
        emails = [e.lower() for e in EMAIL_RE.findall(text) if not is_admin_email(e)]
        if emails:
            return {
                "name": normalize_name(name),
                "email": emails[0],
                "department": dept,
                "title": extract_title(text),
                "url": url,
            }
    except Exception:
        pass
    return None


async def scrape_dept_v2(context, dept_name: str, domain: str,
                          main_urls: list[str], extra_urls: list[str]) -> list[dict]:
    """V2版学院爬取：支持两级导航"""
    print(f"\n{'='*50}")
    print(f"[{dept_name}] 开始爬取...")
    sys.stdout.flush()

    page = await context.new_page()
    profile_page = await context.new_page()
    results = []
    seen_emails = set()
    all_teacher_links = []

    try:
        # === 第一步：收集所有可能的教师列表 URL ===
        sub_urls = list(extra_urls)

        for url in main_urls[:2]:
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                await asyncio.sleep(2)

                # 先尝试查找子分类链接
                subcats = await find_subcategory_links(page)
                print(f"  [{url.rsplit('/',1)[-1]}] 找到 {len(subcats)} 个子分类")

                for sc in subcats:
                    href = sc["href"]
                    if href not in sub_urls and "list.htm" in href:
                        sub_urls.append(href)

                # 同时提取当前页面的教师链接
                teacher_links = await extract_teacher_links_from_page(page)
                if teacher_links:
                    print(f"  直接教师链接: {len(teacher_links)}")
                all_teacher_links.extend(teacher_links)

            except Exception as e:
                print(f"  主列表页错误: {e}")

        # === 第二步：访问所有子分类页面，收集教师链接 ===
        visited_sub = set()
        for sub_url in sub_urls:
            if sub_url in visited_sub:
                continue
            visited_sub.add(sub_url)
            try:
                await page.goto(sub_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                await asyncio.sleep(2)

                # 在这个子页面中找教师链接
                sub_links = await extract_teacher_links_from_page(page)
                print(f"  [sub] {sub_url.rsplit('/',1)[-1]}: {len(sub_links)} 教师链接")
                all_teacher_links.extend(sub_links)

                # 也搜索内联邮箱
                inline = await extract_emails_inline(page)
                for item in inline:
                    if item["email"] not in seen_emails and not is_admin_email(item["email"]):
                        seen_emails.add(item["email"])
                        results.append({
                            "name": normalize_name(item.get("name", "")),
                            "email": item["email"],
                            "department": dept_name,
                            "title": "",
                            "url": sub_url,
                        })
                if inline:
                    print(f"  [sub] 内联邮箱: {len(inline)}")

            except Exception as e:
                print(f"  子页面错误: {e}")

        # === 第三步：去重所有教师链接 ===
        seen_urls = set()
        unique_links = []
        for l in all_teacher_links:
            if l["url"] not in seen_urls:
                seen_urls.add(l["url"])
                unique_links.append(l)

        print(f"  去重后教师链接: {len(unique_links)}")

        # === 第四步：访问教师详情页 ===
        if unique_links:
            count = 0
            for i, entry in enumerate(unique_links[:MAX_TEACHERS_PER_DEPT]):
                result = await scrape_profile(profile_page, entry["url"], entry["name"], dept_name)
                if result and result.get("email") and result["email"] not in seen_emails:
                    seen_emails.add(result["email"])
                    results.append(result)
                    count += 1
                    if count <= 3 or count % 30 == 0:
                        print(f"    [{count}] {result['name']} <{result['email']}> {result['title']}")
                        sys.stdout.flush()
                if (i + 1) % 40 == 0:
                    await asyncio.sleep(0.5)

        print(f"  完成: {len(results)} 位教师(有邮箱)")

    except Exception as e:
        print(f"  学院异常: {e}")
    finally:
        await page.close()
        await profile_page.close()

    sys.stdout.flush()
    return results


async def main():
    progress = load_progress()
    all_teachers = list(progress.get("teachers", []))

    # 先加载原有已完成学院的数据
    old_progress_file = OUTPUT_DIR / "seu_progress.json"
    if old_progress_file.exists():
        old_data = json.loads(old_progress_file.read_text(encoding="utf-8"))
        existing_emails = {t["email"] for t in all_teachers if t.get("email")}
        for t in old_data.get("teachers", []):
            if t.get("email") and t["email"] not in existing_emails:
                existing_emails.add(t["email"])
                all_teachers.append(t)
        print(f"从旧数据加载了 {len(all_teachers)} 位教师")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
        )

        for dept_name, domain, main_urls, extra_urls in PROBLEM_DEPTS:
            if dept_name in progress.get("done_depts", {}):
                prev = progress["done_depts"][dept_name]
                print(f"跳过已完成: {dept_name} ({prev} 位)")
                continue

            try:
                teachers = await scrape_dept_v2(context, dept_name, domain, main_urls, extra_urls)
                if teachers:
                    # 合并到总数据
                    existing_emails = {t["email"] for t in all_teachers if t.get("email")}
                    for t in teachers:
                        if t.get("email") and t["email"] not in existing_emails:
                            existing_emails.add(t["email"])
                            all_teachers.append(t)
                save_progress(dept_name, teachers, progress)
                # Update progress in-memory
                progress = load_progress()
            except Exception as e:
                print(f"学院异常: {dept_name}: {e}")
                save_progress(dept_name, [], progress)
                progress = load_progress()

        await context.close()
        await browser.close()

    # 最终去重
    seen = set()
    unique = []
    for t in all_teachers:
        if t.get("email") and t["email"] not in seen:
            seen.add(t["email"])
            unique.append(t)

    print(f"\n{'='*50}")
    print(f"总计: {len(unique)} 位教师 (去重后)")
    print(f"{'='*50}")

    if unique:
        export_results(unique)
    else:
        print("无新数据")


def export_results(records: list[dict]):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = OUTPUT_DIR / f"东南大学_教师邮箱_{ts}.xlsx"
    csv_path = OUTPUT_DIR / f"东南大学_教师邮箱_{ts}.csv"

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "东南大学教师邮箱"

    headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
    hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    cf = Font(name="微软雅黑", size=10)
    ca = Alignment(vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = thin_border

    for i, r in enumerate(records, 1):
        for col, v in enumerate([i, r.get("name",""), r.get("email",""),
                                  r.get("department",""), r.get("title",""), r.get("url","")], 1):
            cell = ws.cell(row=i+1, column=col, value=v)
            cell.font = cf; cell.alignment = ca; cell.border = thin_border

    for col, w in zip("ABCDEF", [8, 18, 32, 20, 16, 55]):
        ws.column_dimensions[col].width = w

    wb.save(xlsx_path)
    print(f"XLSX: {xlsx_path}")

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for i, r in enumerate(records, 1):
            w.writerow([i, r.get("name",""), r.get("email",""),
                        r.get("department",""), r.get("title",""), r.get("url","")])
    print(f"CSV: {csv_path}")

    dept_counts = Counter(r["department"] for r in records)
    print(f"\n各学院统计:")
    for dept, cnt in dept_counts.most_common():
        print(f"  {dept}: {cnt} 位")

    (OUTPUT_DIR / "seu_final_v2.json").write_text(
        json.dumps({"total": len(records), "xlsx": str(xlsx_path), "csv": str(csv_path),
                    "dept_counts": dict(dept_counts), "teachers": records},
                   ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\n[FILES]")
    print(f"{xlsx_path.name}|东南大学教师邮箱XLSX（V2完整版）")
    print(f"{csv_path.name}|东南大学教师邮箱CSV（V2完整版）")
    print(f"[/FILES]")


if __name__ == "__main__":
    asyncio.run(main())
