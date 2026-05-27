"""南京大学全学院教师信息爬虫 — 完整版。

特性：
- 覆盖约33个院系
- 多策略DOM提取（表格/列表/卡片/图文）
- 自动翻页处理
- 增量保存（断点续传友好）
- 包含无邮箱教师
- 输出XLSX + CSV
"""

import asyncio
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import Counter

OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)
PROGRESS_FILE = OUTPUT_DIR / "nju_full_progress.json"

PAGE_TIMEOUT = 30000
PROFILE_TIMEOUT = 15000
MAX_PROFILES_PER_DEPT = 200  # 每个学院最多访问的详情页数

EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

# 公共邮箱前缀（学院级别，非个人）
ADMIN_PREFIXES = {
    "wxyxz", "xwcb", "bgs", "dangzheng", "office", "yuanban",
    "webmaster", "admin", "info", "master", "root", "postmaster",
    "gcglxydw", "njugcglxy", "njuart", "njulaw", "njumed",
    "njuphysics", "njumath", "njucs", "njusoftware",
}

# ======================== 南京大学院系列表 ========================
# 格式: (学院名, 域名, [可能的师资页面URL列表])
NJU_DEPARTMENTS = [
    # ——— 人文科学部 ———
    ("文学院", "https://chin.nju.edu.cn", [
        "https://chin.nju.edu.cn/szdw.htm",
        "https://chin.nju.edu.cn/szdw/xrjs/zggdwx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/zggdwxx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/zgxddwx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/xjyysx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/bjwxysjwx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/hyywzx/index.html",
        "https://chin.nju.edu.cn/szdw/xrjs/yyxjyyyyx/index.html",
    ]),
    ("历史学院", "https://history.nju.edu.cn", [
        "https://history.nju.edu.cn/szdw/list.htm",
        "https://history.nju.edu.cn/28475/list.htm",
    ]),
    ("哲学系", "https://philo.nju.edu.cn", [
        "https://philo.nju.edu.cn/szdw/list.htm",
        "https://philo.nju.edu.cn/szdw/list2.htm",
    ]),
    ("外国语学院", "https://sfs.nju.edu.cn", [
        "https://sfs.nju.edu.cn/szdw/list.htm",
        "https://sfs.nju.edu.cn/xyjj/list.htm",
    ]),
    ("新闻传播学院", "https://jc.nju.edu.cn", [
        "https://jc.nju.edu.cn/szdw/list.htm",
        "https://jc.nju.edu.cn/3658/list.htm",
    ]),

    # ——— 社会科学部 ———
    ("商学院", "https://nubs.nju.edu.cn", [
        "https://nubs.nju.edu.cn/szdw/list.htm",
        "https://nubs.nju.edu.cn/1737/list.htm",
        "https://nubs.nju.edu.cn/1738/list.htm",
    ]),
    ("法学院", "https://law.nju.edu.cn", [
        "https://law.nju.edu.cn/szdw/list.htm",
        "https://law.nju.edu.cn/1920/list.htm",
    ]),
    ("政府管理学院", "https://public.nju.edu.cn", [
        "https://public.nju.edu.cn/szdw/list.htm",
        "https://public.nju.edu.cn/xygk/list.htm",
    ]),
    ("信息管理学院", "https://im.nju.edu.cn", [
        "https://im.nju.edu.cn/szdw/list.htm",
        "https://im.nju.edu.cn/10595/list.htm",
    ]),
    ("社会学院", "https://sociology.nju.edu.cn", [
        "https://sociology.nju.edu.cn/szdw/list.htm",
        "https://sociology.nju.edu.cn/16269/list.htm",
    ]),
    ("国际关系学院", "https://sir.nju.edu.cn", [
        "https://sir.nju.edu.cn/szdw/list.htm",
    ]),
    ("马克思主义学院", "https://marxism.nju.edu.cn", [
        "https://marxism.nju.edu.cn/szdw/list.htm",
    ]),

    # ——— 理学部 ———
    ("数学学院", "https://math.nju.edu.cn", [
        "https://math.nju.edu.cn/szdw/list.htm",
        "https://math.nju.edu.cn/23673/list.htm",
    ]),
    ("物理学院", "https://physics.nju.edu.cn", [
        "https://physics.nju.edu.cn/szdw/list.htm",
        "https://physics.nju.edu.cn/zzjs/list.htm",
        "https://physics.nju.edu.cn/fjs/list.htm",
        "https://physics.nju.edu.cn/js/list.htm",
    ]),
    ("化学化工学院", "https://chem.nju.edu.cn", [
        "https://chem.nju.edu.cn/szll/list.htm",
        "https://chem.nju.edu.cn/szll/list2.htm",
        "https://chem.nju.edu.cn/szll/list3.htm",
    ]),
    ("天文与空间科学学院", "https://astronomy.nju.edu.cn", [
        "https://astronomy.nju.edu.cn/szdw/list.htm",
        "https://astronomy.nju.edu.cn/26193/list.htm",
    ]),
    ("地球科学与工程学院", "https://es.nju.edu.cn", [
        "https://es.nju.edu.cn/szdw/list.htm",
        "https://es.nju.edu.cn/szdw/list2.htm",
    ]),
    ("地理与海洋科学学院", "https://sgos.nju.edu.cn", [
        "https://sgos.nju.edu.cn/szdw/list.htm",
        "https://sgos.nju.edu.cn/26389/list.htm",
    ]),
    ("大气科学学院", "https://atmos.nju.edu.cn", [
        "https://atmos.nju.edu.cn/szdw/list.htm",
        "https://atmos.nju.edu.cn/26458/list.htm",
    ]),
    ("生命科学学院", "https://life.nju.edu.cn", [
        "https://life.nju.edu.cn/szdw/list.htm",
        "https://life.nju.edu.cn/26622/list.htm",
    ]),

    # ——— 工学部 ———
    ("电子科学与工程学院", "https://ese.nju.edu.cn", [
        "https://ese.nju.edu.cn/szdw/list.htm",
    ]),
    ("计算机科学与技术系", "https://cs.nju.edu.cn", [
        "https://cs.nju.edu.cn/szdw/list.htm",
        "https://cs.nju.edu.cn/22842/list.htm",
    ]),
    ("软件学院", "https://software.nju.edu.cn", [
        "https://software.nju.edu.cn/szdw/list.htm",
        "https://software.nju.edu.cn/23470/list.htm",
    ]),
    ("人工智能学院", "https://ai.nju.edu.cn", [
        "https://ai.nju.edu.cn/szdw/list.htm",
        "https://ai.nju.edu.cn/21872/list.htm",
    ]),
    ("现代工程与应用科学学院", "https://eng.nju.edu.cn", [
        "https://eng.nju.edu.cn/szdw/list.htm",
        "https://eng.nju.edu.cn/24680/list.htm",
    ]),
    ("建筑与城市规划学院", "https://arch.nju.edu.cn", [
        "https://arch.nju.edu.cn/szdw/list.htm",
        "https://arch.nju.edu.cn/16373/list.htm",
    ]),
    ("环境学院", "https://environ.nju.edu.cn", [
        "https://environ.nju.edu.cn/szdw/list.htm",
        "https://environ.nju.edu.cn/24938/list.htm",
    ]),
    ("工程管理学院", "https://sme.nju.edu.cn", [
        "https://sme.nju.edu.cn/gygcyyyglx/list.htm",
        "https://sme.nju.edu.cn/fzgcglx/list.htm",
        "https://sme.nju.edu.cn/jrkjygcx/list.htm",
        "https://sme.nju.edu.cn/2031/list.htm",
    ]),
    ("能源与资源学院", "https://energy.nju.edu.cn", [
        "https://energy.nju.edu.cn/ktzcy/js/index.html",
        "https://energy.nju.edu.cn/ktzcy/zl/index.html",
        "https://energy.nju.edu.cn/ktzcy/bsh/index.html",
    ]),
    ("匡亚明学院", "https://dii.nju.edu.cn", [
        "https://dii.nju.edu.cn/lsjs/list.htm",
    ]),

    # ——— 医学与生命科学 ———
    ("医学院", "https://med.nju.edu.cn", [
        "https://med.nju.edu.cn/szdw/list.htm",
        "https://med.nju.edu.cn/26018/list.htm",
    ]),

    # ——— 其他 ———
    ("教育研究院", "https://edu.nju.edu.cn", [
        "https://edu.nju.edu.cn/8746/list.htm",
        "https://edu.nju.edu.cn/szdw/list.htm",
    ]),
    ("艺术学院", "https://art.nju.edu.cn", [
        "https://art.nju.edu.cn/ysllycyx/list.htm",
        "https://art.nju.edu.cn/msysjx/list.htm",
        "https://art.nju.edu.cn/whysjyzx/list.htm",
    ]),
    ("中美文化研究中心", "https://hnc.nju.edu.cn", [
        "https://hnc.nju.edu.cn/szll/list.htm",
    ]),
    ("体育部", "https://tyb.nju.edu.cn", [
        "https://tyb.nju.edu.cn/szdw/list.htm",
        "https://tyb.nju.edu.cn/12911/list.htm",
    ]),

    # ——— 交叉学科 ———
    ("国际地球系统科学研究所", "https://giess.nju.edu.cn", [
        "https://giess.nju.edu.cn/szdw/list.htm",
    ]),
]


def ts():
    return datetime.now().strftime("%H:%M:%S")


def is_admin_email(email: str) -> bool:
    email_lower = email.lower()
    local = email_lower.split("@")[0]
    if len(local) <= 2 and local.isalpha():
        return True
    for prefix in ADMIN_PREFIXES:
        if email_lower.startswith(prefix + "@"):
            return True
    return False


def recover_email(text: str) -> str:
    patterns = [
        (r"\s*\[at\]\s*", "@"),
        (r"\s*\(at\)\s*", "@"),
        (r"\s*#@\s*", "@"),
        (r"\s*\[@\]\s*", "@"),
    ]
    for pat, rep in patterns:
        text = re.sub(pat, rep, text, flags=re.IGNORECASE)
    return text


def extract_title(text: str) -> str:
    """从文本中提取职称。"""
    keywords = [
        "教授", "副教授", "助理教授", "讲师", "研究员", "副研究员",
        "助理研究员", "工程师", "高级工程师", "院士", "博导", "硕导",
        "长江学者", "杰青", "优青", "博士后", "特聘教授", "讲座教授",
        "长聘教授", "准聘教授", "准聘副教授",
    ]
    found = []
    for kw in keywords:
        if kw in text:
            found.append(kw)
    return "、".join(found)


async def goto_safe(page, url: str, timeout=PAGE_TIMEOUT) -> bool:
    try:
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        if resp and resp.status >= 500:
            return False
        await asyncio.sleep(1.5)
        return True
    except Exception:
        return False


async def find_faculty_pages(page, domain: str) -> list[dict]:
    """在学院主页查找师资相关子页面。"""
    links = await page.evaluate("""(domain) => {
        const keywords = ['师资', '教师', 'faculty', 'staff', '人员', '教授', '博导',
            '硕导', '导师', '教职工', '名录', '教师名录', '教师主页', '导师介绍',
            '教师名单', '在职教师', '专任教师', '师资力量', '师资概况',
            'szdw', 'szll', 'teacher', 'teachers', 'faculty', 'js', 'jzg',
        ];
        const results = [];
        const seen = new Set();
        document.querySelectorAll('a[href]').forEach(a => {
            const text = (a.textContent || '').trim().toLowerCase();
            const href = (a.href || '').toLowerCase();
            if (!href || href.startsWith('javascript:') || href.startsWith('#')) return;
            if (href.endsWith('.pdf') || href.endsWith('.doc')) return;
            if (seen.has(href)) return;
            for (const kw of keywords) {
                if (text.includes(kw) || href.includes(kw)) {
                    seen.add(href);
                    results.push({text: a.textContent.trim(), href: a.href});
                    break;
                }
            }
        });
        return results.slice(0, 20);
    }""", domain)
    return links


async def extract_teacher_links_from_page(page, domain: str) -> list[dict]:
    """从当前页面提取教师姓名链接。返回 [{name, url}]。"""
    entries = await page.evaluate("""(domain) => {
        const navSet = new Set([
            '学校主页','学院首页','网站首页','返回首页','设为首页','加入收藏',
            '联系我们','后台管理','登录系统','学院概况','历史沿革','组织机构',
            '院长致辞','现任领导','院系领导','党群组织','行政部门','规章制度',
            '办事指南','诚聘英才','人才招聘','校友名录','校友风采','校友活动',
            '百年院庆','校友之家','人才培养','本科生','研究生','博士后','留学生',
            '暑期班','培训班','海外班','专业设置','学位授予','课程设置','教学大纲',
            '招生简章','招生培养','科学研究','学术研究','科研项目','科研奖励',
            '院办刊物','学术会议','学术机构','研究方向','成果速递','成果展示',
            '学生工作','党团建设','学工布告','学生活动','学子风采','学科建设',
            '合作交流','国际合作','社会服务','新闻速递','通知公告','学术动态',
            '人才队伍','党建思政','团学工作','校友工作','信息公开','下载中心',
            '资料下载','教学资源','相关下载','新闻中心','学院新闻','学术活动',
            '报告讲座','教务通知','学工通知','公示公告','招聘信息','访问学者',
            '讲座教授','友情链接','网站地图','版权信息','English','中文','英文',
            '首页','返回','更多','查看详情','职称','学历','研究方向','主讲课程',
            '教师名录','现任教师','退休教师','荣休教师','教师主页','导师介绍',
            '导师简介','师资队伍','院士','教职工','专任教师','双一流','211','985',
            '无机化学','分析化学','有机化学','物理化学','高分子化学',
            '生物化学','环境化学','化学工程','材料化学','电子工程','通信工程',
            '光电工程','微电子','环境工程','给排水','环境科学','地理科学',
            '遥感科学','地图学','GIS','大气科学','气象学','气候学','临床医学',
            '基础医学','预防医学','口腔医学','人工智能','机器学习','计算机科学',
            '软件工程','建筑学','城市规划','风景园林',
        ]);

        const results = [];
        const seen = new Set();

        // 策略1: 表格行中的教师链接（最常见）
        document.querySelectorAll('table tr, .wp_list, .list_item, .teacher_item').forEach(row => {
            const links = row.querySelectorAll('a[href]');
            const cells = row.querySelectorAll('td, th, .wp_col');
            if (links.length >= 1) {
                const firstLink = links[0];
                const text = (firstLink.textContent || '').trim();
                const href = firstLink.href || '';
                if (!href || href.startsWith('javascript:') || href.startsWith('#')) return;
                if (href.endsWith('.pdf') || href.endsWith('.doc')) return;
                if (href.includes('beian.miit.gov.cn')) return;
                if (seen.has(href)) return;
                if (!/^[一-鿿]{2,4}$/.test(text)) return;
                if (navSet.has(text)) return;
                seen.add(href);

                let title = '';
                if (cells.length >= 2) {
                    for (let i = 1; i < Math.min(cells.length, 6); i++) {
                        const ct = (cells[i].textContent || '').trim();
                        const tm = ct.match(/(教授|副教授|助理教授|讲师|研究员|副研究员|助理研究员|工程师|高级工程师|院士)/);
                        if (tm) { title = tm[1]; break; }
                    }
                }
                results.push({name: text, url: href, title});
            }
        });

        // 策略2: 列表/卡片中的链接
        if (results.length < 5) {
            document.querySelectorAll('a[href]').forEach(a => {
                const text = (a.textContent || '').trim();
                const href = a.href || '';
                if (!href || seen.has(href)) return;
                if (href.startsWith('javascript:') || href.startsWith('#')) return;
                if (href.endsWith('.pdf') || href.endsWith('.doc')) return;
                if (href.includes('beian.miit.gov.cn')) return;
                if (!/^[一-鿿]{2,4}$/.test(text)) return;
                if (navSet.has(text)) return;
                // 检查不在导航区域
                const parent = a.closest('nav, .nav, .header, .footer, .navi, .menu, .sidebar, .top_nav, .bottom_nav');
                if (parent) return;
                seen.add(href);
                results.push({name: text, url: href, title: ''});
            });
        }

        return results.slice(0, 200);
    }""", domain)
    return entries


async def find_next_page(page) -> str | None:
    """查找下一页URL。"""
    next_url = await page.evaluate("""() => {
        const links = document.querySelectorAll('a');
        for (const a of links) {
            const text = a.textContent.trim();
            if (text === '下一页' || text === '>' || text === '>>' ||
                text === 'next' || text === '下页') {
                return a.href;
            }
        }
        // 查找分页中的"当前页+1"数字
        let currentPage = null;
        for (const a of links) {
            const text = a.textContent.trim();
            if (a.classList.contains('current') || a.classList.contains('active') ||
                a.parentElement && (a.parentElement.classList.contains('current') || a.parentElement.classList.contains('active'))) {
                const m = text.match(/^\\d+$/);
                if (m) currentPage = parseInt(m[0]);
            }
        }
        if (currentPage !== null) {
            for (const a of links) {
                const text = a.textContent.trim();
                const m = text.match(/^\\d+$/);
                if (m && parseInt(m[0]) === currentPage + 1) {
                    return a.href;
                }
            }
        }
        return null;
    }""")
    return next_url


async def scrape_profile(context, entry: dict, dept_name: str) -> dict | None:
    """访问教师详情页，提取邮箱和职称。"""
    name = entry.get("name", "")
    url = entry.get("url", "")
    title_hint = entry.get("title", "")

    try:
        profile_page = await context.new_page()
        try:
            await profile_page.goto(url, wait_until="domcontentloaded", timeout=PROFILE_TIMEOUT)
            await asyncio.sleep(0.6)

            body_text = await profile_page.evaluate("() => document.body?.innerText || ''")
            body_text = recover_email(body_text)

            # 提取邮箱
            all_emails = set()
            for e in EMAIL_RE.findall(body_text):
                all_emails.add(e.lower())

            # mailto链接
            mailto_emails = await profile_page.evaluate("""() => {
                const emails = [];
                document.querySelectorAll('a[href^="mailto:"]').forEach(a => {
                    const e = a.getAttribute('href').replace('mailto:', '').split('?')[0].trim();
                    if (e) emails.push(e);
                });
                return emails;
            }""")
            for e in mailto_emails:
                all_emails.add(e.lower())

            valid_emails = [e for e in all_emails if EMAIL_RE.match(e) and not is_admin_email(e)]
            email = valid_emails[0] if valid_emails else ""

            # 提取职称
            title = title_hint
            if not title:
                title = extract_title(body_text[:3000])

            # 尝试从meta标签或特定位置提取姓名确认
            name_confirm = ""
            name_patterns = [
                r'姓名[：:]\s*([一-鿿]{2,4})',
                r'教师姓名[：:]\s*([一-鿿]{2,4})',
            ]
            for pat in name_patterns:
                m = re.search(pat, body_text[:2000])
                if m:
                    name_confirm = m.group(1)
                    break

            final_name = name_confirm if name_confirm else name

            return {
                "name": final_name,
                "email": email,
                "department": dept_name,
                "title": title,
                "url": url,
            }
        finally:
            await profile_page.close()
    except Exception:
        # 即使失败也返回基本信息
        return {
            "name": name,
            "email": "",
            "department": dept_name,
            "title": title_hint,
            "url": url,
        }


async def scrape_department(context, dept_name: str, domain: str, list_urls: list[str]) -> list[dict]:
    """爬取单个学院的全部教师。"""
    print(f"\n{'='*55}")
    print(f"[{ts()}] 🏫 {dept_name}")

    page = await context.new_page()
    all_results = []
    all_teacher_links = []
    seen_urls = set()

    try:
        # --- Phase 1: 收集所有教师链接 ---
        # 方法A: 使用预配置的URL列表
        for list_url in list_urls:
            if not await goto_safe(page, list_url):
                continue
            print(f"  📄 列表页: {list_url[:80]}")

            # 处理分页
            page_num = 1
            while page_num <= 20:
                if page_num > 1:
                    await asyncio.sleep(1.5)

                entries = await extract_teacher_links_from_page(page, domain)
                new_count = 0
                for e in entries:
                    if e["url"] not in seen_urls:
                        seen_urls.add(e["url"])
                        all_teacher_links.append(e)
                        new_count += 1
                print(f"    第{page_num}页: +{new_count} 个新教师链接")

                if new_count < 3 and page_num > 1:
                    break

                # 找下一页
                next_url = await find_next_page(page)
                if next_url and next_url != page.url:
                    if not await goto_safe(page, next_url):
                        break
                    page_num += 1
                else:
                    break

        # 方法B: 如果预配置URL不够，在首页搜索师资入口
        if len(all_teacher_links) < 5:
            print(f"  🔍 预配置URL获得 {len(all_teacher_links)} 个链接，在首页搜索...")
            if await goto_safe(page, domain):
                faculty_links = await find_faculty_pages(page, domain)
                for fl in faculty_links[:10]:
                    if fl["href"] in seen_urls:
                        continue
                    if not await goto_safe(page, fl["href"]):
                        continue
                    entries = await extract_teacher_links_from_page(page, domain)
                    new_count = 0
                    for e in entries:
                        if e["url"] not in seen_urls:
                            seen_urls.add(e["url"])
                            all_teacher_links.append(e)
                            new_count += 1
                    print(f"    {fl['text'][:20]}: +{new_count}")
                    if len(all_teacher_links) >= 10:
                        break

        print(f"  🔗 总计: {len(all_teacher_links)} 个教师链接")

        if len(all_teacher_links) > MAX_PROFILES_PER_DEPT:
            print(f"  ⚠️ 超过 {MAX_PROFILES_PER_DEPT} 个，截断")
            all_teacher_links = all_teacher_links[:MAX_PROFILES_PER_DEPT]

        # --- Phase 2: 访问详情页 ---
        for i, entry in enumerate(all_teacher_links):
            result = await scrape_profile(context, entry, dept_name)
            if result:
                all_results.append(result)
                status = "✅" if result["email"] else "❌"
                if (i + 1) % 30 == 0:
                    has_email = sum(1 for r in all_results if r["email"])
                    print(f"    进度: {i+1}/{len(all_teacher_links)}, 已提取邮箱 {has_email} 个")

    finally:
        await page.close()

    # 去重
    seen_keys = set()
    unique = []
    for r in all_results:
        key = (r["name"], r["url"])
        if key not in seen_keys:
            seen_keys.add(key)
            unique.append(r)

    with_email = sum(1 for r in unique if r["email"])
    print(f"  📊 {dept_name}: {len(unique)}人, {with_email}有邮箱")
    return unique


async def main():
    print(f"[{ts()}] 🚀 南京大学全学院教师信息爬虫")
    print(f"[{ts()}] 目标: {len(NJU_DEPARTMENTS)} 个院系")
    print()

    # 加载进度
    progress = {"completed": [], "results": []}
    if PROGRESS_FILE.exists():
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                progress = json.load(f)
            print(f"[{ts()}] 恢复进度: {len(progress['completed'])} 院系已完成, {len(progress['results'])} 条记录")
        except Exception:
            pass

    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )

        try:
            for idx, (dept_name, domain, list_urls) in enumerate(NJU_DEPARTMENTS):
                if dept_name in progress["completed"]:
                    print(f"\n[{ts()}] ⏭️ 跳过 {dept_name} (已完成)")
                    continue

                print(f"\n[{ts()}] [{idx+1}/{len(NJU_DEPARTMENTS)}]")

                try:
                    results = await scrape_department(context, dept_name, domain, list_urls)
                except Exception as e:
                    print(f"  ❌ 异常: {e}")
                    import traceback
                    traceback.print_exc()
                    results = []

                progress["completed"].append(dept_name)
                progress["results"].extend(results)

                # 增量保存
                with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
                    json.dump(progress, f, ensure_ascii=False)
                print(f"  💾 进度: {len(progress['completed'])}/{len(NJU_DEPARTMENTS)} ({len(progress['results'])} 条)")

        finally:
            await context.close()
            await browser.close()

    # ===================== 最终处理 =====================
    ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    all_results = progress["results"]

    # 去重（按 name+url）
    seen = set()
    final = []
    for r in all_results:
        key = (r["name"], r["url"])
        if key not in seen:
            seen.add(key)
            final.append(r)

    # 清洗：移除邮箱和职称都为空的非教师条目
    before = len(final)
    final = [r for r in final if r.get("name") and len(r["name"]) >= 2]
    print(f"\n[{ts()}] 清洗: {before} → {len(final)} (移除空姓名等)")

    # 统计
    print(f"\n{'='*55}")
    print(f"[{ts()}] 🎉 爬取完成!")
    print(f"[{ts()}] 院系数: {len(progress['completed'])}")
    print(f"[{ts()}] 教师总数: {len(final)}")
    with_email = sum(1 for r in final if r["email"])
    print(f"[{ts()}] 有邮箱: {with_email}/{len(final)}")

    # 按学院统计
    dept_counts = Counter(r["department"] for r in final)
    print(f"\n=== 各学院教师数量 ===")
    for dept, count in dept_counts.most_common():
        has_email = sum(1 for r in final if r["department"] == dept and r["email"])
        print(f"  {dept}: {count}人 ({has_email}有邮箱)")

    # 保存CSV
    csv_path = OUTPUT_DIR / f"南京大学_教师邮箱_完整版_{ts_str}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "姓名", "邮箱", "学院", "职称", "主页链接"])
        for i, r in enumerate(final, 1):
            writer.writerow([
                i, r["name"], r["email"], r["department"], r["title"], r["url"],
            ])
    print(f"\n💾 CSV: {csv_path}")

    # 保存XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "教师邮箱"

        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )
        cell_font = Font(name="微软雅黑", size=10)
        cell_align = Alignment(vertical="center")
        link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

        headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for i, r in enumerate(final, 1):
            values = [i, r["name"], r["email"], r["department"], r["title"], r["url"]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.font = link_font if col == 6 and val else cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 55

        xlsx_path = OUTPUT_DIR / f"南京大学_教师邮箱_完整版_{ts_str}.xlsx"
        wb.save(xlsx_path)
        print(f"💾 XLSX: {xlsx_path}")
    except Exception as e:
        print(f"⚠️ XLSX导出失败: {e}")

    print(f"\n[{ts()}] ✅ 全部完成!")
    return csv_path, final


if __name__ == "__main__":
    asyncio.run(main())
