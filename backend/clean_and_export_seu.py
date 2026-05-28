"""数据清理 + 最终导出 — 东南大学教师邮箱"""
import json
import csv
import re
from datetime import datetime
from pathlib import Path
from collections import Counter

OUTPUT_DIR = Path(r"D:\Work\test\UniEmailAgent\backend\outputs\bda95480-ec96-4bd7-bc18-05f797e28dd4")

# 加载V2数据
v2 = json.loads((OUTPUT_DIR / "seu_final_v2.json").read_text(encoding="utf-8"))
teachers = v2["teachers"]
print(f"加载: {len(teachers)} 条记录")

# 清理：过滤name为导航词/非人名的条目
BAD_NAMES = {
    '建筑学院','博士后','诚聘英才','本科招生','研究生院','信息门户','教师名录',
    '师资队伍','师资力量','师资概况','学院概况','机构设置','联系我们','下载专区',
    '学院新闻','通知公告','学术活动','人才培养','科学研究','党群工作','学生工作',
    '校友工作','学院简介','历史沿革','现任领导','历任领导','组织框架','院长信箱',
    '首页','概况','新闻','通知','公告','招生','培养','就业','英文版','English',
    'Version','返回','更多','详情','查看','下载','登录','管理入口','师资维护',
    '师资修改','个人中心','中文',
    # V2 内联提取的非人名
    '东南大学','南京', '江苏省','联系方式','邮政编码','通信地址','办公地点',
}

# 清理admin邮箱
ADMIN_PATTERNS = [
    'deanoffice', 'yuanban', 'bgs@', 'office@', 'admin', 'webmaster',
    'info@seu', 'master@', 'root@', 'postmaster@', 'xwcb@', 'wxyxz@',
    'dzxy@pub', 'ysxy@pub',
]

def is_bad_name(name):
    if not name or len(name) < 2:
        return True
    if name in BAD_NAMES:
        return True
    # 包含特殊字符或纯数字
    if re.match(r'^[\d\s\-_\.,;:()（）]+$', name):
        return True
    # 必须是2-4个汉字或包含英文名
    if not re.match(r'^[一-鿿]{2,4}$', name) and not re.match(r'^[A-Z][a-z]+', name):
        # 宽松模式：至少包含一个汉字
        if not re.search(r'[一-鿿]', name):
            return True
    return False

def is_admin_email(email):
    low = email.lower()
    for p in ADMIN_PATTERNS:
        if p in low:
            return True
    return False

# 过滤
cleaned = []
for t in teachers:
    name = t.get("name", "").strip()
    email = t.get("email", "").strip().lower()

    if is_bad_name(name):
        continue
    if is_admin_email(email):
        continue
    if not email or "@" not in email:
        continue

    cleaned.append(t)

print(f"清理后: {len(cleaned)} 条记录 (移除 {len(teachers) - len(cleaned)} 条)")

# 去重(按email)
seen = set()
unique = []
for t in cleaned:
    e = t["email"].lower()
    if e not in seen:
        seen.add(e)
        unique.append(t)

print(f"去重后: {len(unique)} 条记录")

# ==== 导出 XLSX ====
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
xlsx_path = OUTPUT_DIR / f"东南大学_教师邮箱_完整版_{ts}.xlsx"
csv_path = OUTPUT_DIR / f"东南大学_教师邮箱_完整版_{ts}.csv"

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "东南大学教师邮箱"

headers = ["序号", "姓名", "邮箱", "学院", "职称", "主页链接"]
hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="10A37F", end_color="10A37F", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
cf = Font(name="微软雅黑", size=10)
ca = Alignment(vertical="center")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hf; cell.fill = hfill; cell.alignment = halign; cell.border = thin_border

# 按学院排序
unique.sort(key=lambda r: (r.get("department",""), r.get("name","")))

for i, r in enumerate(unique, 1):
    for col, v in enumerate([i, r.get("name",""), r.get("email",""),
                              r.get("department",""), r.get("title",""), r.get("url","")], 1):
        cell = ws.cell(row=i+1, column=col, value=v)
        cell.font = cf; cell.alignment = ca; cell.border = thin_border

for col, w in zip("ABCDEF", [8, 18, 32, 20, 16, 55]):
    ws.column_dimensions[col].width = w

# 冻结首行
ws.freeze_panes = "A2"

wb.save(xlsx_path)
print(f"XLSX: {xlsx_path}")

# ==== 导出 CSV ====
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(headers)
    for i, r in enumerate(unique, 1):
        w.writerow([i, r.get("name",""), r.get("email",""),
                    r.get("department",""), r.get("title",""), r.get("url","")])
print(f"CSV: {csv_path}")

# ==== 学院统计 ====
dept_counts = Counter(r["department"] for r in unique)
print(f"\n各学院统计 ({len(unique)} 位教师):")
for dept, cnt in dept_counts.most_common():
    print(f"  {dept}: {cnt} 位")

# 保存最终JSON
final_json = OUTPUT_DIR / "seu_final_clean.json"
final_json.write_text(json.dumps({
    "total": len(unique),
    "xlsx": str(xlsx_path),
    "csv": str(csv_path),
    "dept_counts": dict(dept_counts),
    "teachers": unique
}, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"\n[FILES]")
print(f"{xlsx_path.name}|东南大学教师邮箱XLSX（清理版）")
print(f"{csv_path.name}|东南大学教师邮箱CSV（清理版）")
print(f"[/FILES]")
